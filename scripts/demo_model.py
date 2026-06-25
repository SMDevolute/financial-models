"""Reference model showing the house style end-to-end. Run: python3 scripts/demo_model.py"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import house_style as hs
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "models", "demo_styled_model.xlsx")
YEARS = ["FY24A", "FY25E", "FY26E", "FY27E", "FY28E"]
C0 = 3                       # first number column (C)
cols = [get_column_letter(C0 + i) for i in range(len(YEARS))]

wb = hs.workbook()
ws = hs.sheet(wb, "Operating Model")

hs.title_block(ws, "Acme SaaS Inc.", "Illustrative operating model · $000s · fiscal year ending Dec")
hs.legend(ws, 5)

r = 7
hs.section(ws, r, "Revenue build", span=len(YEARS) + 1); r += 2
hs.col_headers(ws, r, YEARS, start_col=C0, label="$000s"); r += 1

row_custs = r
hs.label(ws, r, "Customers (avg)")
hs.row_inputs(ws, r, [1200, 1620, 2100, 2650, 3250], start_col=C0, fmt="num"); r += 1

row_arpa = r
hs.label(ws, r, "ARPA ($/yr)")
hs.row_inputs(ws, r, [4800, 5100, 5350, 5550, 5700], start_col=C0, fmt="num"); r += 1

row_rev = r
hs.label(ws, r, "Revenue", bold=True)
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"={col}{row_custs}*{col}{row_arpa}/1000", fmt="num", kind="formula", bold=True, top=True)
r += 1
row_growth = r
hs.label(ws, r, "  growth %", muted=True, italic=True)
for i, col in enumerate(cols):
    val = "" if i == 0 else f"={col}{row_rev}/{cols[i-1]}{row_rev}-1"
    hs.cell(ws, r, C0 + i, val, fmt="pct", kind="formula")
r += 2

hs.section(ws, r, "Profit & loss", span=len(YEARS) + 1); r += 2
hs.col_headers(ws, r, YEARS, start_col=C0, label="$000s"); r += 1

row_rev2 = r
hs.label(ws, r, "Revenue")
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"={col}{row_rev}", fmt="num", kind="link")
r += 1
row_cogs = r
hs.label(ws, r, "Cost of revenue")
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"=-{col}{row_rev2}*0.22", fmt="num", kind="formula")
r += 1
row_gp = r
hs.label(ws, r, "Gross profit", bold=True)
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"={col}{row_rev2}+{col}{row_cogs}", fmt="num", kind="formula", bold=True, top=True)
r += 1
row_gm = r
hs.label(ws, r, "  gross margin %", muted=True, italic=True)
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"={col}{row_gp}/{col}{row_rev2}", fmt="pct", kind="formula")
r += 1
row_opex = r
hs.label(ws, r, "Operating expenses")
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"=-{col}{row_rev2}*0.55", fmt="num", kind="formula")
r += 1
row_ebitda = r
hs.label(ws, r, "EBITDA", bold=True)
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"={col}{row_gp}+{col}{row_opex}", fmt="num", kind="formula", bold=True, double_top=True)
r += 1
hs.label(ws, r, "  EBITDA margin %", muted=True, italic=True)
for i, col in enumerate(cols):
    hs.cell(ws, r, C0 + i, f"={col}{row_ebitda}/{col}{row_rev2}", fmt="pct", kind="formula")
r += 2

# a self-documenting check
hs.check(ws, r, 2, f"=IF(ABS(C{row_rev}-C{row_rev2})<0.5,\"OK — revenue ties\",\"CHECK\")")

hs.band_rows(ws, [row_arpa, row_growth, row_gm], start_col=2, end_col=C0 + len(YEARS) - 1)
hs.set_widths(ws, {1: 2.5, 2: 26, **{C0 + i: 12 for i in range(len(YEARS))}})
hs.finish(ws, freeze=f"C{8}")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("wrote", OUT)
