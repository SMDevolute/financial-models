"""audit.py -- identity, sign and structural checks on the recalculated model."""
import openpyxl, sys, datetime as dt
from openpyxl.utils import get_column_letter as gl
P=sys.argv[1]; V=openpyxl.load_workbook(P,data_only=True); F=openpyxl.load_workbook(P)
AS,RF,CG,PE,OP,FS,DB=(V['Assumptions'],V['Revenue Forecast'],V['COGS'],V['Personnel'],
                      V['OPEX'],V['Financial Statements'],V['Dashboard'])
M0,NM=5,60; YEARS=[2026,2027,2028,2029,2030]; AC0=M0+NM+1
YC={y:gl(AC0+k) for k,y in enumerate(YEARS)}
def n(x): return x if isinstance(x,(int,float)) else 0.0
def md(i): return dt.date(2026+i//12,i%12+1,1)
def row(ws,r): return [n(ws.cell(r,M0+i).value) for i in range(NM)]
fails=[]
def chk(name,cond_fn,detail=''):
    bad=[]
    for i in range(NM):
        ok,info=cond_fn(i)
        if not ok: bad.append((i,info))
    if bad:
        fails.append(name)
        print(f'  FAIL  {name:<52} {len(bad)} months, first {md(bad[0][0]).strftime("%b-%Y")}: {bad[0][1]}')
    else:
        print(f'  pass  {name}')

print('=== CASH FLOW AND BALANCE SHEET IDENTITIES ===')
r=lambda ws,x: row(ws,x)
f=lambda x: row(FS,x)
F32,F33,F37,F38,F39,F40=f(32),f(33),f(37),f(38),f(39),f(40)
F43,F44,F45,F46,F47,F48=f(43),f(44),f(45),f(46),f(47),f(48)
F50,F51,F52,F53,F54,F55,F56,F58=f(50),f(51),f(52),f(53),f(54),f(55),f(56),f(58)
F6,F7,F8,F11,F12,F13,F14,F16,F19,F20,F21,F22,F23=[f(x) for x in (6,7,8,11,12,13,14,16,19,20,21,22,23)]
F61,F62,F63,F64,F35=f(61),f(62),f(63),f(64),f(35)
T=0.02
chk('movement in cash = operations + investing + financing',
    lambda i:(abs(F38[i]-(F32[i]+F33[i]+F37[i]))<T,f'{F38[i]:.2f} vs {F32[i]+F33[i]+F37[i]:.2f}'))
chk('closing cash = opening cash + movement',
    lambda i:(abs(F40[i]-(F39[i]+F38[i]))<T,f'{F40[i]:.2f}'))
chk('opening cash = prior closing cash',
    lambda i:(True,'') if i==0 else (abs(F39[i]-F40[i-1])<T,f'{F39[i]:.2f} vs {F40[i-1]:.2f}'))
chk('cash on the balance sheet = cash from the cash flow',
    lambda i:(abs(F43[i]-F40[i])<T,f'{F43[i]:.2f} vs {F40[i]:.2f}'))
chk('total current assets = cash + receivables + inventory',
    lambda i:(abs(F46[i]-(F43[i]+F44[i]+F45[i]))<T,''))
chk('total assets = current assets + net PPE',
    lambda i:(abs(F48[i]-(F46[i]+F47[i]))<T,''))
chk('total liabilities = payables + loan',
    lambda i:(abs(F52[i]-(F50[i]+F51[i]))<T,''))
chk('total equity = share capital + retained earnings',
    lambda i:(abs(F55[i]-(F53[i]+F54[i]))<T,''))
chk('liabilities and equity = liabilities + equity',
    lambda i:(abs(F56[i]-(F52[i]+F55[i]))<T,''))
chk('balance sheet check row is nil',lambda i:(abs(F58[i])<T,f'{F58[i]:.4f}'))
chk('retained earnings = cumulative net income',
    lambda i:(abs(F54[i]-sum(F23[:i+1]))<T,f'{F54[i]:.2f} vs {sum(F23[:i+1]):.2f}'))
chk('net PPE = gross book value less accumulated depreciation',
    lambda i:(abs(F47[i]-(F63[i]-F62[i]))<T,''))
chk('accumulated depreciation never exceeds gross book value',
    lambda i:(F62[i]<=F63[i]+T,f'accdep {F62[i]:.0f} > gross {F63[i]:.0f}'))
chk('gross book value = cumulative capex',
    lambda i:(abs(F63[i]-sum(F61[:i+1]))<T,''))
chk('loan balance = cumulative drawdowns',
    lambda i:(abs(F51[i]-sum(row(FS,36)[:i+1]))<T,''))

print('\n=== SIGN CONVENTIONS ===')
for nm,vals in (('revenue is never negative',F6),('gross profit rows',None)):
    pass
chk('revenue is never negative',lambda i:(F6[i]>=-T,f'{F6[i]:.2f}'))
chk('cost of goods sold is shown negative',lambda i:(F7[i]<=T,f'{F7[i]:.2f}'))
for lbl,v in (('R&D',F11),('sales and marketing',F12),('G&A',F13),('total',F14)):
    chk(f'operating expenses shown negative: {lbl}',lambda i,v=v:(v[i]<=T,f'{v[i]:.2f}'))
chk('depreciation is shown negative',lambda i:(F19[i]<=T,f'{F19[i]:.2f}'))
chk('interest is shown negative',lambda i:(F20[i]<=T,f'{F20[i]:.2f}'))
chk('tax is a charge, never a credit',lambda i:(F22[i]<=T,f'{F22[i]:.2f}'))
chk('capex line in the cash flow is negative',lambda i:(F33[i]<=T,f'{F33[i]:.2f}'))
chk('capex schedule itself is positive',lambda i:(F61[i]>=-T,f'{F61[i]:.2f}'))
chk('receivables never negative',lambda i:(F44[i]>=-T,''))
chk('inventory never negative',lambda i:(F45[i]>=-T,''))
chk('payables never negative',lambda i:(F50[i]>=-T,''))
chk('tax losses carried forward never negative',lambda i:(F64[i]>=-T,''))
chk('equity raised never negative',lambda i:(F35[i]>=-T,''))

print('\n=== OPERATING LOGIC ===')
R11,R24,R31,R34,R16,R20,R47,R41=[row(RF,x) for x in (12,22,31,34,16,19,47,41)]
C16=row(CG,16)
AS_lbl={}
for rr in range(1,200):
    b=AS.cell(rr,2).value
    if isinstance(b,str): AS_lbl.setdefault(b.strip(),rr)
sell_from=AS.cell(AS_lbl['First month we can sell'],6).value.date()
hire_from=AS.cell(AS_lbl['Hiring starts from'],6).value.date()
freeze_to=AS.cell(AS_lbl['Committed 2026 plan holds until'],6).value.date()
chk('units never exceed demand',lambda i:(R34[i]<=R11[i]+0.5,f'{R34[i]} vs demand {R11[i]:.1f}'))
chk('units never exceed selling capacity',lambda i:(R34[i]<=R24[i]+0.5,f'{R34[i]} vs {R24[i]:.1f}'))
chk('units never exceed build capacity',lambda i:(R34[i]<=R31[i]+0.5,f'{R34[i]} vs {R31[i]:.1f}'))
chk('no units before the first sellable month',
    lambda i:(True,'') if md(i)>=sell_from else (abs(R34[i])<T,f'{R34[i]}'))
chk('no reps hired before the first sellable month',
    lambda i:(True,'') if md(i)>=sell_from else (abs(R16[i])<T,f'{R16[i]}'))
chk('no partners signed before the first sellable month',
    lambda i:(True,'') if md(i)>=sell_from else (abs(R20[i])<T,f'{R20[i]}'))
chk('installation is a true pass-through, revenue equals cost',
    lambda i:(abs(R47[i]-C16[i])<T,f'{R47[i]:.2f} vs {C16[i]:.2f}'))
chk('installed base never falls',lambda i:(True,'') if i==0 else (R41[i]>=R41[i-1]-T,''))
for rr in range(6,24):
    vals=row(PE,rr)
    chk(f'headcount never negative: {str(PE.cell(rr,2).value)[:34]}',
        lambda i,v=vals:(v[i]>=-T,f'{v[i]}'))

print('\n=== THE FROZEN 2026 PLAN ===')
KNOWN={30:[128520-0,0,0,0,0,0,0,0,0,0]}
TH_RND=[50700+25000,50700+25000,50700+25000,50700+25000,50700+25000,50700+25000,
        50700+25000,50700+25000,54700+25000,61620+125000]
TH_SM =[0+21000,10800+21000,10800+21000,10800+21000,10800+31000,10800+31000,
        10800+21000,10800+21000,10800+21000,19800+24000]
TH_GA =[11820+20000,11820+20000,11820+20000,11820+20000,11820+20000,11820+20000,
        11820+20000,11820+20000,25740+20000,21420+20000]
O30,O31,O32=row(OP,30),row(OP,31),row(OP,32)
for lbl,got,want in (('R&D',O30,TH_RND),('sales and marketing',O31,TH_SM),('G&A',O32,TH_GA)):
    chk(f'frozen months match the existing plan: {lbl}',
        lambda i,g=got,w=want:(True,'') if md(i)>freeze_to else
        (abs(g[i]-w[i])<T,f'{g[i]:,.0f} vs {w[i]:,.0f}'))

print('\n=== ANNUAL COLUMNS ===')
def annual_kind(ws,r):
    fm=F[ws.title].cell(r,AC0).value
    if not isinstance(fm,str): return None
    if fm.startswith('=SUM('): return 'sum'
    if fm.startswith('=AVERAGE') or 'AVERAGE(' in fm: return 'avg'
    if fm.startswith('=MIN('): return 'min'
    return 'end'
bad_ann=0; classified={'sum':[],'end':[],'avg':[],'min':[]}
for ws in (RF,CG,PE,OP,FS):
    for r in range(4,70):
        k=annual_kind(ws,r)
        if k is None: continue
        mv=row(ws,r)
        for y in YEARS:
            got=n(ws[f'{YC[y]}{r}'].value)
            sl=mv[(y-2026)*12:(y-2026)*12+12]
            want={'sum':sum(sl),'end':sl[-1],
                  'avg':(sum(sl)/12 if sl else 0),'min':min(sl)}[k]
            if abs(got-want)>max(0.05,abs(want)*1e-9):
                bad_ann+=1
                print(f'  FAIL  {ws.title}!{YC[y]}{r} ({k}) got {got:,.2f} want {want:,.2f}')
        classified[k].append(f'{ws.title} r{r} {str(ws.cell(r,2).value)[:30]}')
print(f'  {"pass" if bad_ann==0 else "FAIL"}  every annual column matches its stated basis '
      f'({sum(len(v) for v in classified.values())} rows checked)')
if bad_ann: fails.append('annual columns')

print('\n=== DASHBOARD RECONCILIATIONS ===')
d=lambda ref: n(DB[ref].value)
ok=abs(d('D43')-d('D47'))<T
print(f'  {"pass" if ok else "FAIL"}  sources equal uses: {d("D43"):,.2f} vs {d("D47"):,.2f}')
if not ok: fails.append('sources and uses')
ok2=abs(d('D34'))<T
print(f'  {"pass" if ok2 else "FAIL"}  balance sheet check headline is nil: {d("D34"):.6f}')
if not ok2: fails.append('bs headline')
for r,lbl in ((6,'units'),(14,'revenue'),(18,'EBITDA'),(24,'headcount'),(31,'cash')):
    src={6:(RF,34),14:(FS,6),18:(FS,16),24:(PE,20),31:(FS,40)}[r]
    bad=[y for y in YEARS if abs(n(DB[f'{gl(4+YEARS.index(y))}{r}'].value)
                                -n(src[0][f'{YC[y]}{src[1]}'].value))>T]
    print(f'  {"pass" if not bad else "FAIL"}  dashboard {lbl} ties to its source sheet'
          + (f' (off in {bad})' if bad else ''))
    if bad: fails.append(f'dashboard {lbl}')

print('\n=== STOCK ROWS THAT USE A SUM, AND FLOW ROWS THAT USE DECEMBER ===')
print('  (for review: a stock summed over 12 months, or a flow taking only December, is a bug)')
for k in ('sum','end'):
    print(f'  {k}: ' + ' | '.join(x.split(' ',1)[1] for x in classified[k][:0]) )
print(f'\n{"ALL CHECKS PASS" if not fails else "FAILURES: " + ", ".join(fails)}')
