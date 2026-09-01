"""
audit_v2.py -- full correctness audit of the Tarnoc v2 model.

    python3 scripts/audit_v2.py models/Tarnoc_v2_2026-09-01.xlsx

Four phases, run against a real LibreOffice recalculation, not against openpyxl's
view of the formulas:

  1. Recalculate every combination of the two switches and look for formula
     errors and a balance sheet that does not tie.
  2. Compare the workbook against an independent Python reimplementation of the
     whole model, cell by cell. Any disagreement is a bug in one of them.
  3. Check the accounting identities and sign conventions that a shadow model
     cannot catch, because a shadow can share the same wrong intent.
  4. Check the structure: no stray hardcodes, every cross-sheet reference lands
     on the row it claims, the live column reads its own row, no orphans.

Exits non-zero if anything fails.
"""
import os, subprocess, sys, tempfile, shutil
import openpyxl

SOFFICE = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
HERE = os.path.dirname(os.path.abspath(__file__))
ERRS = ('#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A', '#NULL!', '#NUM!', 'Err:')
CASE_CELL, TIER_CELL, CHECK_ROW = 'E5', 'E6', 58


def recalc(path, outdir):
    os.makedirs(outdir, exist_ok=True)
    subprocess.run([SOFFICE, '--headless', '--norestore', '--convert-to', 'xlsx',
                    '--outdir', outdir, path], check=True, capture_output=True)
    return os.path.join(outdir, os.path.basename(path))


def variant(src, tmp, case, tier):
    wb = openpyxl.load_workbook(src)
    wb['Assumptions'][CASE_CELL] = case
    wb['Assumptions'][TIER_CELL] = tier
    p = os.path.join(tmp, f'v_{case}{tier}.xlsx')
    wb.save(p)
    return recalc(p, os.path.join(tmp, f'o{case}{tier}'))


def phase1(recalced):
    print('PHASE 1  formula errors and the balance sheet check')
    ok = True
    for (case, tier), path in sorted(recalced.items()):
        v = openpyxl.load_workbook(path, data_only=True)
        errs = [f'{ws.title}!{c.coordinate}={c.value}' for ws in v for row in ws.iter_rows()
                for c in row if isinstance(c.value, str) and any(e in c.value for e in ERRS)]
        fs = v['Financial Statements']
        chk = [fs.cell(CHECK_ROW, c).value for c in range(2, fs.max_column + 1)]
        worst = max((abs(x) for x in chk if isinstance(x, (int, float))), default=0.0)
        good = not errs and worst < 0.01
        ok &= good
        print(f'  {"pass" if good else "FAIL"}  case {case}, tier basis {tier}: '
              f'{len(errs)} errors, worst balance check {worst:.4f}')
        for e in errs[:5]:
            print('        ', e)
    return ok


def run(script, args, keep=('MISMATCH', 'FAIL', 'agree', 'ALL ', 'FAILURES')):
    r = subprocess.run([sys.executable, os.path.join(HERE, script)] + args,
                       capture_output=True, text=True)
    out = [l for l in r.stdout.splitlines() if any(k in l for k in keep)]
    for l in out:
        print('  ' + l.strip())
    return 'FAIL' not in r.stdout and 'MISMATCH' not in r.stdout


def main():
    src = os.path.abspath(sys.argv[1])
    tmp = tempfile.mkdtemp(prefix='audit_v2_')
    try:
        recalced = {(c, t): variant(src, tmp, c, t) for c in (1, 2) for t in (1, 2)}
        ok = phase1(recalced)
        print('\nPHASE 2  independent shadow model, cell by cell')
        for case in (1, 2):
            print(f'  case {case}:')
            ok &= run('audit_v2_shadow.py', [recalced[(case, 1)], str(case)])
        print('\nPHASE 3  accounting identities, signs and operating logic')
        ok &= run('audit_v2_identities.py', [recalced[(2, 1)]])
        ok &= run('audit_v2_identities.py', [recalced[(1, 1)]])
        print('\nPHASE 4  structure')
        ok &= run('audit_v2_structure.py', [src])
        print('\n' + ('AUDIT PASSED' if ok else 'AUDIT FAILED'))
        sys.exit(0 if ok else 1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == '__main__':
    main()
