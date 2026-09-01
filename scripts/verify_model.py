"""
verify_model.py — recalculate a model with LibreOffice and check it.

    python3 scripts/verify_model.py models/Some_Model.xlsx [--switch Assumptions!D82=1,2]

openpyxl never evaluates formulas, so a model can look fine and be full of
errors. This drives LibreOffice headless to do a real recalculation, then
reports formula errors and the balance-sheet check row. Pass --switch to run
every combination of one or more scenario switches.

Requires LibreOffice, and its 'always recalculate on load' setting:
  org.openoffice.Office.Calc/Formula/Load/OOXMLRecalcMode = 0
(set once in ~/Library/Application Support/LibreOffice/4/user/registrymodifications.xcu)
"""
import argparse, itertools, os, shutil, subprocess, sys, tempfile
import openpyxl

SOFFICE = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
ERRS = ('#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A', '#NULL!', '#NUM!', 'Err:')
CHECK_ROW = 91          # 'CHECK (Total Assets - Total E&L = 0)' on Financial Statements


def recalc(path, workdir):
    subprocess.run([SOFFICE, '--headless', '--norestore', '--convert-to', 'xlsx',
                    '--outdir', workdir, path],
                   check=True, capture_output=True)
    return os.path.join(workdir, os.path.basename(path))


def inspect(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    errors = [f'{ws.title}!{c.coordinate} {c.value}'
              for ws in wb for row in ws.iter_rows() for c in row
              if isinstance(c.value, str) and any(e in c.value for e in ERRS)]
    worst = 0.0
    if 'Financial Statements' in wb.sheetnames:
        fs = wb['Financial Statements']
        vals = [fs.cell(CHECK_ROW, c).value for c in range(2, fs.max_column + 1)]
        worst = max((abs(v) for v in vals if isinstance(v, (int, float))), default=0.0)
    return errors, worst


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('model')
    ap.add_argument('--switch', action='append', default=[],
                    help="e.g. 'Assumptions!D82=1,2' — repeatable; all combinations are run")
    args = ap.parse_args()

    switches = []
    for spec in args.switch:
        ref, vals = spec.split('=')
        sheet, cell = ref.split('!')
        switches.append((sheet, cell, [int(v) for v in vals.split(',')]))

    combos = list(itertools.product(*[s[2] for s in switches])) or [()]
    tmp = tempfile.mkdtemp(prefix='verify_')
    failed = False
    try:
        for combo in combos:
            src = args.model
            if combo:
                wb = openpyxl.load_workbook(src)
                for (sheet, cell, _), val in zip(switches, combo):
                    wb[sheet][cell] = val
                src = os.path.join(tmp, 'variant_' + '_'.join(map(str, combo)) + '.xlsx')
                wb.save(src)
            out = os.path.join(tmp, 'out_' + '_'.join(map(str, combo)) or 'out')
            os.makedirs(out, exist_ok=True)
            errors, worst = inspect(recalc(src, out))
            tag = ', '.join(f'{c}={v}' for (_, c, _), v in zip(switches, combo)) or 'as shipped'
            ok = not errors and worst < 0.01
            failed |= not ok
            print(f'{"PASS" if ok else "FAIL"}  {tag:<24} errors={len(errors):<5} worst BS check={worst:.4f}')
            for e in errors[:8]:
                print(f'         {e}')
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    sys.exit(1 if failed else 0)


if __name__ == '__main__':
    main()
