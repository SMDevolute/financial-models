"""
build_tarnoc_growth_engine.py
=============================
Rebuild the Tarnoc aggressive case so that UNITS SOLD is an OUTPUT of a real
operating plan, not a hardcoded wish.

Before:  Assumptions!D21:G21 = typed numbers (1,800 / 5,200 / 11,000 / 22,000).
         Everything downstream was arithmetic on an assertion.

After:   a new 'Growth Engine' tab computes, month by month Jan-2027..Dec-2030:

           demand      = marketing spend / CPL x lead-qual% x qual-won%
           sales cap   = ramped direct reps x quota x attainment
                         + ramped installer partners x units/partner
           build cap   = assembly-partner capacity + in-house lines online
           units sold  = MIN(demand, sales capacity, build capacity, market cap)

         and names the BINDING CONSTRAINT in every month. Assumptions!D21:G21
         then reads that output, so the whole model is driven by the plan.

Also:
  - capex is derived from the in-house line schedule (online date less lead
    time), replacing the invented flat EUR5.5m
  - the incremental sales org the plan requires (reps + partner managers) is
    costed into OPEX, so growth is not free
  - implied market share is computed against the ~400k/yr NL gas-boiler
    REPLACEMENT market (the real TAM for a 1-for-1 boiler swap), not the
    125k/yr heat-pump market
  - a 'Diligence' tab surfaces the numbers a VC asks for first

Source: models/Tarnoc_LIVE_2026-08-31_10m-aggressive.xlsx
"""
import copy, os, sys
import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import house_style as hs

SRC = os.environ.get('SRC', 'working.xlsx')
OUT = os.environ.get('OUT', 'growth.xlsx')

# ===========================================================================
# DRIVER VALUES  (Base, Aggressive) — the plan, in one editable place
# ===========================================================================
D = {
    # --- market -----------------------------------------------------------
    'nl_dwellings'      : (8_200_000, 8_200_000),
    'nl_boiler_repl'    : (400_000,   400_000),   # gas boilers sold/yr in NL
    'suitable_pct'      : (0.55,      0.55),      # suitable for 1-for-1 swap
    'exp_boiler_repl'   : (0,         750_000),   # DE + BE replacement market
    'exp_entry_year'    : (2031,      2029),
    'max_share'         : (0.03,      0.08),      # sanity ceiling on share
    # --- demand funnel ----------------------------------------------------
    'mkt_2027'          : (30_000,    66_000),    # EUR / month
    'mkt_2028'          : (72_000,    244_000),
    'mkt_2029'          : (136_000,   410_000),
    'mkt_2030'          : (236_000,   670_000),
    'cpl'               : (120,       120),       # EUR cost per lead
    'cpl_improve'       : (0.10,      0.10),      # per installed-base doubling
    'lead_to_qual'      : (0.50,      0.50),
    'qual_to_won'       : (0.40,      0.40),      # => 20% lead->sale, EUR600 CAC
    # --- sales motion -----------------------------------------------------
    'direct_2027'       : (0.80,      0.70),      # share of units sold direct
    'direct_2028'       : (0.60,      0.40),
    'direct_2029'       : (0.45,      0.22),
    'direct_2030'       : (0.35,      0.15),
    'reps_start'        : (3,         3),         # quota carriers at Jan-2027
    'reps_hire_2027'    : (0.25,      0.75),      # FTE added per month
    'reps_hire_2028'    : (0.25,      0.75),
    'reps_hire_2029'    : (0.25,      0.50),
    'reps_hire_2030'    : (0.25,      0.50),
    'quota'             : (20,        20),        # units/ramped rep/month
    'rep_ramp'          : (6,         6),         # months to full productivity
    'attainment'        : (0.85,      0.85),
    'ptr_start'         : (2,         2),         # active installer partners
    'ptr_hire_2027'     : (1.0,       2.0),       # partners recruited / month
    'ptr_hire_2028'     : (2.0,       4.0),
    'ptr_hire_2029'     : (3.0,       8.0),
    'ptr_hire_2030'     : (3.8,       10.0),
    'ptr_ramp'          : (4,         4),
    'units_per_ptr'     : (8,         8),         # units/ramped partner/month
    'ptr_per_pm'        : (18,        18),        # partners per Partner Mgr
    'cost_rep'          : (7_500,     7_500),     # loaded EUR/month
    'cost_pm'           : (8_000,     8_000),
    # --- manufacturing ----------------------------------------------------
    'partner_cap'       : (650,       1_000),     # assembly partner units/mo
    'line1_online'      : ('2031-01-01', '2028-07-01'),
    'line2_online'      : ('2031-01-01', '2029-07-01'),
    'line_cap'          : (1_000,     1_000),     # units/month per line
    'line_capex'        : (2_500_000, 2_500_000),
    'tooling_capex'     : (1_000_000, 1_000_000), # one-off with line 1
    'line_leadtime'     : (12,        12),        # months capex before online
    'line_opex'         : (90_000,    90_000),    # facility+maintenance per live line
    # --- service (check only; delivery cost already in service COGS) ------
    'visits_per_tech'   : (750,       750),
    # --- organisation ratios: headcount falls out of volume & installed base
    'units_per_sc_fte'  : (900,       900),     # buying EUR114m of tiered components + inbound QC
    'operators_per_line': (35,        35),      # assembly, balancing, leak + run-in + electrical test
    'ib_per_support'    : (2_500,     2_500),   # heating appliance, winter peak, first-generation product
    'ib_per_esc'        : (15_000,    15_000),  # installed units per escalation FTE
    'ptr_per_trainer'   : (40,        40),      # new partners/yr per installer trainer
    'units_per_salesops': (3_000,     3_000),   # units/yr per order-desk FTE
    'mkt_per_mkt_fte'   : (3_000_000, 3_000_000), # marketing spend/yr per marketer
    'qa_base'           : (3,         3),       # quality & certification, base
    'qa_per_market'     : (2,         2),       # extra QA/cert FTE per new market
    'fte_per_ga'        : (9,         9),       # three countries, a plant, investor reporting
    'cost_commercial'   : (7_000,     7_000),   # loaded EUR/month
    'cost_ops_fte'      : (5_000,     5_000),
    'cost_support'      : (4_800,     4_800),
    'cost_ga_fte'       : (7_000,     7_000),
    # --- non-personnel costs that scale with the org and the installed base
    'fac_per_fte'       : (700,       700),     # facilities EUR/FTE/month
    'it_per_fte'        : (250,       250),     # IT & software EUR/FTE/month
    'travel_per_fte'    : (300,       300),     # travel EUR/FTE/month
    'recruit_per_hire'  : (8_000,     8_000),   # cost per net new hire
    'enable_per_partner': (3_500,     3_500),   # training + demo unit per new partner
    'warranty_per_unit' : (200,       200),     # reserve ABOVE the 3% in the BOM
    'entry_per_market'  : (400_000,   400_000), # one-off per new country
    'cost_rnd'          : (8_000,     8_000),   # loaded EUR/month per engineer
    'rnd_fte_2027'      : (0,         0),       # R&D engineers ABOVE the roster
    'rnd_fte_2028'      : (4,         15),
    'rnd_fte_2029'      : (8,         40),
    'rnd_fte_2030'      : (12,        60),
}

# Assumptions row for each driver, laid out in a new block
ROW = {}
_r = 132
_LAYOUT = [
    ('sec', 'MARKET  —  the TAM is the gas-boiler REPLACEMENT market, not the heat-pump market'),
    ('nl_dwellings',   'NL dwellings',                              'units',    'num'),
    ('nl_boiler_repl', 'NL gas boilers sold per year (replacement market)', 'units/yr', 'num'),
    ('suitable_pct',   'Share suitable for a 1-for-1 turbineketel swap', '%',   'pct'),
    ('exp_boiler_repl','Expansion market (BE+DE) boilers per year',  'units/yr', 'num'),
    ('exp_entry_year', 'Expansion market entry year',                'year',     'year'),
    ('max_share',      'Ceiling on share of serviceable market',     '%',        'pct'),
    ('sec', 'DEMAND FUNNEL  —  marketing spend drives units, not the other way round'),
    ('mkt_2027',       'Marketing spend — 2027',                     'EUR/month','num'),
    ('mkt_2028',       'Marketing spend — 2028',                     'EUR/month','num'),
    ('mkt_2029',       'Marketing spend — 2029',                     'EUR/month','num'),
    ('mkt_2030',       'Marketing spend — 2030',                     'EUR/month','num'),
    ('cpl',            'Cost per lead (CPL)',                        'EUR',      'num'),
    ('cpl_improve',    'CPL improvement per installed-base doubling','%',        'pct'),
    ('lead_to_qual',   'Lead -> qualified',                          '%',        'pct'),
    ('qual_to_won',    'Qualified -> won',                           '%',        'pct'),
    ('sec', 'SALES MOTION  —  shifts from direct selling to an enabled installer channel'),
    ('direct_2027',    'Direct motion share — 2027',                 '%',        'pct'),
    ('direct_2028',    'Direct motion share — 2028',                 '%',        'pct'),
    ('direct_2029',    'Direct motion share — 2029',                 '%',        'pct'),
    ('direct_2030',    'Direct motion share — 2030',                 '%',        'pct'),
    ('reps_start',     'Quota-carrying reps at Jan-2027',            'FTE',      'num1'),
    ('reps_hire_2027', 'Reps hired per month — 2027',                'FTE/month','num2'),
    ('reps_hire_2028', 'Reps hired per month — 2028',                'FTE/month','num2'),
    ('reps_hire_2029', 'Reps hired per month — 2029',                'FTE/month','num2'),
    ('reps_hire_2030', 'Reps hired per month — 2030',                'FTE/month','num2'),
    ('quota',          'Quota per fully-ramped rep',                 'units/mo', 'num'),
    ('rep_ramp',       'Rep ramp to full productivity',              'months',   'num'),
    ('attainment',     'Quota attainment',                           '%',        'pct'),
    ('ptr_start',      'Active installer partners at Jan-2027',      'partners', 'num'),
    ('ptr_hire_2027',  'Installer partners recruited per month — 2027','/month', 'num1'),
    ('ptr_hire_2028',  'Installer partners recruited per month — 2028','/month', 'num1'),
    ('ptr_hire_2029',  'Installer partners recruited per month — 2029','/month', 'num1'),
    ('ptr_hire_2030',  'Installer partners recruited per month — 2030','/month', 'num1'),
    ('ptr_ramp',       'Partner ramp to full productivity',          'months',   'num'),
    ('units_per_ptr',  'Units per fully-ramped partner per month',   'units/mo', 'num'),
    ('ptr_per_pm',     'Installer partners per Partner Manager',     'partners', 'num'),
    ('cost_rep',       'Loaded cost per sales rep',                  'EUR/month','num'),
    ('cost_pm',        'Loaded cost per partner manager',            'EUR/month','num'),
    ('sec', 'MANUFACTURING  —  assembly partner carries the ramp, in-house takes over'),
    ('partner_cap',    'Assembly partner capacity',                  'units/mo', 'num'),
    ('line1_online',   'In-house line 1 online',                     'date',     'date'),
    ('line2_online',   'In-house line 2 online',                     'date',     'date'),
    ('line_cap',       'Capacity per in-house line',                 'units/mo', 'num'),
    ('line_capex',     'Capex per in-house line',                    'EUR',      'num'),
    ('tooling_capex',  'Tooling & automation (one-off, with line 1)','EUR',      'num'),
    ('line_leadtime',  'Capex lead time before a line goes live',    'months',   'num'),
    ('line_opex',      'Facility & maintenance per live in-house line','EUR/month','num'),
    ('sec', 'SERVICE  —  check only; field labour already sits in service COGS'),
    ('visits_per_tech','Service visits per technician per year',     'visits',   'num'),
]
_LAYOUT2 = [
    ('sec', 'ORGANISATION RATIOS  —  headcount is an output of volume and installed base'),
    ('units_per_sc_fte',  'Units per supply-chain / planning / logistics FTE', 'units/yr','num'),
    ('operators_per_line','Production operators per live in-house line', 'FTE',      'num'),
    ('ib_per_support',    'Installed units per customer-support agent',  'units',    'num'),
    ('ib_per_esc',        'Installed units per technical-escalation FTE','units',    'num'),
    ('ptr_per_trainer',   'New partners per year per installer trainer', 'partners', 'num'),
    ('units_per_salesops','Units per year per order-desk FTE',           'units/yr', 'num'),
    ('mkt_per_mkt_fte',   'Marketing spend per year per marketer',       'EUR/yr',   'num'),
    ('qa_base',           'Quality & certification — base team',         'FTE',      'num'),
    ('qa_per_market',     'Extra quality & certification per new market','FTE',      'num'),
    ('fte_per_ga',        'Staff per G&A FTE (finance, HR, IT, legal)',  'FTE',      'num'),
    ('sec', 'LOADED COST PER PERSON'),
    ('cost_commercial',   'Commercial (trainers, order desk, marketing)','EUR/month','num'),
    ('cost_ops_fte',      'Supply chain & production',                   'EUR/month','num'),
    ('cost_support',      'Customer support & escalation',               'EUR/month','num'),
    ('cost_ga_fte',       'Quality, certification and G&A',              'EUR/month','num'),
    ('sec', 'COSTS THAT SCALE WITH THE ORGANISATION'),
    ('fac_per_fte',       'Facilities per person',                       'EUR/month','num'),
    ('it_per_fte',        'IT & software per person',                    'EUR/month','num'),
    ('travel_per_fte',    'Travel per person',                           'EUR/month','num'),
    ('recruit_per_hire',  'Recruitment cost per net new hire',           'EUR',      'num'),
    ('enable_per_partner','Installer training + demo unit per new partner','EUR',    'num'),
    ('warranty_per_unit', 'Warranty reserve per unit, above the 3% in the BOM','EUR','num'),
    ('entry_per_market',  'One-off cost of entering a new country',      'EUR',      'num'),
    ('cost_rnd',          'R&D engineer',                                'EUR/month','num'),
    ('rnd_fte_2027',      'Extra R&D engineers above the roster — 2027',  'FTE',      'num'),
    ('rnd_fte_2028',      'Extra R&D engineers above the roster — 2028',  'FTE',      'num'),
    ('rnd_fte_2029',      'Extra R&D engineers above the roster — 2029',  'FTE',      'num'),
    ('rnd_fte_2030',      'Extra R&D engineers above the roster — 2030',  'FTE',      'num'),
]
SECROWS = []
for item in _LAYOUT:
    if item[0] == 'sec':
        SECROWS.append((_r, item[1])); _r += 1
    else:
        ROW[item[0]] = _r; _r += 1
_r += 1
for item in _LAYOUT2:
    if item[0] == 'sec':
        SECROWS.append((_r, item[1])); _r += 1
    else:
        ROW[item[0]] = _r; _r += 1
LAST_ASSUMP_ROW = _r
print(f'Assumptions growth-engine block: rows 132..{LAST_ASSUMP_ROW-1}')


wb = openpyxl.load_workbook(SRC)
A  = wb['Assumptions']; FS = wb['Financial Statements']; RF = wb['Revenue Forecast']
CG = wb['COGS'];        PE = wb['Personnel'];            OP = wb['OPEX']
HR = wb['How to read me']

# ---- style helpers: clone the look of an existing comparable cell ---------
def restyle(ws, dst, src):
    s, d = ws[src], ws[dst]
    d.font = copy.copy(s.font); d.fill = copy.copy(s.fill)
    d.border = copy.copy(s.border); d.alignment = copy.copy(s.alignment)
    d.number_format = s.number_format

def put(ws, ref, val, style_from=None, fmt=None):
    ws[ref] = val
    if style_from: restyle(ws, ref, style_from)
    if fmt: ws[ref].number_format = fmt

NF = {'num': '#,##0;(#,##0)', 'num1': '#,##0.0;(#,##0.0)', 'num2': '#,##0.00;(#,##0.00)',
      'pct': '0.0%;(0.0%)', 'pct0': '0%;(0%)', 'year': '0', 'date': 'mmm-yy',
      'eur': '#,##0;(#,##0)', 'txt': '@'}

# ===========================================================================
# 1. ASSUMPTIONS — the growth-engine driver block
# ===========================================================================
put(A, 'B130', 'GROWTH ENGINE DRIVERS', 'B65')
for c in 'CDEFG': restyle(A, f'{c}130', 'C65')
for lbl, col in (('Driver', 'B'), ('Unit', 'C'), ('Base', 'D'),
                 ('Aggressive', 'E'), ('LIVE (per case)', 'F')):
    put(A, f'{col}131', lbl, 'B66')

for rownum, text in SECROWS:
    put(A, f'B{rownum}', text, 'B66')
    A[f'B{rownum}'].font = Font(name='Arial', size=9, bold=True, italic=True, color='2A3A55')

import datetime as _dt
def _parse(v):
    return _dt.datetime.strptime(v, '%Y-%m-%d') if isinstance(v, str) else v

FMT_OF = {}
for item in _LAYOUT + _LAYOUT2:
    if item[0] != 'sec': FMT_OF[item[0]] = item[3]
LBL_OF = {i[0]: (i[1], i[2]) for i in _LAYOUT + _LAYOUT2 if i[0] != 'sec'}

for key, r in ROW.items():
    lbl, unit = LBL_OF[key]
    base, aggr = D[key]
    fmt = NF[FMT_OF[key]]
    put(A, f'B{r}', lbl, 'B66')
    put(A, f'C{r}', unit, 'B66'); A[f'C{r}'].font = Font(name='Arial', size=9, color='6B7686')
    put(A, f'D{r}', _parse(base), 'D15', fmt)
    put(A, f'E{r}', _parse(aggr), 'D15', fmt)
    put(A, f'F{r}', f'=IF($D$82=2,E{r},D{r})', 'D21', fmt)

def live(key):
    """Absolute reference to the case-driven value of a driver."""
    return f'Assumptions!$F${ROW[key]}'

# a few explanatory notes in the notes column
NOTES = {
 'nl_boiler_repl': 'the real TAM: a turbineketel replaces a gas boiler 1-for-1, so the market is boiler '
                   'replacements (~400k/yr in NL), not heat pumps (~125k/yr in 2024)',
 'cpl':            'EUR120 CPL x 20% lead-to-sale = EUR600 marketing per customer — which is exactly the '
                   'EUR600 already assumed in row 102. The funnel is a bridge to the existing number, not a new one',
 'qual_to_won':    '50% x 40% = 20% lead-to-sale, in line with HVAC benchmarks (18-32% close on high-intent leads)',
 'direct_2030':    'the core business change: at scale Tarnoc cannot close 22k units itself. Selling shifts to '
                   'an enabled installer channel, which is why partner managers replace reps',
 'partner_cap':    'the assembly partner carries the ramp; in-house lines take over as they come live',
 'line_leadtime':  'capex is spent this many months BEFORE a line produces — that is what makes the raise urgent',
 'max_share':      'a hard sanity ceiling: units can never exceed this share of the serviceable market',
}
for k, txt in NOTES.items():
    put(A, f'L{ROW[k]}', txt, 'L15')

# the old hardcoded volume rows become a memo line, kept for comparison
put(A, 'B84', 'MEMO — previous hardcoded aggressive plan (no longer drives anything)', 'B66')
put(A, 'B85', 'MEMO — previous hardcoded base plan (no longer drives anything)', 'B66')
put(A, 'L84', 'superseded by the Growth Engine tab: units are now an output of the operating plan', 'L15')
print('assumptions block written')

# ===========================================================================
# 2. GROWTH ENGINE TAB
# ===========================================================================
if 'Growth Engine' in wb.sheetnames: del wb['Growth Engine']
GE = wb.create_sheet('Growth Engine', wb.sheetnames.index('Assumptions') + 1)
GE.sheet_view.showGridLines = False

M0, NM = 6, 48                      # F = Jan-2027, 48 months to Dec-2030
MC = [gl(M0 + i) for i in range(NM)]
YCOL = {2027: 55, 2028: 56, 2029: 57, 2030: 58}     # BC..BF annual
YR_MONTHS = {y: MC[i * 12:(i + 1) * 12] for i, y in enumerate((2027, 2028, 2029, 2030))}
NOTE_COL = 'BH'

ink, navy, muted, blue, green = '18222F', '2A3A55', '6B7686', '2F6FB0', '2F8F63'
def F(sz=10, b=False, i=False, c=ink): return Font(name='Arial', size=sz, bold=b, italic=i, color=c)

GE['B2'] = 'Tarnoc B.V. — Growth Engine'; GE['B2'].font = F(14, True, c=navy)
GE['B3'] = ('Units sold are an OUTPUT of demand, sales capacity and build capacity — not an input. '
            'Every driver lives on the Assumptions tab (rows 130+). Case switch: Assumptions D82.')
GE['B3'].font = F(9, i=True, c=muted)

def section(r, text):
    GE.cell(r, 2, text).font = F(10, True, c='FFFFFF')
    for c in range(2, 60):
        GE.cell(r, c).fill = PatternFill('solid', fgColor=navy)
        if c > 2: GE.cell(r, c).font = F(10, True, c='FFFFFF')

def rowdef(r, label, unit, formula_fn, fmt='num', bold=False, kind='formula', note=None, annual='sum'):
    """formula_fn(col_letter, i) -> formula string for month i."""
    GE.cell(r, 2, label).font = F(10, bold, c=ink)
    GE.cell(r, 3, unit).font = F(9, c=muted)
    colr = {'formula': ink, 'input': blue, 'link': green}[kind]
    for i, cl in enumerate(MC):
        cell = GE[f'{cl}{r}']
        cell.value = formula_fn(cl, i)
        cell.number_format = NF[fmt]; cell.font = F(10, bold, c=colr)
        cell.alignment = Alignment(horizontal='right')
    for y, cidx in YCOL.items():
        cl = gl(cidx); ms = YR_MONTHS[y]
        if annual == 'sum':      f = f'=SUM({ms[0]}{r}:{ms[-1]}{r})'
        elif annual == 'end':    f = f'={ms[-1]}{r}'
        elif annual == 'avg':    f = f'=AVERAGE({ms[0]}{r}:{ms[-1]}{r})'
        elif annual == 'max':    f = f'=MAX({ms[0]}{r}:{ms[-1]}{r})'
        else:                    f = None
        if f:
            c = GE[f'{cl}{r}']; c.value = f; c.number_format = NF[fmt]
            c.font = F(10, True, c=navy); c.alignment = Alignment(horizontal='right')
    if note:
        GE[f'{NOTE_COL}{r}'] = note; GE[f'{NOTE_COL}{r}'].font = F(9, i=True, c=muted)

# ---- header row ----------------------------------------------------------
GE['B5'] = 'Month'; GE['B5'].font = F(9, True, c=muted)
for i, cl in enumerate(MC):
    y, m = 2027 + i // 12, i % 12 + 1
    c = GE[f'{cl}5']; c.value = _dt.datetime(y, m, 1)
    c.number_format = 'mmm-yy'; c.font = F(9, True, c=navy)
    c.alignment = Alignment(horizontal='right')
for y, cidx in YCOL.items():
    c = GE.cell(5, cidx, y); c.font = F(10, True, c='FFFFFF')
    c.fill = PatternFill('solid', fgColor=navy); c.alignment = Alignment(horizontal='right')
GE[f'{NOTE_COL}5'] = 'what this row does'; GE[f'{NOTE_COL}5'].font = F(9, True, c=muted)

def yr(cl, per_year_keys):
    """Nested IF picking the driver for the year of this column."""
    k27, k28, k29, k30 = per_year_keys
    return (f'IF(YEAR({cl}$5)=2027,{live(k27)},IF(YEAR({cl}$5)=2028,{live(k28)},'
            f'IF(YEAR({cl}$5)=2029,{live(k29)},{live(k30)})))')

R = {}
# ---- market --------------------------------------------------------------
section(7, 'MARKET  —  serviceable demand for a 1-for-1 gas-boiler replacement')
R['mkt_nl'] = 8; rowdef(8, 'NL serviceable market', 'units/mo',
    lambda cl, i: f'={live("nl_boiler_repl")}*{live("suitable_pct")}/12', annual='sum',
    note='~400k gas boilers replaced a year in NL x the share suitable for a 1-for-1 swap')
R['mkt_exp'] = 9; rowdef(9, 'Expansion market (BE+DE)', 'units/mo',
    lambda cl, i: (f'=IF(YEAR({cl}$5)>={live("exp_entry_year")},'
                   f'{live("exp_boiler_repl")}*{live("suitable_pct")}/12,0)'),
    note='opens in the entry year set on Assumptions')
R['mkt_tot'] = 10; rowdef(10, 'Total serviceable market', 'units/mo',
    lambda cl, i: f'={cl}8+{cl}9', bold=True)
R['mkt_cap'] = 11; rowdef(11, 'Share ceiling — max units allowed', 'units/mo',
    lambda cl, i: f'={cl}10*{live("max_share")}',
    note='a hard ceiling, so the plan can never quietly assume we own an unrealistic slice of the market')

# ---- demand --------------------------------------------------------------
section(13, 'DEMAND GENERATION  —  marketing spend drives orders')
R['spend'] = 14; rowdef(14, 'Marketing spend', 'EUR/mo',
    lambda cl, i: '=' + yr(cl, ('mkt_2027', 'mkt_2028', 'mkt_2029', 'mkt_2030')), kind='link')
R['ib_open'] = 15; rowdef(15, 'Installed base, start of month', 'units',
    lambda cl, i: '=0' if i == 0 else f'={MC[i-1]}46', annual='end')
R['cpl_eff'] = 16; rowdef(16, 'Effective CPL', 'EUR',
    lambda cl, i: (f'={live("cpl")}*(1-{live("cpl_improve")})^'
                   f'IF({cl}15<400,0,LOG({cl}15/400,2))'), annual='avg',
    note='cost per lead falls 10% per doubling of installed base — brand, referral and channel pull')
R['leads'] = 17; rowdef(17, 'Leads', 'leads',
    lambda cl, i: f'=IFERROR({cl}14/{cl}16,0)')
R['sql'] = 18; rowdef(18, 'Qualified leads', 'leads',
    lambda cl, i: f'={cl}17*{live("lead_to_qual")}')
R['demand'] = 19; rowdef(19, 'DEMAND — orders generated', 'units/mo',
    lambda cl, i: f'={cl}18*{live("qual_to_won")}', bold=True)
R['cac'] = 20; rowdef(20, 'Implied CAC', 'EUR/unit',
    lambda cl, i: f'=IFERROR({cl}14/{cl}19,0)', annual='avg',
    note='reconciles to the EUR600/customer already assumed on Assumptions row 102')

# ---- sales capacity ------------------------------------------------------
section(22, 'SALES CAPACITY  —  direct reps early, enabled installer channel at scale')
R['dshare'] = 23; rowdef(23, 'Direct motion share', '%',
    lambda cl, i: '=' + yr(cl, ('direct_2027', 'direct_2028', 'direct_2029', 'direct_2030')),
    fmt='pct', kind='link', annual='avg')
R['rep_add'] = 24; rowdef(24, 'Reps hired in month', 'FTE',
    lambda cl, i: '=' + yr(cl, ('reps_hire_2027', 'reps_hire_2028', 'reps_hire_2029', 'reps_hire_2030')),
    fmt='num2', kind='link')
R['rep_hc'] = 25; rowdef(25, 'Reps on payroll', 'FTE',
    lambda cl, i: (f'={live("reps_start")}+{cl}24' if i == 0 else f'={MC[i-1]}25+{cl}24'),
    fmt='num1', annual='end')
R['rep_ramped'] = 26; rowdef(26, 'Reps fully ramped', 'FTE',
    lambda cl, i: (f'={live("reps_start")}' if i == 0 else
                   f'=IF({i}<{live("rep_ramp")},{live("reps_start")},'
                   f'INDEX($F25:{MC[-1]}25,{i+1}-{live("rep_ramp")}))'),
    fmt='num1', annual='end',
    note='a rep carries no quota until fully ramped (6 months) — deliberately conservative')
R['dcap'] = 27; rowdef(27, 'Direct capacity', 'units/mo',
    lambda cl, i: f'={cl}26*{live("quota")}*{live("attainment")}')
R['ptr_add'] = 28; rowdef(28, 'Installer partners recruited', 'partners',
    lambda cl, i: '=' + yr(cl, ('ptr_hire_2027', 'ptr_hire_2028', 'ptr_hire_2029', 'ptr_hire_2030')),
    fmt='num1', kind='link')
R['ptr_hc'] = 29; rowdef(29, 'Installer partners on books', 'partners',
    lambda cl, i: (f'={live("ptr_start")}+{cl}28' if i == 0 else f'={MC[i-1]}29+{cl}28'),
    fmt='num1', annual='end')
R['ptr_ramped'] = 30; rowdef(30, 'Installer partners productive', 'partners',
    lambda cl, i: (f'={live("ptr_start")}' if i == 0 else
                   f'=IF({i}<{live("ptr_ramp")},{live("ptr_start")},'
                   f'INDEX($F29:{MC[-1]}29,{i+1}-{live("ptr_ramp")}))'),
    fmt='num1', annual='end')
R['ccap'] = 31; rowdef(31, 'Channel capacity', 'units/mo',
    lambda cl, i: f'={cl}30*{live("units_per_ptr")}')
R['scap'] = 32; rowdef(32, 'TOTAL sales capacity', 'units/mo',
    lambda cl, i: (f'=MIN(IFERROR({cl}27/{cl}23,1000000),'
                   f'IFERROR({cl}31/(1-{cl}23),1000000))'), bold=True,
    note='how many we can actually sell: our own reps have to cover the direct share, installers the rest')
R['pm'] = 33; rowdef(33, 'Partner managers required', 'FTE',
    lambda cl, i: f'=ROUNDUP({cl}29/{live("ptr_per_pm")},0)', annual='end')

# ---- build capacity ------------------------------------------------------
section(36, 'BUILD CAPACITY  —  assembly partner first, in-house lines take over')
R['pcap'] = 37; rowdef(37, 'Assembly partner capacity', 'units/mo',
    lambda cl, i: f'={live("partner_cap")}', kind='link')
R['lines'] = 38; rowdef(38, 'In-house lines live', 'lines',
    lambda cl, i: (f'=IF({cl}$5>={live("line1_online")},1,0)+IF({cl}$5>={live("line2_online")},1,0)'),
    annual='end')
R['icap'] = 39; rowdef(39, 'In-house capacity', 'units/mo',
    lambda cl, i: f'={cl}38*{live("line_cap")}')
R['bcap'] = 40; rowdef(40, 'TOTAL build capacity', 'units/mo',
    lambda cl, i: f'={cl}37+{cl}39', bold=True)

# ---- the answer ----------------------------------------------------------
section(42, 'UNITS SOLD  —  and what stopped us selling more, month by month')
R['units'] = 43; rowdef(43, 'UNITS SOLD', 'units',
    lambda cl, i: f'=ROUND(MIN({cl}19,{cl}11,{cl}32,{cl}40),0)', bold=True,
    note='we sell the smallest of: what people want, what the market allows, what we can sell, what we can build')
R['bind'] = 44; rowdef(44, 'Why we did not sell more', '',
    lambda cl, i: (f'=IF({cl}43>={cl}40-0.5,"Build capacity",'
                   f'IF({cl}43>={cl}32-0.5,"Sales capacity",'
                   f'IF({cl}43>={cl}11-0.5,"Market ceiling","Demand")))'),
    fmt='txt', annual=None)
R['util'] = 45; rowdef(45, 'Build capacity utilisation', '%',
    lambda cl, i: f'=IFERROR({cl}43/{cl}40,0)', fmt='pct', annual='avg')

section(47, 'CONSEQUENCES  —  market share, installed base, and the org the plan needs')
R['share'] = 48; rowdef(48, 'Implied share of serviceable market', '%',
    lambda cl, i: f'=IFERROR({cl}43/{cl}10,0)', fmt='pct', annual='avg',
    note='the first number a VC computes — kept on the face of the model')
R['ib_close'] = 46; rowdef(46, 'Installed base, end of month', 'units',
    lambda cl, i: f'={cl}15+{cl}43', annual='end')
R['techs'] = 49; rowdef(49, 'Service technicians implied', 'FTE',
    lambda cl, i: (f'=ROUNDUP({cl}46*(Assumptions!$E$117+Assumptions!$E$118)/'
                   f'{live("visits_per_tech")},0)'), annual='end',
    note='check only — field labour is already inside the ~50% service COGS, so it is not charged twice')
# ---- the organisation the plan actually needs ----------------------------
def PECOL(i):
    """Personnel monthly column for Growth Engine month i (I = Jan-2025);
    Personnel stops at Dec-2029, so 2030 holds the Dec-2029 roster flat."""
    return gl(min(33 + i, 68))

MARKETS = f'(1+IF(YEAR({{cl}}$5)>={live("exp_entry_year")},2,0))'

section(56, 'ORGANISATION  —  the people this plan actually needs')
rowdef(57, 'Sales reps', 'FTE', lambda cl, i: f'={cl}25', fmt='num1', annual='end')
rowdef(58, 'Partner managers', 'FTE', lambda cl, i: f'={cl}33', annual='end')
rowdef(59, 'Installer trainers', 'FTE',
    lambda cl, i: f'=ROUNDUP({cl}28*12/{live("ptr_per_trainer")},0)', annual='end',
    note='290 installer partners have to be trained and certified before they can sell anything')
rowdef(60, 'Order desk / sales ops', 'FTE',
    lambda cl, i: f'=ROUNDUP({cl}43*12/{live("units_per_salesops")},0)', annual='end')
rowdef(61, 'Marketing team', 'FTE',
    lambda cl, i: f'=2+ROUNDUP({cl}14*12/{live("mkt_per_mkt_fte")},0)', annual='end')
rowdef(62, 'Commercial team (S&M)', 'FTE',
    lambda cl, i: f'=SUM({cl}57:{cl}61)', bold=True, annual='end')
rowdef(63, 'Supply chain, planning & logistics', 'FTE',
    lambda cl, i: f'=ROUNDUP({cl}43*12/{live("units_per_sc_fte")},0)', annual='end',
    note='someone has to buy, plan and move EUR114m of components a year')
rowdef(64, 'Production operators', 'FTE',
    lambda cl, i: f'={cl}38*{live("operators_per_line")}', annual='end')
rowdef(65, 'Customer support', 'FTE',
    lambda cl, i: f'=ROUNDUP({cl}46/{live("ib_per_support")},0)', annual='end',
    note='inbound calls scale with the installed base, not with sales')
rowdef(66, 'Technical escalation', 'FTE',
    lambda cl, i: f'=ROUNDUP({cl}46/{live("ib_per_esc")},0)', annual='end')
rowdef(67, 'Quality & certification', 'FTE',
    lambda cl, i: (f'={live("qa_base")}+{live("qa_per_market")}*'
                   + MARKETS.format(cl=cl) + '-' + live("qa_per_market")), annual='end',
    note='every new country needs its own certification work')
rowdef(68, 'R&D — roster plus new engineers', 'FTE',
    lambda cl, i: f"=Personnel!{PECOL(i)}68+" + yr(cl, ('rnd_fte_2027', 'rnd_fte_2028',
                  'rnd_fte_2029', 'rnd_fte_2030')), annual='end',
    note='a novel turbine product, the Twincycle, and certification for two new countries')
rowdef(69, 'G&A — finance, HR, IT, legal, office', 'FTE',
    lambda cl, i: (f'=ROUNDUP(({cl}62+SUM({cl}63:{cl}67)+{cl}68)/{live("fte_per_ga")},0)'),
    annual='end')
rowdef(70, 'Operations, support & G&A', 'FTE',
    lambda cl, i: f'=SUM({cl}63:{cl}67)+{cl}69', bold=True, annual='end')
rowdef(71, 'TOTAL PEOPLE ON PAYROLL', 'FTE',
    lambda cl, i: f'={cl}62+{cl}70+{cl}68', bold=True, annual='end')
rowdef(72, 'Field service technicians', 'FTE',
    lambda cl, i: f'={cl}49', annual='end',
    note='counted here but NOT charged again — their cost is already inside the ~50% service COGS')
rowdef(73, 'TOTAL HEADCOUNT', 'FTE',
    lambda cl, i: f'={cl}71+{cl}72', bold=True, annual='end')
rowdef(74, 'People in the Personnel tab today', 'FTE',
    lambda cl, i: f"=Personnel!{PECOL(i)}71", kind='link', annual='end')
rowdef(75, 'Extra people the plan needs', 'FTE',
    lambda cl, i: f'=MAX(0,{cl}71-{cl}74)', bold=True, annual='end',
    note='the gap between the roster in the Personnel tab and what this volume actually requires')

section(77, 'COST  —  what that organisation costs, on top of the roster')
rowdef(78, 'Commercial team cost required', 'EUR/mo',
    lambda cl, i: (f'={cl}57*{live("cost_rep")}+{cl}58*{live("cost_pm")}'
                   f'+({cl}59+{cl}60+{cl}61)*{live("cost_commercial")}'))
rowdef(79, 'S&M cost already in the roster', 'EUR/mo',
    lambda cl, i: f"=Personnel!{PECOL(i)}63", kind='link')
rowdef(80, 'S&M personnel uplift', 'EUR/mo',
    lambda cl, i: f'=MAX(0,{cl}78-{cl}79)', bold=True)
rowdef(81, 'Operations, support & G&A cost required', 'EUR/mo',
    lambda cl, i: (f'=({cl}63+{cl}64)*{live("cost_ops_fte")}'
                   f'+({cl}65+{cl}66)*{live("cost_support")}'
                   f'+({cl}67+{cl}69)*{live("cost_ga_fte")}'))
rowdef(82, 'G&A cost already in the roster', 'EUR/mo',
    lambda cl, i: f"=Personnel!{PECOL(i)}64", kind='link')
rowdef(83, 'G&A personnel uplift', 'EUR/mo',
    lambda cl, i: f'=MAX(0,{cl}81-{cl}82)', bold=True)
rowdef(84, 'Facilities, IT & travel', 'EUR/mo',
    lambda cl, i: (f'={cl}71*({live("fac_per_fte")}+{live("it_per_fte")}'
                   f'+{live("travel_per_fte")})'))
rowdef(85, 'Recruitment', 'EUR/mo',
    lambda cl, i: ('=0' if i == 0 else
                   f'=MAX(0,{cl}71-{MC[i-1]}71)*{live("recruit_per_hire")}'))
rowdef(86, 'Installer training & demo units', 'EUR/mo',
    lambda cl, i: f'={cl}28*{live("enable_per_partner")}')
rowdef(87, 'In-house line facility & maintenance', 'EUR/mo',
    lambda cl, i: f'={cl}38*{live("line_opex")}')
rowdef(88, 'Entering a new country', 'EUR/mo',
    lambda cl, i: (f'=IF(AND(YEAR({cl}$5)={live("exp_entry_year")},MONTH({cl}$5)=1),'
                   f'{live("entry_per_market")}*2,0)'))
rowdef(89, 'R&D — new engineers', 'EUR/mo',
    lambda cl, i: '=(' + yr(cl, ('rnd_fte_2027', 'rnd_fte_2028', 'rnd_fte_2029', 'rnd_fte_2030'))
                  + f')*{live("cost_rnd")}')
rowdef(90, 'TOTAL EXTRA OPEX', 'EUR/mo',
    lambda cl, i: f'={cl}80+{cl}83+{cl}84+{cl}85+{cl}86+{cl}87+{cl}88+{cl}89', bold=True)
rowdef(91, 'Warranty reserve above the BOM (goes to COGS)', 'EUR/mo',
    lambda cl, i: f'={cl}43*{live("warranty_per_unit")}',
    note='the BOM already carries a 3% yield/warranty provision; this is the reserve on top')

GE.column_dimensions['A'].width = 2.5
GE.column_dimensions['B'].width = 38
GE.column_dimensions['C'].width = 10
for cl in MC: GE.column_dimensions[cl].width = 10
for cidx in YCOL.values(): GE.column_dimensions[gl(cidx)].width = 12
GE.column_dimensions[NOTE_COL].width = 70
GE.freeze_panes = 'F6'
print('growth engine tab written')

# ===========================================================================
# 3. REWIRE — units, monthly revenue shape, capex, opex
# ===========================================================================
# 3a. the units dial now READS the Growth Engine instead of being typed
for c, y in zip('DEFG', (2027, 2028, 2029, 2030)):
    A[f'{c}21'] = f"='Growth Engine'!{gl(YCOL[y])}43"
    A[f'{c}21'].font = Font(name='Arial', size=10, color='2F8F63')   # green = link
put(A, 'L21', 'OUTPUT, not an input — units come from the Growth Engine tab '
              '(demand vs sales capacity vs build capacity)', 'L15')

# 3b. monthly units take the Growth Engine's real shape, replacing the
#     synthetic exponential interpolation that used to spread the annual total
for j in range(12):
    for yoff, rfbase in ((0, 29), (1, 41)):        # AC.. = 2027, AO.. = 2028
        rc = gl(rfbase + j); gc = MC[yoff * 12 + j]
        RF[f'{rc}6']  = f"='Growth Engine'!{gc}43*Assumptions!$H$23"
        RF[f'{rc}12'] = f"='Growth Engine'!{gc}43*Assumptions!$H$24"
for c in range(5, 29):                              # E..AB = 2025-2026: no units
    RF[f'{gl(c)}6'] = 0; RF[f'{gl(c)}12'] = 0
RF['BH6']  = 'units from the Growth Engine tab x the TTK share of the mix (Assumptions H23)'
RF['BH12'] = 'units from the Growth Engine tab x the Combi+ share of the mix (Assumptions H24)'

# 3c. capex is now the in-house line schedule, spent a lead time before a
#     line produces — not a flat invented number
put(A, 'B125', 'In-house line capex (per line, at order date)', 'B66')
put(A, 'B126', 'Tooling & automation (one-off, with line 1)', 'B66')
put(A, 'B127', 'Total capex', 'B66')
for c in 'DEFG':
    A[f'{c}125'] = (f'=(IF(YEAR(EDATE({live("line1_online")},-{live("line_leadtime")}))={c}124,1,0)'
                    f'+IF(YEAR(EDATE({live("line2_online")},-{live("line_leadtime")}))={c}124,1,0))'
                    f'*{live("line_capex")}')
    A[f'{c}126'] = (f'=IF(YEAR(EDATE({live("line1_online")},-{live("line_leadtime")}))={c}124,'
                    f'{live("tooling_capex")},0)')
    A[f'{c}127'] = f'={c}125+{c}126'
    for rr in (125, 126, 127): A[f'{c}{rr}'].number_format = NF['num']
put(A, 'L125', 'a line is paid for when it is ordered, which is the lead time before it can build anything', 'L15')
put(A, 'L127', 'what the model actually spends — driven entirely by the in-house line dates', 'L15')

# 3d. OPEX — the full cost of the organisation the plan needs
OP['B45'] = 'GROWTH ENGINE — THE ORGANISATION THIS PLAN NEEDS (case-driven)'
OP['B45'].font = Font(name='Arial', size=10, bold=True, color='2A3A55')
OP['B46'] = 'S&M — people uplift, installer training & demo units'
OP['B47'] = 'G&A — people uplift, facilities, IT, travel, recruitment, plant, new countries'
OP['B49'] = 'R&D — uplift'
OP['B48'] = 'Total growth-engine OPEX'
OP['B48'].font = Font(name='Arial', size=10, bold=True)
OPEX_M0 = 31                           # OPEX AE = Growth Engine F = Jan-2027
for c in range(7, 55):                 # G..BB monthly
    cl = gl(c)
    if c < OPEX_M0 or c >= OPEX_M0 + 24:
        OP[f'{cl}46'] = 0; OP[f'{cl}47'] = 0; OP[f'{cl}49'] = 0
    else:
        g = MC[c - OPEX_M0]
        OP[f'{cl}46'] = f"='Growth Engine'!{g}80+'Growth Engine'!{g}86"
        OP[f'{cl}47'] = (f"='Growth Engine'!{g}83+'Growth Engine'!{g}84"
                         f"+'Growth Engine'!{g}85+'Growth Engine'!{g}87+'Growth Engine'!{g}88")
        OP[f'{cl}49'] = f"='Growth Engine'!{g}89"
    OP[f'{cl}48'] = f'={cl}46+{cl}47+{cl}49'
for cl, y in (('BD', 2025), ('BE', 2026), ('BF', 2027), ('BG', 2028), ('BH', 2029)):
    if y in (2027, 2028):
        a, b = ('AE', 'AP') if y == 2027 else ('AQ', 'BB')
        for r in (46, 47, 49): OP[f'{cl}{r}'] = f'=SUM({a}{r}:{b}{r})'
    elif y == 2029:
        g = gl(YCOL[2029])
        OP[f'{cl}46'] = f"='Growth Engine'!{g}80+'Growth Engine'!{g}86"
        OP[f'{cl}47'] = (f"='Growth Engine'!{g}83+'Growth Engine'!{g}84"
                         f"+'Growth Engine'!{g}85+'Growth Engine'!{g}87+'Growth Engine'!{g}88")
        OP[f'{cl}49'] = f"='Growth Engine'!{g}89"
    else:
        for r in (46, 47, 49): OP[f'{cl}{r}'] = 0
    OP[f'{cl}48'] = f'={cl}46+{cl}47+{cl}49'

# fold the new lines into the department totals the P&L reads
for c in list(range(7, 55)) + [56, 57, 58, 59, 60]:      # monthly + BD..BH
    cl = gl(c)
    OP[f'{cl}38'] = f'={cl}19+{cl}6+{cl}49'      # Total R&D
    OP[f'{cl}39'] = f'={cl}26+{cl}7+{cl}46'      # Total S&M
    OP[f'{cl}40'] = f'={cl}32+{cl}8+{cl}47'      # Total G&A
    OP[f'{cl}36'] = f'={cl}34+{cl}9+{cl}48'      # TOTAL OPEX
OP['BK46'] = 'only the cost ABOVE what the Personnel tab already pays is charged here, so nobody is paid twice'
OP['BK47'] = 'overheads scale with headcount; plant and country-entry costs come off the capacity plan'
print('rewiring done')

# 3e. COGS — warranty reserve on top of the 3% already in the BOM
CG['B21'] = 'Warranty reserve (above the 3% in the BOM)'
for c in range(5, 53):                             # E..AZ monthly
    cl = gl(c)
    CG[f'{cl}21'] = (0 if c < 29 else f"='Growth Engine'!{MC[c-29]}91")
for cl, y in (('BB', 2025), ('BC', 2026), ('BD', 2027), ('BE', 2028), ('BF', 2029)):
    if y == 2027:   CG[f'{cl}21'] = '=SUM(AC21:AN21)'
    elif y == 2028: CG[f'{cl}21'] = '=SUM(AO21:AZ21)'
    elif y == 2029: CG[f'{cl}21'] = f"='Growth Engine'!{gl(YCOL[2029])}91"
    else:           CG[f'{cl}21'] = 0
for c in list(range(5, 53)) + [54, 55, 56, 57, 58]:
    cl = gl(c)
    CG[f'{cl}22'] = f'={cl}9+{cl}14+{cl}17+{cl}20+{cl}21'
CG['BH21'] = 'a real warranty reserve on a first-generation turbine machine, on top of the BOM provision'

# ===========================================================================
# 4. DILIGENCE TAB — what a VC asks for first, on one page
# ===========================================================================
if 'Diligence' in wb.sheetnames: del wb['Diligence']
DG = wb.create_sheet('Diligence')
DG.sheet_view.showGridLines = False
DG['B2'] = 'Tarnoc B.V. — Diligence Pack'; DG['B2'].font = F(14, True, c=navy)
DG['B3'] = ('Everything on this tab is a formula off the Growth Engine and the Financial Statements. '
            'Change a driver on Assumptions and every number here moves.')
DG['B3'].font = F(9, i=True, c=muted)

GY = {y: gl(YCOL[y]) for y in (2027, 2028, 2029, 2030)}   # Growth Engine annual
FY = {2027: 'BD', 2028: 'BE', 2029: 'BF'}                 # Financial Statements annual
DC = {2027: 'D', 2028: 'E', 2029: 'F', 2030: 'G'}

def dsec(r, text):
    DG.cell(r, 2, text).font = F(10, True, c='FFFFFF')
    for c in range(2, 9):
        DG.cell(r, c).fill = PatternFill('solid', fgColor=navy)
        DG.cell(r, c).font = F(10, True, c='FFFFFF')

def drow(r, label, fn, fmt='num', bold=False, note=None):
    DG.cell(r, 2, label).font = F(10, bold, c=ink)
    for y in (2027, 2028, 2029, 2030):
        v = fn(y)
        c = DG[f'{DC[y]}{r}']
        c.value = v if v is not None else 'n/a'
        c.number_format = NF[fmt] if v is not None else '@'
        c.font = F(10, bold, c=ink if v is not None else muted)
        c.alignment = Alignment(horizontal='right')
    if note:
        DG[f'I{r}'] = note; DG[f'I{r}'].font = F(9, i=True, c=muted)

for y in (2027, 2028, 2029, 2030):
    c = DG[f'{DC[y]}5']; c.value = y; c.font = F(10, True, c='FFFFFF')
    c.fill = PatternFill('solid', fgColor=navy); c.alignment = Alignment(horizontal='right')

dsec(7, 'VOLUME & MARKET')
drow(8,  'Units sold',                    lambda y: f"='Growth Engine'!{GY[y]}43", bold=True)
drow(9,  'Serviceable market (units)',    lambda y: f"='Growth Engine'!{GY[y]}10")
drow(10, 'Implied market share',          lambda y: f"='Growth Engine'!{GY[y]}48", fmt='pct',
     note='against the ~400k/yr NL boiler REPLACEMENT market, plus BE+DE once open')
drow(11, 'Installed base (year end)',     lambda y: f"='Growth Engine'!{GY[y]}46")
drow(12, 'Why we did not sell more (Dec)', lambda y: f"='Growth Engine'!{YR_MONTHS[y][-1]}44", fmt='txt',
     note='the reason we could not sell more in December of that year')
drow(13, 'Build capacity utilisation',    lambda y: f"='Growth Engine'!{GY[y]}45", fmt='pct')

dsec(15, 'UNIT ECONOMICS')
drow(16, 'Marketing spend',               lambda y: f"='Growth Engine'!{GY[y]}14")
drow(17, 'CAC (marketing / unit)',        lambda y: f"=IFERROR({DC[y]}16/{DC[y]}8,0)")
drow(18, 'Revenue',                       lambda y: f"='Financial Statements'!{FY[y]}11" if y in FY else None)
drow(19, 'Revenue per unit',              lambda y: f"=IFERROR({DC[y]}18/{DC[y]}8,0)" if y in FY else None)
drow(20, 'Gross profit',                  lambda y: f"='Financial Statements'!{FY[y]}23" if y in FY else None)
drow(21, 'Gross margin',                  lambda y: f"=IFERROR({DC[y]}20/{DC[y]}18,0)" if y in FY else None, fmt='pct')
drow(22, 'Gross profit per unit',         lambda y: f"=IFERROR({DC[y]}20/{DC[y]}8,0)" if y in FY else None)
drow(23, 'CAC payback (units of GP)',     lambda y: f"=IFERROR({DC[y]}17/{DC[y]}22,0)" if y in FY else None,
     fmt='pct', note='CAC as a share of the gross profit the unit earns on day one')

dsec(25, 'ORGANISATION')
drow(26, 'Quota-carrying reps (year end)',    lambda y: f"='Growth Engine'!{GY[y]}25", fmt='num1')
drow(27, 'Installer partners (year end)',     lambda y: f"='Growth Engine'!{GY[y]}29", fmt='num1')
drow(28, 'Partner managers required',         lambda y: f"='Growth Engine'!{GY[y]}33")
drow(29, 'Units per rep per year',            lambda y: f"=IFERROR({DC[y]}8*'Growth Engine'!{GY[y]}23/{DC[y]}26,0)",
     note='direct units only — the channel is carried by partners, not reps')
drow(30, 'Units per partner per year',        lambda y: f"=IFERROR({DC[y]}8*(1-'Growth Engine'!{GY[y]}23)/{DC[y]}27,0)")
drow(31, 'Field service technicians',         lambda y: f"='Growth Engine'!{GY[y]}72")
drow(32, 'TOTAL HEADCOUNT',                   lambda y: f"='Growth Engine'!{GY[y]}73", bold=True,
     note='people on payroll plus field service technicians')
drow(33, 'People in the Personnel tab today', lambda y: f"='Growth Engine'!{GY[y]}74")
drow(44, 'Extra people the plan needs',       lambda y: f"='Growth Engine'!{GY[y]}75", bold=True)
drow(45, 'Revenue per employee',              lambda y: f"=IFERROR({DC[y]}18/{DC[y]}32,0)" if y in FY else None,
     note='Viessmann runs at ~EUR276k, Vaillant ~EUR200k. Far above that means the org is understated')
drow(46, 'Extra cost of the organisation',    lambda y: f"='Growth Engine'!{GY[y]}90")

dsec(34, 'CAPACITY & CAPEX')
drow(35, 'Assembly partner capacity (units/yr)', lambda y: f"='Growth Engine'!{GY[y]}37")
drow(36, 'In-house lines live (year end)',       lambda y: f"='Growth Engine'!{GY[y]}38")
drow(37, 'Total build capacity (units/yr)',      lambda y: f"='Growth Engine'!{GY[y]}40")
drow(38, 'Capex',                                lambda y: f"=Assumptions!{DC[y]}127")

dsec(40, 'PROFIT & FUNDING')
drow(41, 'EBITDA',        lambda y: f"='Financial Statements'!{FY[y]}32" if y in FY else None, bold=True)
drow(42, 'EBITDA margin', lambda y: f"='Financial Statements'!{FY[y]}33" if y in FY else None, fmt='pct')
drow(43, 'Net income',    lambda y: f"='Financial Statements'!{FY[y]}41" if y in FY else None)
drow(44, 'Ending cash',   lambda y: f"='Financial Statements'!{FY[y]}62" if y in FY else None)

DG['B48'] = 'Minimum cash across the whole horizon'
DG['B48'].font = F(10, True, c=ink)
DG['D48'] = "=MIN('Financial Statements'!$E$62:$AZ$62,'Financial Statements'!$BF$62)"
DG['D48'].number_format = NF['num']; DG['D48'].font = F(10, True, c=ink)
DG['B49'] = 'Equity raised (cumulative)'
DG['D49'] = "='Financial Statements'!$AZ$86"
DG['D49'].number_format = NF['num']
DG['I48'] = 'if this is comfortably positive the raise is bigger than the plan needs — say so before a VC does'
DG['I48'].font = F(9, i=True, c=muted)

DG.column_dimensions['A'].width = 2.5
DG.column_dimensions['B'].width = 46
for cl in 'DEFG': DG.column_dimensions[cl].width = 15
DG.column_dimensions['I'].width = 72
DG.freeze_panes = 'D6'

# ===========================================================================
# 5. HOW TO READ ME — document the rebuild, drop the scratch note
# ===========================================================================
for r in (43, 44, 45):
    HR[f'B{r}'] = None
put(HR, 'B43', 'THE GROWTH ENGINE (new tab) — why the volume plan is credible', 'B33')
put(HR, 'B44',
    "Units sold used to be typed into Assumptions row 21. They are now an OUTPUT. The Growth Engine tab "
    "works out four numbers every month: how many people want one (marketing spend / cost per lead x "
    "conversion), how many we can sell (ramped reps x quota, plus trained installer partners x units each), "
    "how many we can build (assembly partner plus in-house lines), and the most the market allows. We sell "
    "the smallest of the four, and the tab writes down which one it was.", 'B34')
put(HR, 'B45',
    "The business genuinely changes in the aggressive case. Tarnoc cannot close 20,000+ units a year with a "
    "direct sales team - that would need roughly 90 reps. So the motion shifts from direct selling to an "
    "enabled installer channel: the direct share falls from 70% to 15% while installer partners scale, and "
    "partner managers replace quota carriers. Manufacturing shifts the same way - the assembly partner "
    "carries the ramp, then in-house lines come live, each paid for a lead time before it can build.", 'B34')
put(HR, 'B46',
    "Market share is on the face of the model. The turbineketel replaces a gas boiler one-for-one, so the "
    "market is the ~400,000 gas boilers replaced in the Netherlands each year - not the ~125,000 heat pumps "
    "sold in 2024. Share is measured against that, plus Belgium and Germany once the expansion market opens.", 'B34')

# ===========================================================================
# 6. PRE-EXISTING DEFECT — bare IFNA() renders as #NAME? in Excel
# ===========================================================================
# IFNA is a post-2007 function: OOXML requires it stored as _xlfn.IFNA.
# Written bare (as Google Sheets exports it, and as the previous build script
# wrote it) Excel and LibreOffice both show #NAME? and every dependent cell
# cascades. IFERROR is in the 2007 baseline and needs no prefix, so swap it.
swapped = 0
for ws in wb:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            t = getattr(v, 'text', v)
            if isinstance(t, str) and 'IFNA(' in t:
                nt = t.replace('IFNA(', 'IFERROR(')
                if hasattr(v, 'text'): v.text = nt
                else: c.value = nt
                swapped += 1
print(f'IFNA -> IFERROR: {swapped} formulas repaired')

wb.save(OUT)
print('saved', OUT)
