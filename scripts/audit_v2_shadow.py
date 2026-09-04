"""
shadow.py -- independent reimplementation of the Tarnoc v2 model in Python,
compared cell by cell against the LibreOffice-recalculated workbook.

The point is to catch formula and reference mistakes: wrong signs, off-by-one
ranges, a row pointing at the wrong row, a stock that does not roll forward.
Any disagreement is a bug in the workbook or in my understanding of it, and
either way it needs looking at.
"""
import openpyxl, math, sys, datetime as dt
from openpyxl.utils import get_column_letter as gl

PATH = sys.argv[1]
CASE = int(sys.argv[2]) if len(sys.argv) > 2 else 2
V = openpyxl.load_workbook(PATH, data_only=True)
AS, RF, CG, PE, OP, FS = (V['Assumptions'], V['Revenue Forecast'], V['COGS'],
                          V['Personnel'], V['OPEX'], V['Financial Statements'])
M0, NM = 5, 60
YEARS = [2026, 2027, 2028, 2029, 2030]
def n(x): return x if isinstance(x, (int, float)) else 0.0

# ---- read the assumptions by label, so a moved row cannot silently break it
LBL = {}
for r in range(1, 200):
    b = AS.cell(r, 2).value
    if isinstance(b, str):
        LBL.setdefault(b.strip(), r)

def live(label):
    """The live (per case) value of a single-value driver."""
    r = LBL[label]
    return AS.cell(r, 6).value

def yearvals(label):
    """The live row of a year table: label row, then Base, Aggressive, Live."""
    r = LBL[label]
    return {y: n(AS.cell(r + 3, 4 + k).value) for k, y in enumerate(YEARS)}

def dat(label):
    v = live(label)
    return v.date() if isinstance(v, dt.datetime) else v

open_cash = n(live('Opening cash at Jan-2026'))
sal_infl  = n(live('Annual salary increase'))
taxr      = n(live('Corporate income tax rate'))
loan_rate = n(live('Interest on the working capital loan'))
sell_from = dat('First month we can sell')
hire_from = dat('Hiring starts from')
freeze_to = dat('Committed 2026 plan holds until')
grant_a   = n(live('Subsidy received'));  grant_d = dat('Subsidy received in')
dso, dpo, dio = (n(live('Days sales outstanding')), n(live('Days payable outstanding')),
                 n(live('Days inventory outstanding')))
p_ttk = n(live('Turbineketel list price')); p_odu = n(live('Outdoor unit, Combi+ only'))
p_cmb = n(live('Combi+ price (ketel plus outdoor unit)'))
i_ttk = n(live('Installation, TTK'));      i_cmb = n(live('Installation, Combi+'))
mix_t = n(live('TTK share of units'));     mix_c = n(live('Combi+ share of units'))
ups_r = n(live('Upsell revenue per unit sold')); ups_c = n(live('Upsell cost per unit sold'))
svc_r = n(live('Service revenue per installed unit per year'))
svc_c = n(live('Service cost per installed unit per year'))
svc_a = n(live('Share of the installed base on a contract'))
mkt   = yearvals('Marketing spend')
cpl   = n(live('Cost per lead'))
cpld  = n(live('Cost per lead improvement per doubling of installed base'))
l2q   = n(live('Lead to qualified')); q2w = n(live('Qualified to won'))
direct = yearvals('Share of units sold direct')
repadd = yearvals('Reps hired per month')
ptradd = yearvals('Installer partners signed per month')
rep_s = n(live('Reps in post at Jan-2026')); quota = n(live('Quota per rep'))
ptr_s = n(live('Installer partners at Jan-2026'))
per_ptr = n(live('Units per partner per month'))
comm = n(live('Installer partner commission, share of boiler price'))
ptr_ord = yearvals('Orders an installer partner brings in per month')
ptr_pm  = n(live('Partners per partner manager'))
pcap = n(live('Assembly partner capacity'))
line1 = dat('In-house line 1 producing from'); line2 = dat('In-house line 2 producing from')
lcap = n(live('Capacity per in-house line')); lcapex = n(live('Capex per in-house line'))
tcapex = n(live('Tooling and automation, one-off with line 1'))
lead_m = int(n(live('Months from paying for a line to it producing')))
ops_line = n(live('Production operators per live line'))
line_run = n(live('Facility and maintenance per live line'))
dep_life = n(live('Depreciation life, straight line'))
ship = n(live('Inbound shipping, Combi+ outdoor unit'))
tiers = []
for lb in ('Tier 1', 'Tier 2', 'Tier 3'):
    r = LBL[lb]
    tiers.append((n(AS.cell(r, 4).value), n(AS.cell(r, 5).value), n(AS.cell(r, 6).value)))
tier_basis = n(AS.cell(LBL['BOM tier basis   1 = this year only,  2 = this year plus next'], 5).value)
u_sc = n(live('Units per supply chain and logistics FTE'))
ib_sup = n(live('Installed units per support agent'))
ib_esc = n(live('Installed units per technical escalation FTE'))
ptr_tr = n(live('New partners per year per installer trainer'))
u_desk = n(live('Units per order desk FTE'))
mkt_fte = n(live('Marketing spend per marketer'))
visits = n(live('Boilers one field engineer can look after'))
rnd_s = n(live('R&D engineers carried into 2027')); mkt_base = n(live('Marketing team floor'))
rndadd = yearvals('R&D engineers hired in the year')
c_rep = n(live('Sales rep')); c_pm = n(live('Partner manager'))
c_comm = n(live('Trainer, order desk, marketing')); c_ops = n(live('Supply chain and production'))
c_tech = n(live('Field service engineer'))
c_sup = n(live('Support and escalation')); c_ga = n(live('Leadership and back office'))
c_rnd = n(live('R&D engineer'))
fac = n(live('Offices and facilities per person')); it = n(live('IT and software per person'))
trav = n(live('Travel per person')); recr = n(live('Recruitment per net new hire'))
enable = n(live('Installer training and demo unit per new partner'))
g_fin = n(live('Finance and legal')); g_oth = n(live('Other general'))
r_dev = n(live('Ongoing development')); r_thr = n(live('Third party product development'))
cinfl = n(live('Annual increase on the costs above'))
eq1_d = dat('First round, money in'); eq1_a = n(live('First round, amount'))
eq2_d = dat('Second round, money in'); eq2_a = n(live('Second round, amount'))
loan_a = n(live('Convertible loan drawn')); loan_d = dat('Loan drawn on')

FROZEN = {}
for key, row in (('p_rnd', 38), ('p_sm', 39), ('p_ga', 40),
                 ('np_rnd', 41), ('np_sm', 42), ('np_ga', 43)):
    FROZEN[key] = [n(OP.cell(row, M0 + i).value) for i in range(NM)]

def xround(x):
    """Excel ROUND: half away from zero, unlike Python's round to even."""
    return math.floor(abs(x) + 0.5 + 1e-9) * (1 if x >= 0 else -1)

def md(i): return dt.date(2026 + i // 12, i % 12 + 1, 1)
def yr(i): return 2026 + i // 12
def edate_back(d, months):
    m = d.month - months; y = d.year + (m - 1) // 12
    return dt.date(y, (m - 1) % 12 + 1, 1)
def vlook(key, col):
    """VLOOKUP with TRUE: the last tier whose threshold does not exceed the key."""
    pick = tiers[0]
    for t in tiers:
        if key >= t[0]: pick = t
    return pick[col]

# ---- the shadow model -----------------------------------------------------
S = {k: [0.0] * NM for k in (
    'spend','ib_open','cpl_eff','leads','sql','demand_f','demand_p','demand','direct','rep_h','rep_hc',
    'dcap','ptr_h','ptr_hc','ccap','scap','pm','pcap','lines','icap','bcap',
    'units','ttk_u','cmb_u','ib_close','r_ttk','r_cmb','r_ups','r_ins','r_svc','r_grant',
    'r_tot','uy','uny','tkey','c_ttk_u','c_odu_u','c_ttk','c_cmb','c_ups','c_ins','c_svc',
    'c_comm','c_tot','hc_rep','hc_pm','hc_tr','hc_desk','hc_mkt','hc_sm','hc_sc','hc_op',
    'hc_sup','hc_esc','hc_rnd','hc_core','hc_ops','hc_pay','hc_tech','hc_tot',
    'pc_rnd','pc_sm','pc_ga','pc_tot','o6','o7','o8','o12','o13','o14','o17','o18','o19',
    'o22','o23','o24','o25','o26','o27','o30','o31','o32','o33','f6','f7','f8','f11','f12',
    'f13','f14','f16','f19','f20','f21','f22','f23','f29','f30','f31','f32','f33','f35',
    'f36','f37','f38','f39','f40','f44','f45','f47','f50','f51','f53','f54','f58','f61',
    'f62','f63','f64')}

# annual unit totals need a two-pass approach because the BOM tier key uses them
for _pass in range(3):
    for i in range(NM):
        d, y = md(i), yr(i)
        S['spend'][i] = mkt[y]
        S['ib_open'][i] = 0.0 if i == 0 else S['ib_close'][i-1]
        ibo = S['ib_open'][i]
        S['cpl_eff'][i] = cpl * (1 - cpld) ** (0 if ibo < 250 else math.log(ibo/250, 2))
        S['leads'][i] = S['spend'][i]/S['cpl_eff'][i] if S['cpl_eff'][i] else 0.0
        S['sql'][i] = S['leads'][i]*l2q
        S['demand_f'][i] = S['sql'][i]*q2w
        S['direct'][i] = direct[y]
        S['rep_h'][i] = 0.0 if d < sell_from else repadd[y]
        S['rep_hc'][i] = (rep_s + S['rep_h'][i]) if i == 0 else S['rep_hc'][i-1] + S['rep_h'][i]
        S['dcap'][i] = S['rep_hc'][i]*quota
        S['ptr_h'][i] = 0.0 if d < sell_from else ptradd[y]
        S['ptr_hc'][i] = (ptr_s + S['ptr_h'][i]) if i == 0 else S['ptr_hc'][i-1] + S['ptr_h'][i]
        S['ccap'][i] = S['ptr_hc'][i]*per_ptr
        S['demand_p'][i] = S['ptr_hc'][i]*ptr_ord[y]
        S['demand'][i] = S['demand_f'][i] + S['demand_p'][i]
        ds = S['direct'][i]
        a = S['dcap'][i]/ds if ds else 1_000_000
        b = S['ccap'][i]/(1-ds) if (1-ds) else 1_000_000
        S['scap'][i] = min(a, b)
        S['pm'][i] = xround(S['ptr_hc'][i]/ptr_pm)
        S['pcap'][i] = pcap
        S['lines'][i] = (1 if d >= line1 else 0) + (1 if d >= line2 else 0)
        S['icap'][i] = S['lines'][i]*lcap
        S['bcap'][i] = S['pcap'][i] + S['icap'][i]
        S['units'][i] = 0.0 if d < sell_from else float(xround(min(
            S['demand'][i], S['scap'][i], S['bcap'][i])))
        S['ttk_u'][i] = S['units'][i]*mix_t
        S['cmb_u'][i] = S['units'][i]*mix_c
        S['ib_close'][i] = ibo + S['units'][i]

ANN = {y: sum(S['units'][k] for k in range(NM) if yr(k) == y) for y in YEARS}
for i in range(NM):
    d, y = md(i), yr(i)
    infl_s = (1+sal_infl)**(y-2026)
    infl_c = (1+cinfl)**(y-2026)
    S['r_ttk'][i] = S['ttk_u'][i]*p_ttk
    S['r_cmb'][i] = S['cmb_u'][i]*p_cmb
    S['r_ups'][i] = S['units'][i]*ups_r
    S['r_ins'][i] = S['ttk_u'][i]*i_ttk + S['cmb_u'][i]*i_cmb
    S['r_svc'][i] = S['ib_open'][i]*svc_r/12
    S['r_grant'][i] = grant_a if d == grant_d else 0.0
    S['r_tot'][i] = sum(S[k][i] for k in ('r_ttk','r_cmb','r_ups','r_ins','r_svc','r_grant'))
    S['uy'][i] = ANN.get(y, 0.0)
    S['uny'][i] = ANN.get(y+1, S['uy'][i])
    S['tkey'][i] = S['uy'][i] + S['uny'][i] if tier_basis == 2 else S['uy'][i]
    S['c_ttk_u'][i] = vlook(S['tkey'][i], 1)
    S['c_odu_u'][i] = vlook(S['tkey'][i], 2)
    S['c_ttk'][i] = S['ttk_u'][i]*S['c_ttk_u'][i]
    S['c_cmb'][i] = S['cmb_u'][i]*(S['c_ttk_u'][i]+S['c_odu_u'][i]+ship)
    S['c_ups'][i] = S['units'][i]*ups_c
    S['c_ins'][i] = S['r_ins'][i]
    S['c_svc'][i] = S['ib_open'][i]*svc_c/12
    S['c_comm'][i] = (S['r_ttk'][i]+S['r_cmb'][i])*(1-S['direct'][i])*comm
    S['c_tot'][i] = sum(S[k][i] for k in ('c_ttk','c_cmb','c_ups','c_ins','c_svc','c_comm'))
    S['hc_rep'][i] = S['rep_hc'][i]; S['hc_pm'][i] = S['pm'][i]
    S['hc_tr'][i] = xround(S['ptr_h'][i]*12/ptr_tr)
    S['hc_desk'][i] = xround(S['units'][i]*12/u_desk)
    S['hc_mkt'][i] = mkt_base + xround(S['spend'][i]*12/mkt_fte)
    S['hc_sm'][i] = sum(S[k][i] for k in ('hc_rep','hc_pm','hc_tr','hc_desk','hc_mkt'))
    S['hc_sc'][i] = xround(S['units'][i]*12/u_sc)
    S['hc_op'][i] = S['lines'][i]*ops_line
    S['hc_sup'][i] = xround(S['ib_close'][i]/ib_sup)
    S['hc_esc'][i] = xround(S['ib_close'][i]/ib_esc)
    S['hc_rnd'][i] = rnd_s + sum(rndadd[k] for k in YEARS if k <= y)
    S['hc_core'][i] = n(PE.cell(17, M0 + i).value)   # typed input on the Personnel tab
    S['hc_tech'][i] = xround(S['ib_close'][i]*svc_a/visits)
    S['hc_ops'][i] = sum(S[k][i] for k in ('hc_sc','hc_op','hc_sup','hc_esc','hc_rnd','hc_core','hc_tech'))
    S['hc_pay'][i] = S['hc_sm'][i] + S['hc_ops'][i]
    S['hc_tot'][i] = S['hc_pay'][i]
    S['pc_rnd'][i] = S['hc_rnd'][i]*c_rnd*infl_s
    S['pc_sm'][i] = (S['hc_rep'][i]*c_rep + S['hc_pm'][i]*c_pm
                     + (S['hc_tr'][i]+S['hc_desk'][i]+S['hc_mkt'][i])*c_comm)*infl_s
    S['pc_ga'][i] = ((S['hc_sc'][i]+S['hc_op'][i])*c_ops
                     + (S['hc_sup'][i]+S['hc_esc'][i])*c_sup
                     + S['hc_core'][i]*c_ga + S['hc_tech'][i]*c_tech)*infl_s
    S['pc_tot'][i] = S['pc_rnd'][i]+S['pc_sm'][i]+S['pc_ga'][i]
    frozen = d <= freeze_to
    S['o6'][i] = FROZEN['p_rnd'][i] if frozen else S['pc_rnd'][i]
    S['o7'][i] = FROZEN['p_sm'][i] if frozen else S['pc_sm'][i]
    S['o8'][i] = FROZEN['p_ga'][i] if frozen else S['pc_ga'][i]
    S['o12'][i] = S['spend'][i]
    S['o13'][i] = S['ptr_h'][i]*enable
    S['o14'][i] = FROZEN['np_sm'][i] if frozen else S['o12'][i]+S['o13'][i]
    S['o17'][i] = r_dev*infl_c
    S['o18'][i] = r_thr*infl_c
    S['o19'][i] = FROZEN['np_rnd'][i] if frozen else S['o17'][i]+S['o18'][i]
    S['o22'][i] = S['hc_pay'][i]*(fac+it+trav)
    S['o23'][i] = 0.0 if i == 0 else max(0.0, S['hc_pay'][i]-S['hc_pay'][i-1])*recr
    S['o24'][i] = S['lines'][i]*line_run
    S['o25'][i] = g_fin*infl_c
    S['o26'][i] = g_oth*infl_c
    S['o27'][i] = FROZEN['np_ga'][i] if frozen else sum(S[k][i] for k in ('o22','o23','o24','o25','o26'))
    S['o30'][i] = S['o6'][i]+S['o19'][i]
    S['o31'][i] = S['o7'][i]+S['o14'][i]
    S['o32'][i] = S['o8'][i]+S['o27'][i]
    S['o33'][i] = S['o30'][i]+S['o31'][i]+S['o32'][i]
    S['f6'][i] = S['r_tot'][i]; S['f7'][i] = -S['c_tot'][i]
    S['f8'][i] = S['f6'][i]+S['f7'][i]
    S['f11'][i] = -S['o30'][i]; S['f12'][i] = -S['o31'][i]; S['f13'][i] = -S['o32'][i]
    S['f14'][i] = S['f11'][i]+S['f12'][i]+S['f13'][i]
    S['f16'][i] = S['f8'][i]+S['f14'][i]
    S['f61'][i] = ((lcapex+tcapex) if md(i) == edate_back(line1, lead_m) else 0.0) \
                + (lcapex if md(i) == edate_back(line2, lead_m) else 0.0)
    S['f63'][i] = S['f61'][i] if i == 0 else S['f63'][i-1]+S['f61'][i]
    S['f19'][i] = 0.0 if i == 0 else -S['f63'][i-1]/(dep_life*12)
    S['f62'][i] = -S['f19'][i] if i == 0 else S['f62'][i-1]-S['f19'][i]
    S['f36'][i] = loan_a if md(i) == loan_d else 0.0
    S['f51'][i] = S['f36'][i] if i == 0 else S['f51'][i-1]+S['f36'][i]
    S['f20'][i] = 0.0 if i == 0 else -S['f51'][i-1]*loan_rate/12
    S['f21'][i] = S['f16'][i]+S['f19'][i]+S['f20'][i]
    prev_tlcf = 0.0 if i == 0 else S['f64'][i-1]
    S['f22'][i] = -(max(0.0, S['f21'][i]-prev_tlcf)*taxr) if S['f21'][i] > 0 else 0.0
    S['f23'][i] = S['f21'][i]+S['f22'][i]
    S['f64'][i] = max(0.0, prev_tlcf-max(0.0, S['f21'][i]))+max(0.0, -S['f21'][i])
    S['f44'][i] = dso/30*S['f6'][i]
    S['f45'][i] = dio/30*(-S['f7'][i])
    S['f50'][i] = dpo/30*(-S['f7'][i])
    pa = 0.0 if i == 0 else S['f44'][i-1]
    pi = 0.0 if i == 0 else S['f45'][i-1]
    pp = 0.0 if i == 0 else S['f50'][i-1]
    S['f29'][i] = -(S['f44'][i]-pa); S['f30'][i] = -(S['f45'][i]-pi)
    S['f31'][i] = S['f50'][i]-pp
    S['f32'][i] = S['f23'][i] + (-S['f19'][i]) + S['f29'][i]+S['f30'][i]+S['f31'][i]
    S['f33'][i] = -S['f61'][i]
    S['f35'][i] = (eq1_a if md(i) == eq1_d else 0.0)+(eq2_a if md(i) == eq2_d else 0.0)
    S['f37'][i] = S['f35'][i]+S['f36'][i]
    S['f38'][i] = S['f32'][i]+S['f33'][i]+S['f37'][i]
    S['f39'][i] = open_cash if i == 0 else S['f40'][i-1]
    S['f40'][i] = S['f39'][i]+S['f38'][i]
    S['f47'][i] = S['f63'][i]-S['f62'][i]
    S['f53'][i] = (open_cash+S['f35'][i]) if i == 0 else S['f53'][i-1]+S['f35'][i]
    S['f54'][i] = S['f23'][i] if i == 0 else S['f54'][i-1]+S['f23'][i]

# ---- compare -------------------------------------------------------------
CHECKS = [
 ('RF spend',RF,6,'spend'),('RF ib open',RF,7,'ib_open'),('RF cpl',RF,8,'cpl_eff'),
 ('RF qualified',RF,9,'sql'),('RF funnel orders',RF,10,'demand_f'),
 ('RF partner orders',RF,11,'demand_p'),('RF demand',RF,12,'demand'),
 ('RF direct share',RF,15,'direct'),('RF reps hired',RF,16,'rep_h'),
 ('RF reps in post',RF,17,'rep_hc'),('RF direct cap',RF,18,'dcap'),
 ('RF ptr signed',RF,19,'ptr_h'),('RF ptr on books',RF,20,'ptr_hc'),
 ('RF channel cap',RF,21,'ccap'),('RF selling cap',RF,22,'scap'),
 ('RF partner mgrs',RF,23,'pm'),('RF lines',RF,29,'lines'),('RF build cap',RF,31,'bcap'),
 ('RF units',RF,34,'units'),('RF ttk units',RF,39,'ttk_u'),('RF cmb units',RF,40,'cmb_u'),
 ('RF ib close',RF,41,'ib_close'),('RF rev ttk',RF,44,'r_ttk'),('RF rev cmb',RF,45,'r_cmb'),
 ('RF rev upsell',RF,46,'r_ups'),('RF rev install',RF,47,'r_ins'),
 ('RF rev service',RF,48,'r_svc'),('RF rev grant',RF,49,'r_grant'),
 ('RF rev total',RF,50,'r_tot'),
 ('CG units yr',CG,6,'uy'),('CG units next yr',CG,7,'uny'),('CG tier key',CG,8,'tkey'),
 ('CG ttk unit cost',CG,9,'c_ttk_u'),('CG odu unit cost',CG,10,'c_odu_u'),
 ('CG ttk',CG,13,'c_ttk'),('CG cmb',CG,14,'c_cmb'),('CG upsell',CG,15,'c_ups'),
 ('CG install',CG,16,'c_ins'),('CG service',CG,17,'c_svc'),('CG commission',CG,18,'c_comm'),
 ('CG total',CG,19,'c_tot'),
 ('PE reps',PE,6,'hc_rep'),('PE partner mgrs',PE,7,'hc_pm'),('PE trainers',PE,8,'hc_tr'),
 ('PE order desk',PE,9,'hc_desk'),('PE marketing',PE,10,'hc_mkt'),('PE S&M hc',PE,11,'hc_sm'),
 ('PE supply chain',PE,12,'hc_sc'),('PE operators',PE,13,'hc_op'),
 ('PE support',PE,14,'hc_sup'),('PE escalation',PE,15,'hc_esc'),
 ('PE R&D',PE,16,'hc_rnd'),('PE core',PE,17,'hc_core'),
 ('PE techs',PE,18,'hc_tech'),('PE ops hc',PE,19,'hc_ops'),('PE total hc',PE,20,'hc_tot'),('PE cost R&D',PE,26,'pc_rnd'),('PE cost S&M',PE,27,'pc_sm'),
 ('PE cost G&A',PE,28,'pc_ga'),('PE cost total',PE,29,'pc_tot'),
 ('OP people R&D',OP,6,'o6'),('OP people S&M',OP,7,'o7'),('OP people G&A',OP,8,'o8'),
 ('OP marketing',OP,12,'o12'),('OP enablement',OP,13,'o13'),('OP S&M other',OP,14,'o14'),
 ('OP dev',OP,17,'o17'),('OP third party',OP,18,'o18'),('OP R&D other',OP,19,'o19'),
 ('OP facilities',OP,22,'o22'),('OP recruitment',OP,23,'o23'),('OP plant',OP,24,'o24'),
 ('OP fin legal',OP,25,'o25'),('OP other gen',OP,26,'o26'),('OP G&A other',OP,27,'o27'),
 ('OP total R&D',OP,30,'o30'),('OP total S&M',OP,31,'o31'),('OP total G&A',OP,32,'o32'),
 ('OP total',OP,33,'o33'),
 ('FS revenue',FS,6,'f6'),('FS cogs',FS,7,'f7'),('FS gross profit',FS,8,'f8'),
 ('FS opex R&D',FS,11,'f11'),('FS opex S&M',FS,12,'f12'),('FS opex G&A',FS,13,'f13'),
 ('FS total opex',FS,14,'f14'),('FS EBITDA',FS,16,'f16'),('FS depreciation',FS,19,'f19'),
 ('FS interest',FS,20,'f20'),('FS PBT',FS,21,'f21'),('FS tax',FS,22,'f22'),
 ('FS net income',FS,23,'f23'),('FS d AR',FS,29,'f29'),('FS d inventory',FS,30,'f30'),
 ('FS d AP',FS,31,'f31'),('FS CF ops',FS,32,'f32'),('FS capex CF',FS,33,'f33'),
 ('FS equity',FS,35,'f35'),('FS loan',FS,36,'f36'),('FS CF fin',FS,37,'f37'),
 ('FS d cash',FS,38,'f38'),('FS cash open',FS,39,'f39'),('FS cash close',FS,40,'f40'),
 ('FS AR',FS,44,'f44'),('FS inventory',FS,45,'f45'),('FS net PPE',FS,47,'f47'),
 ('FS AP',FS,50,'f50'),('FS loan bal',FS,51,'f51'),('FS share cap',FS,53,'f53'),
 ('FS retained',FS,54,'f54'),('FS capex',FS,61,'f61'),('FS accum dep',FS,62,'f62'),
 ('FS gross PPE',FS,63,'f63'),('FS tax losses',FS,64,'f64'),
]
TOL = 0.02
bad = 0
print(f'shadow model vs workbook, case {CASE}, {len(CHECKS)} rows x {NM} months '
      f'= {len(CHECKS)*NM} cells\n')
for name, ws, row, key in CHECKS:
    worst, wi = 0.0, None
    for i in range(NM):
        x = n(ws.cell(row, M0+i).value); y = S[key][i]
        diff = abs(x-y)
        rel = diff/max(1.0, abs(y))
        if diff > TOL and rel > 1e-6 and diff > worst:
            worst, wi = diff, i
    if wi is not None:
        bad += 1
        print(f'  MISMATCH  {name:<22} row {row:<3} worst {worst:>14,.2f} at '
              f'{md(wi).strftime("%b-%Y")}  workbook {n(ws.cell(row,M0+wi).value):>16,.2f}  '
              f'shadow {S[key][wi]:>16,.2f}')
print(f'\n{len(CHECKS)-bad} of {len(CHECKS)} rows agree.  {bad} mismatched.')
