"""
expand_tarnoc_personnel.py
==========================
Take the Growth Engine's *required* headcount and materialise it as real rows in
the existing Personnel tab — same columns, same formula pattern, same summary
block — so cost flows to OPEX and the P&L through the wiring that is already
there. Nothing parallel, nothing bypassed.

Why a second pass: the required headcount is computed by the model itself, so we
recalculate it (LibreOffice, both scenario cases) and read the answer rather than
re-deriving it in Python and risking a different number.

    python3 scripts/expand_tarnoc_personnel.py <in.xlsx> <out.xlsx>

What it does
  1. recalculates the model for Case 1 (Base) and Case 2 (Aggressive)
  2. reads required headcount per function per month, Jan-2027..Dec-2029
  3. works out which of those people the existing roster already covers
  4. adds one row per missing person, with a CASE-DEPENDENT start date so the
     same roster serves both scenarios (the pattern already used on rows 28-55)
  5. shifts the how-to text and the two summary blocks down, and widens every
     SUMIF / COUNTIFS range to cover the new rows
  6. repoints OPEX and the Growth Engine at the moved summary rows
  7. removes the temporary OPEX personnel-uplift lines, because Personnel now
     carries that cost
"""
import copy, os, re, shutil, subprocess, sys, tempfile
import datetime as dt
import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.styles import Font

SOFFICE = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
SRC, OUT = sys.argv[1], sys.argv[2]

PE_M0, PE_MONTHS = 33, 36          # Personnel AG = Jan-2027 .. BP = Dec-2029
GE_M0 = 6                          # Growth Engine F = Jan-2027
LAST_ROSTER_ROW = 56
NEVER = dt.date(2035, 1, 1)        # a start date outside the model horizon

# function -> (Growth Engine row, dept, role label, EUR/month, roster roles it replaces)
FUNCS = [
    ('reps',      57, 'S&M', 'Sales',                     7500, ['Sales']),
    ('pm',        58, 'S&M', 'Partner Manager',           8000, ['Partner Manager', 'Account Manager']),
    ('trainer',   59, 'S&M', 'Installer Trainer',         7000, []),
    ('orderdesk', 60, 'S&M', 'Order Desk',                7000, []),
    ('mktg',      61, 'S&M', 'Marketing',                 7000, ['Marketing', 'Growth']),
    ('supply',    63, 'G&A', 'Supply Chain & Logistics',  5000, ['Logistics', 'Operations']),
    ('ops',       64, 'G&A', 'Production Operator',       5000, []),
    ('support',   65, 'G&A', 'Customer Support',          4800, []),
    ('esc',       66, 'G&A', 'Technical Escalation',      4800, []),
    ('qa',        67, 'G&A', 'Quality & Certification',   7000, []),
    ('ga',        69, 'G&A', 'Finance / HR / IT / Legal', 7000, ['CFO', 'CEO', 'COO',
                                                                 'Office Support', 'Finance']),
    ('rnd',       68, 'R&D', 'R&D Engineer',              8000, ['R&D Engineer', 'Field Engineer',
                             'Firmware Engineer', 'Mechanical Engineer', 'Electrical Engineer',
                             'CTO', 'Software Engineer', 'Maintenance Engineer', 'Product']),
]


def recalc_case(case, tmp):
    """Write the model with D82=case, recalculate it, return the value workbook."""
    wb = openpyxl.load_workbook(SRC)
    wb['Assumptions']['D82'] = case
    p = os.path.join(tmp, f'case{case}.xlsx'); wb.save(p)
    out = os.path.join(tmp, f'rc{case}'); os.makedirs(out, exist_ok=True)
    subprocess.run([SOFFICE, '--headless', '--norestore', '--convert-to', 'xlsx',
                    '--outdir', out, p], check=True, capture_output=True)
    return openpyxl.load_workbook(os.path.join(out, f'case{case}.xlsx'), data_only=True)


def month_of(i):
    return dt.date(2027 + i // 12, i % 12 + 1, 1)


def roster_starts(pe):
    """Existing roster: [(role, dept, base_start, aggr_start, end)] for rows 3..56."""
    out = []
    for r in range(3, LAST_ROSTER_ROW + 1):
        role = pe.cell(r, 4).value
        dept = pe.cell(r, 3).value
        if not role or not dept:
            continue
        f = pe.cell(r, 6).value
        if isinstance(f, str) and f.startswith('='):
            ds = re.findall(r'DATE\((\d+),\s*(\d+),\s*(\d+)\)', f)
            aggr, base = dt.date(*map(int, ds[0])), dt.date(*map(int, ds[1]))
        elif isinstance(f, dt.datetime):
            aggr = base = f.date()
        else:
            continue
        g = pe.cell(r, 7).value
        end = g.date() if isinstance(g, dt.datetime) else None
        out.append((role, dept, base, aggr, end))
    return out


def main():
    tmp = tempfile.mkdtemp(prefix='pers_')
    try:
        vals = {c: recalc_case(c, tmp) for c in (1, 2)}
        book = openpyxl.load_workbook(SRC)
        # guard: this script is not idempotent — running it twice would add the
        # same people again. Row 57 is blank in the source, filled once we run.
        if book['Personnel'].cell(57, 2).value is not None:
            sys.exit('Personnel row 57 is already populated — this model has been '
                     'expanded. Re-run scripts/build_tarnoc_growth_engine.py first.')
        roster = roster_starts(book['Personnel'])

        # ---- required vs already-covered, per function, per month, per case ----
        plan = {}                                  # key -> {case: [hires per month]}
        for key, ge_row, dept, role, cost, covers in FUNCS:
            plan[key] = {}
            for case in (1, 2):
                ge = vals[case]['Growth Engine']
                need = []
                run = 0
                for i in range(PE_MONTHS):
                    v = ge.cell(ge_row, GE_M0 + i).value
                    req = float(v) if isinstance(v, (int, float)) else 0.0
                    when = month_of(i)
                    have = sum(1 for rl, dp, b, a, e in roster
                               if rl in covers
                               and (b if case == 1 else a) <= when
                               and (e is None or e > when))
                    run = max(run, max(0.0, req - have))   # never un-hire
                    need.append(run)
                plan[key][case] = need

        # ---- turn the curves into individual rows -------------------------
        new_rows = []                              # (dept, role, cost, base_start, aggr_start)
        for key, ge_row, dept, role, cost, covers in FUNCS:
            n = int(round(max(plan[key][1][-1], plan[key][2][-1])))
            for j in range(1, n + 1):
                starts = {}
                for case in (1, 2):
                    hit = next((i for i, v in enumerate(plan[key][case]) if v >= j - 0.001), None)
                    starts[case] = month_of(hit) if hit is not None else NEVER
                new_rows.append((dept, role, cost, starts[1], starts[2]))
        N = len(new_rows)
        print(f'people to add to Personnel: {N}')
        for key, *_ in FUNCS:
            n = int(round(max(plan[key][1][-1], plan[key][2][-1])))
            if n:
                print(f'   {key:<10} {n:>4}   (base {plan[key][1][-1]:>5.0f}, '
                      f'aggressive {plan[key][2][-1]:>5.0f} by Dec-2029)')
        if N == 0:
            shutil.copy(SRC, OUT); return

        PE = book['Personnel']
        SHIFT_FROM, MAXR = 58, PE.max_row

        # ---- capture and clear everything below the roster -----------------
        keep = {}
        for r in range(SHIFT_FROM, MAXR + 1):
            for c in range(1, PE.max_column + 1):
                cell = PE.cell(r, c)
                if cell.value is not None or cell.has_style:
                    keep[(r, c)] = (cell.value, copy.copy(cell._style))
        merges = [str(m) for m in PE.merged_cells.ranges]
        for m in merges:
            PE.unmerge_cells(m)
        for (r, c) in list(keep):
            PE.cell(r, c).value = None

        # ---- write the new hire rows, cloning row 21's construction --------
        proto = 21
        seq = max((PE.cell(r, 1).value or 0) for r in range(3, LAST_ROSTER_ROW + 1)
                  if isinstance(PE.cell(r, 1).value, int))
        for k, (dept, role, cost, sb, sa) in enumerate(new_rows):
            r = LAST_ROSTER_ROW + 1 + k
            seq += 1
            vals_ab = [seq, 'New hire', dept, role, 'Planned']
            for c, v in enumerate(vals_ab, start=1):
                PE.cell(r, c, v)._style = copy.copy(PE.cell(proto, c)._style)
            f = PE.cell(r, 6)
            f.value = (f'=IF(Assumptions!$D$82=2,DATE({sa.year},{sa.month},{sa.day}),'
                       f'DATE({sb.year},{sb.month},{sb.day}))')
            f._style = copy.copy(PE.cell(proto, 6)._style)
            PE.cell(r, 7)._style = copy.copy(PE.cell(proto, 7)._style)
            h = PE.cell(r, 8, cost); h._style = copy.copy(PE.cell(proto, 8)._style)
            for c in range(9, 69):                       # I..BP monthly
                src = PE.cell(proto, c).value
                cell = PE.cell(r, c)
                cell.value = src.replace(str(proto), str(r)) if isinstance(src, str) else src
                cell._style = copy.copy(PE.cell(proto, c)._style)

        # ---- put the text and summary blocks back, N rows lower -----------
        for (r, c), (v, st) in keep.items():
            cell = PE.cell(r + N, c)
            cell.value = v; cell._style = st
        for m in merges:
            a, b = m.split(':')
            ac, ar = re.match(r'([A-Z]+)(\d+)', a).groups()
            bc, br = re.match(r'([A-Z]+)(\d+)', b).groups()
            if int(ar) >= SHIFT_FROM:
                PE.merge_cells(f'{ac}{int(ar)+N}:{bc}{int(br)+N}')
            else:
                PE.merge_cells(m)

        # ---- widen the summary ranges and fix their internal row refs ------
        NEW_LAST = LAST_ROSTER_ROW + N
        for r0 in (62, 63, 64, 65, 68, 69, 70, 71):
            r = r0 + N
            for c in range(9, 69):
                v = PE.cell(r, c).value
                if not isinstance(v, str):
                    continue
                v = re.sub(r'\$C\$(\d+):\$C\$56', lambda m: f'$C${m.group(1)}:$C${NEW_LAST}', v)
                v = re.sub(r'([A-Z]{1,2})(\d+):([A-Z]{1,2})56',
                           lambda m: f'{m.group(1)}{m.group(2)}:{m.group(3)}{NEW_LAST}', v)
                v = re.sub(r'SUM\(([A-Z]{1,2})62:([A-Z]{1,2})64\)',
                           lambda m: f'SUM({m.group(1)}{62+N}:{m.group(2)}{64+N})', v)
                v = re.sub(r'SUM\(([A-Z]{1,2})68:([A-Z]{1,2})70\)',
                           lambda m: f'SUM({m.group(1)}{68+N}:{m.group(2)}{70+N})', v)
                PE.cell(r, c).value = v

        # ---- repoint everything that referenced the summary rows ----------
        MOVED = {62: 62 + N, 63: 63 + N, 64: 64 + N, 65: 65 + N,
                 68: 68 + N, 69: 69 + N, 70: 70 + N, 71: 71 + N}
        # must handle RANGES: 'Personnel!BE62:BP62' has only one 'Personnel!' prefix,
        # so repointing the first endpoint alone silently widens the range.
        pat = re.compile(r"(Personnel!)(\$?[A-Z]{1,2}\$?)(\d+)"
                         r"(?::(\$?[A-Z]{1,2}\$?)(\d+))?")
        def repoint(m):
            sheet, c1, r1, c2, r2 = m.groups()
            out = f'{sheet}{c1}{MOVED.get(int(r1), int(r1))}'
            if c2 is not None:
                out += f':{c2}{MOVED.get(int(r2), int(r2))}'
            return out
        for ws in book:
            if ws.title == 'Personnel':
                continue
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    t = getattr(v, 'text', v)
                    if isinstance(t, str) and 'Personnel!' in t:
                        nt = pat.sub(repoint, t)
                        if nt != t:
                            if hasattr(v, 'text'): v.text = nt
                            else: cell.value = nt

        # ---- Personnel now carries the people, so drop the uplift lines ----
        GE, OP = book['Growth Engine'], book['OPEX']
        GEC = [gl(GE_M0 + i) for i in range(48)]
        YCOL2029 = 'BE'
        for c in range(7, 55):
            cl = gl(c)
            if 31 <= c < 55:
                g = GEC[c - 31]
                OP[f'{cl}46'] = f"='Growth Engine'!{g}86"
                OP[f'{cl}47'] = (f"='Growth Engine'!{g}84+'Growth Engine'!{g}85"
                                 f"+'Growth Engine'!{g}87+'Growth Engine'!{g}88")
            else:
                OP[f'{cl}46'] = 0; OP[f'{cl}47'] = 0
            OP[f'{cl}49'] = 0
            OP[f'{cl}48'] = f'={cl}46+{cl}47+{cl}49'
        for cl, y in (('BD', 2025), ('BE', 2026), ('BF', 2027), ('BG', 2028), ('BH', 2029)):
            if y in (2027, 2028):
                a, b = ('AE', 'AP') if y == 2027 else ('AQ', 'BB')
                for rr in (46, 47, 49): OP[f'{cl}{rr}'] = f'=SUM({a}{rr}:{b}{rr})'
            elif y == 2029:
                OP[f'{cl}46'] = f"='Growth Engine'!{YCOL2029}86"
                OP[f'{cl}47'] = (f"='Growth Engine'!{YCOL2029}84+'Growth Engine'!{YCOL2029}85"
                                 f"+'Growth Engine'!{YCOL2029}87+'Growth Engine'!{YCOL2029}88")
                OP[f'{cl}49'] = 0
            else:
                for rr in (46, 47, 49): OP[f'{cl}{rr}'] = 0
            OP[f'{cl}48'] = f'={cl}46+{cl}47+{cl}49'
        OP['B46'] = 'S&M — installer training & demo units'
        OP['B47'] = 'G&A — facilities, IT, travel, recruitment, plant, new countries'
        OP['B49'] = '(people are in the Personnel tab)'
        OP['B49'].font = Font(name='Arial', size=9, italic=True, color='6B7686')
        OP['BK46'] = 'people are NOT here — they are rows in the Personnel tab, like every other employee'

        # R&D headcount now comes wholly from the roster
        for i in range(48):
            g = GEC[i]
            pc = gl(min(PE_M0 + i, 68))
            GE[f'{g}68'] = f'=Personnel!{pc}{68 + N}'
        # row 90 must no longer include the personnel uplift: Personnel carries it
        for i in range(48):
            g = GEC[i]
            GE[f'{g}90'] = f'={g}84+{g}85+{g}86+{g}87+{g}88'
        for y in ('BC', 'BD', 'BE', 'BF'):
            GE[f'{y}90'] = f'={y}84+{y}85+{y}86+{y}87+{y}88'
        GE['B90'] = 'TOTAL EXTRA NON-PEOPLE COST'
        GE['BH90'] = 'people are costed in the Personnel tab; only these overheads are added here'
        GE['B80'] = 'S&M people — required vs roster (check only)'
        GE['B83'] = 'Ops & G&A people — required vs roster (check only)'
        GE['B89'] = 'R&D engineers — now hired in the Personnel tab (check only)'
        book['Diligence']['B46'] = 'Extra non-people cost (overheads, training, plant)'
        GE['B68'] = 'R&D — from the Personnel roster'
        GE['BH75'] = ('should now be at or near zero: the people are in the Personnel tab. '
                      'Anything left is a genuine gap between the plan and the roster.')
        GE['BH78'] = 'required vs roster — a check on the Personnel tab, not a cost'
        for r in (78, 79, 80, 81, 82, 83):
            GE[f'C{r}'] = 'check'

        book.save(OUT)
        print(f'saved {OUT}: Personnel rows 57..{NEW_LAST}, summaries now at '
              f'{62+N}-{65+N} and {68+N}-{71+N}')
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == '__main__':
    main()
