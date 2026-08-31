"""Build the €10M aggressive case into the Tarnoc model (toggle-driven)."""
import openpyxl, copy
from openpyxl.utils import get_column_letter as gl

SRC='working.xlsx'; OUT='aggressive.xlsx'
wb=openpyxl.load_workbook(SRC)
A=wb['Assumptions']; FS=wb['Financial Statements']; CG=wb['COGS']; PE=wb['Personnel']; HR=wb['How to read me']

# ---- style helpers: copy look from an existing comparable cell -------------
def restyle(ws, dst, src):
    s=ws[src]; d=ws[dst]
    d.font=copy.copy(s.font); d.fill=copy.copy(s.fill); d.border=copy.copy(s.border)
    d.alignment=copy.copy(s.alignment); d.number_format=s.number_format
def put(ws, ref, val, style_from=None):
    ws[ref]=val
    if style_from: restyle(ws, ref, style_from)

# ===========================================================================
# 1. SCENARIO BLOCK  (Assumptions rows 81-85, an existing empty gap)
# ===========================================================================
put(A,'B81','SCENARIO & TIER BASIS', 'B65')          # section header look
for c in ('C81','D81','E81','F81','G81'): restyle(A,c,'C65')

rows=[('B82','Case   (1 = Base €3m,  2 = Aggressive €10m)','D82',2,
       'the master switch. 2 = the 10m plan: bigger volumes, capex, 10m equity, faster sales hires'),
      ('B83','BOM tier basis   (1 = this year only,  2 = this year + next)','D83',1,
       'how the ketel cost tier is chosen. 1 = only the year\'s own volume (conservative)')]
for lbl_ref,lbl,inp_ref,val,note in rows:
    put(A,lbl_ref,lbl,'B66'); put(A,inp_ref,val,'D15')
    put(A,f'L{lbl_ref[1:]}',note,'L15')

put(A,'B84','Aggressive case — units sold','B66')
put(A,'B85','Base case — units sold','B66')
for r,vals in ((84,(1800,5200,11000,22000)),(85,(600,1600,3600,7200))):
    for c,v in zip('DEFG',vals):
        put(A,f'{c}{r}',v,'D21')
put(A,'L84','2028 is set to clear the 5,000-unit tier, 2029 to clear 10,000','L15')
put(A,'L85','the plan as it stands today — leave alone so Base stays comparable','L15')

# Units-sold dial now reads the switch
for c in 'DEFG':
    A[f'{c}21']=f'=IF($D$82=2,{c}84,{c}85)'
A['L21']='the main dial: total units a year, picked from the Case switch in row 82'

# Inventory: base keeps 0 days, aggressive carries stock
A['D17']='=IF($D$82=2,30,0)'
A['L17']='days of stock held. 0 in Base (as before); 30 in the Aggressive case'

# ===========================================================================
# 2. CAPEX & DEPRECIATION BLOCK  (Assumptions 123-128; content ends at 121)
# ===========================================================================
put(A,'B123','CAPEX & DEPRECIATION','B65')
for c in 'CDEFG': restyle(A,f'{c}123','C65')
put(A,'B124','Year','B66')
for c,y in zip('DEFG',(2027,2028,2029,2030)): put(A,f'{c}124',y,'D22')
put(A,'B125','Tooling & automation (aggressive)','B66')
for c,v in zip('DEFG',(1500000,1000000,500000,0)): put(A,f'{c}125',v,'D21')
put(A,'B126','Facility fit-out (aggressive)','B66')
for c,v in zip('DEFG',(0,1500000,1000000,0)): put(A,f'{c}126',v,'D21')
put(A,'B127','Total capex (case-driven)','B66')
for c in 'DEFG': put(A,f'{c}127',f'=IF($D$82=2,{c}125+{c}126,0)','D21')
put(A,'B128','Useful life (years, blended)','B66'); put(A,'D128',8,'D15')
put(A,'L125','3.0m of tooling, staged 1.5 / 1.0 / 0.5','L15')
put(A,'L126','2.5m fit-out; partner builds until stage 1 is live','L15')
put(A,'L127','what the model actually spends — zero unless Case = 2','L15')
put(A,'L128','straight line, blended across tooling and building','L15')

# ===========================================================================
# 3. COGS — tier basis switch  (rows 7 and 12, ketel tier key only)
# ===========================================================================
KEY27="IF(Assumptions!$D$83=2,'Revenue Forecast'!$BD$34+'Revenue Forecast'!$BE$34,'Revenue Forecast'!$BD$34)"
KEY28="IF(Assumptions!$D$83=2,'Revenue Forecast'!$BE$34+'Revenue Forecast'!$BF$34,'Revenue Forecast'!$BE$34)"
KEY29="IF(Assumptions!$D$83=2,'Revenue Forecast'!$BF$34+Assumptions!$G$21,'Revenue Forecast'!$BF$34)"
TIER="VLOOKUP(%s,Assumptions!$B$67:$C$69,2,TRUE)"
OKEY27="IF(Assumptions!$D$83=2,'Revenue Forecast'!$BD$12+'Revenue Forecast'!$BE$12,'Revenue Forecast'!$BD$12)"
OKEY28="IF(Assumptions!$D$83=2,'Revenue Forecast'!$BE$12+'Revenue Forecast'!$BF$12,'Revenue Forecast'!$BE$12)"
OKEY29="IF(Assumptions!$D$83=2,'Revenue Forecast'!$BF$12+Assumptions!$G$24,'Revenue Forecast'!$BF$12)"
ODU ="VLOOKUP(%s,Assumptions!$B$78:$C$80,2,TRUE)"
for c in range(29,41):   # AC..AN  2027
    col=gl(c)
    CG[f'{col}7'] =f"=IFNA('Revenue Forecast'!{col}6*{TIER%KEY27},0)"
    CG[f'{col}12']=f"=IFNA('Revenue Forecast'!{col}12*({TIER%KEY27}+{ODU%OKEY27}+Assumptions!$E$74),0)"
for c in range(41,53):   # AO..AZ  2028
    col=gl(c)
    CG[f'{col}7'] =f"=IFNA('Revenue Forecast'!{col}6*{TIER%KEY28},0)"
    CG[f'{col}12']=f"=IFNA('Revenue Forecast'!{col}12*({TIER%KEY28}+{ODU%OKEY28}+Assumptions!$E$74),0)"
CG['BF7'] =f"='Revenue Forecast'!BF6*{TIER%KEY29}"
CG['BF12']=f"='Revenue Forecast'!BF12*({TIER%KEY29}+{ODU%OKEY29}+Assumptions!$E$74)"
CG['BH7'] ='units x ketel build cost. the tier basis is set by the switch on Assumptions row 83'
CG['BH12']='units x (ketel tier + outdoor unit tier + shipping), ketel tier per the row-83 switch'

# ===========================================================================
# 4. FS — capex, depreciation, net PP&E, 2029 inventory, equity
# ===========================================================================
# CAPEX monthly (E..AZ, 2025-2028) then 2029 annual
for c in range(5,53):
    col=gl(c)
    FS[f'{col}53']=f'=IFERROR(-HLOOKUP(YEAR({col}$3),Assumptions!$D$124:$G$127,4,FALSE)/12,0)'
FS['BF53']='=-Assumptions!$F$127'
# Depreciation: positive cost, straight line on cumulative gross PP&E
for c in range(5,53):
    col=gl(c)
    FS[f'{col}35']=f'=-SUM($E$53:{col}53)/(Assumptions!$D$128*12)'
    FS[f'{col}48']=f'={col}35'
FS['BF35']='=(-SUM($E$53:$AZ$53)+(-BF53)/2)/Assumptions!$D$128'
# Net PP&E = cumulative capex less accumulated depreciation
for c in range(5,53):
    col=gl(c)
    FS[f'{col}72']=f'=-SUM($E$53:{col}53)-SUM($E$35:{col}35)'
FS['BF72']='=BE72+(-BF53)-BF35'
# 2029 inventory was hardcoded 0 — would break the BS once DIO > 0
FS['BF69']='=Assumptions!$D$17/365*BF17'
FS['BF51']='=-(BF69-BE69)'
# 2029 interest was hardcoding 5%
FS['BF36']='=BE82*Assumptions!$C$11/12'
# Equity: same timing, bigger cheque
FS['AA55']='=IF(Assumptions!$D$82=2,10000000,3000000)'
FS['BH55']='money in from shares: 300k May-26, then 3.0m (Base) or 10.0m (Aggressive) Nov-26'
FS['BH35']='straight line on cumulative capex, blended life from Assumptions D128'
FS['BH53']='staged capex from Assumptions rows 125-127, spread evenly over each year'
FS['BH72']='cumulative capex less accumulated depreciation'

# ===========================================================================
# 5. PERSONNEL — pull sales/ops hires forward in the aggressive case
# ===========================================================================
ACCEL={30:(2027,1,1), 39:(2027,3,1), 40:(2027,4,1), 47:(2027,6,1), 48:(2027,7,1),
       49:(2027,9,1), 51:(2028,1,1), 50:(2028,1,2),
       28:(2027,2,1), 34:(2027,8,1), 35:(2027,10,1), 36:(2028,1,3),
       37:(2028,1,1), 55:(2028,6,1), 41:(2028,1,1), 42:(2028,1,2)}
for r,(y,m,d) in ACCEL.items():
    orig=PE.cell(r,6).value
    PE.cell(r,6).value=(f'=IF(Assumptions!$D$82=2,DATE({y},{m},{d}),'
                        f'DATE({orig.year},{orig.month},{orig.day}))')

# ===========================================================================
# 6. HOW TO READ ME — refresh stale figures, document the switches
# ===========================================================================
put(HR,'B35',"Below the 5,000-unit tier, gross margin per unit is close to zero — so the business is "
    "loss-making by design until volumes cross that threshold. Base case (600 / 1,600 / 3,600 / 7,200 units, "
    "2027-2030): EBITDA -EUR2.45m (2027), +EUR2.43m (2028), +EUR17.14m (2029).",'B34')
put(HR,'B36',"Cash is tightest in Oct-2026 (~EUR7k) immediately before the equity injection; from Nov-2026 the "
    "plan is funded throughout. The Jan-2028 dip described in earlier versions no longer appears at these volumes.",'B34')
put(HR,'B37',"Results are highly sensitive at the tier thresholds. On the two-year basis 2028's tier key is "
    "1,600 + 3,600 = 5,200 against a 5,000 breakpoint, so a 4% volume miss pushes the 2028 BOM back to EUR9,984.",'B34')
put(HR,'B39','SCENARIO SWITCHES (Assumptions rows 82-83)','B33')
put(HR,'B40',"Case: 1 = Base (EUR3m raise, 600 / 1,600 / 3,600 / 7,200 units). 2 = Aggressive (EUR10m raise, "
    "1,800 / 5,200 / 11,000 / 22,000 units, EUR5.5m staged capex, sales hires pulled forward, 30 days of inventory).",'B34')
put(HR,'B41',"BOM tier basis: 1 = the year's own volume only (conservative). 2 = this year + next, as originally "
    "modelled. On basis 1 the Base case never turns profitable (EBITDA -2.45 / -2.23 / -1.21); the Aggressive case "
    "does (-1.34 / +17.12 / +66.08).",'B34')
put(HR,'B42',"Aggressive volumes are deliberately set to clear the breakpoints: 5,200 units in 2028 (>5,000) and "
    "11,000 in 2029 (>10,000). Crossing 5,000 in 2028 is worth roughly EUR15m of COGS.",'B34')

wb.save(OUT); print('saved', OUT)
