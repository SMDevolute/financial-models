"""
build_tarnoc_v2.py
==================
Tarnoc operating model, built from scratch.

Design follows the house conventions of the existing Tarnoc workbook: Geist 10pt,
white-on-black section bars, pale yellow (#FFF2CC) hardcoded inputs, black
calculated cells and links alike, grey ratio rows, lighter grey balance-sheet checks, parenthesised
negatives, an Arial 9 italic grey notes column on the right, gridlines off and
frozen panes on every time series.

Structural differences from the current workbook, all deliberate:
  * monthly from Jan-2026 to Dec-2030 with no annual-only columns, so the
    fragility around the old annual 2029 column disappears
  * units sold is an output: marketing spend runs a funnel, selling capacity and
    build capacity cap it, and the model says which one bit
  * no market-size or share logic anywhere
  * capacity and capex follow the assembly-partner to in-house handover
  * headcount is driven by the drivers that create the work, inside Personnel

Output: models/Tarnoc_v2_2026-09-01.xlsx
"""
import datetime as dt
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gl

OUT = os.environ.get('OUT', 'models/Tarnoc_v2_2026-09-01.xlsx')

# ---------------------------------------------------------------------------
# house style, lifted from the existing workbook
# ---------------------------------------------------------------------------
FONT = 'Geist'
NOTE_FONT = 'Arial'
INK, WHITE = 'FF000000', 'FFFFFFFF'
GREY, CHECK_GREY, RED = 'FF999999', 'FFCCCCCC', 'FFFF0000'
FILL_BLACK, FILL_SUB, FILL_INPUT = 'FF000000', 'FFF3F3F3', 'FFFFF2CC'
FILL_SUBSEC, FILL_WHITE = 'FFEFEFEF', 'FFFFFFFF'

NUM = r'#,##0;\(#,##0\);\-'
NUM1 = r'#,##0.0;\(#,##0.0\);\-'
NUM2 = r'#,##0.00;\(#,##0.00\);\-'
EUR = r'\€#,##0;"(€"#,##0\);\-'
PCT, PCT1 = '0%', '0.0%'
MONTH_FMT, YEAR_FMT, DATE_FMT, TEXT = 'mmm yyyy', '0', 'mmm yyyy', '@'

def f(bold=False, italic=False, color=INK, size=10, name=FONT):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def fill(rgb):
    return PatternFill('solid', fgColor=rgb)

R, L, C = (Alignment(horizontal='right'), Alignment(horizontal='left'),
           Alignment(horizontal='center'))

# ---------------------------------------------------------------------------
# time grid: 60 months Jan-2026 .. Dec-2030, then five annual columns
# ---------------------------------------------------------------------------
M0, NM = 5, 60                                   # E .. BL
MC = [gl(M0 + i) for i in range(NM)]
YEARS = [2026, 2027, 2028, 2029, 2030]
AC0 = M0 + NM + 1                                # BN
YC = {y: gl(AC0 + k) for k, y in enumerate(YEARS)}
NOTE = gl(AC0 + len(YEARS) + 1)                  # BT
LASTCOL = AC0 + len(YEARS)
YMONTHS = {y: MC[k * 12:(k + 1) * 12] for k, y in enumerate(YEARS)}
def mdate(i):
    return dt.datetime(2026 + i // 12, i % 12 + 1, 1)

wb = openpyxl.Workbook()
wb.remove(wb.active)

def sheet(name, label_w=42, unit_w=13, freeze=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = label_w
    ws.column_dimensions['C'].width = unit_w
    ws.column_dimensions['D'].width = 2
    for cl in MC:
        ws.column_dimensions[cl].width = 10
    ws.column_dimensions[gl(AC0 - 1)].width = 2
    for cl in YC.values():
        ws.column_dimensions[cl].width = 12
    ws.column_dimensions[gl(AC0 + len(YEARS))].width = 2
    ws.column_dimensions[NOTE].width = 62
    if freeze:
        ws.freeze_panes = freeze
    return ws

def title(ws, text, sub=None):
    ws['B1'] = text
    ws['B1'].font = f(bold=True, color=WHITE); ws['B1'].fill = fill(FILL_BLACK)
    ws['B1'].alignment = L
    for c in range(3, LASTCOL + 1):
        ws.cell(1, c).fill = fill(FILL_BLACK)
    if sub:
        ws['B2'] = sub
        ws['B2'].font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)

def datebar(ws, row=3):
    """The month header, white bold on black, with the annual columns beside it."""
    for i, cl in enumerate(MC):
        c = ws[f'{cl}{row}']
        c.value = mdate(i); c.number_format = MONTH_FMT
        c.font = f(color=WHITE); c.fill = fill(FILL_BLACK); c.alignment = R
    for y, cl in YC.items():
        c = ws[f'{cl}{row}']
        c.value = y; c.number_format = YEAR_FMT
        c.font = f(bold=True, color=WHITE); c.fill = fill(FILL_BLACK); c.alignment = R
    for col in (2, 3, 4, AC0 - 1, AC0 + len(YEARS)):
        ws.cell(row, col).fill = fill(FILL_BLACK)
    ws[f'{NOTE}{row}'] = 'what this row does'
    ws[f'{NOTE}{row}'].font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)

def bar(ws, row, text):
    """Section bar: white bold on black, all the way across."""
    ws.cell(row, 2, text).font = f(bold=True, color=WHITE)
    ws.cell(row, 2).alignment = L
    for c in range(2, LASTCOL + 1):
        ws.cell(row, c).fill = fill(FILL_BLACK)
        if c > 2:
            ws.cell(row, c).font = f(bold=True, color=WHITE)

def subbar(ws, row, text):
    """Subsection: bold on light grey."""
    ws.cell(row, 2, text).font = f(bold=True)
    ws.cell(row, 2).alignment = L
    for c in range(2, LASTCOL + 1):
        ws.cell(row, c).fill = fill(FILL_SUBSEC)

KIND_COLOR = {'formula': INK, 'input': INK, 'link': INK,
              'ratio': GREY, 'check': CHECK_GREY}

def line(ws, row, lbl, unit, fn, fmt=NUM, kind='formula', total=False,
         annual='sum', note=None, indent=0, grand=False):
    """One row across the whole grid. fn(col_letter, month_index) -> value."""
    lab_col = WHITE if grand else (GREY if kind == 'ratio' else INK)
    cell = ws.cell(row, 2, ('    ' * indent) + lbl)
    cell.font = f(bold=total or grand, color=lab_col)
    cell.alignment = L
    if grand:
        for c in (2, 3, 4, AC0 - 1, AC0 + len(YEARS)):
            ws.cell(row, c).fill = fill(FILL_BLACK)
    elif total:
        for c in (2, 3):
            ws.cell(row, c).fill = fill(FILL_SUB)
    if unit:
        u = ws.cell(row, 3, unit)
        u.font = f(color=WHITE if grand else GREY, size=9, name=NOTE_FONT); u.alignment = L
    for i, cl in enumerate(MC):
        c = ws[f'{cl}{row}']
        c.value = fn(cl, i)
        c.number_format = fmt
        c.font = f(bold=total or grand, color=WHITE if grand else KIND_COLOR.get(kind, INK))
        c.alignment = R
        if grand:
            c.fill = fill(FILL_BLACK)
        elif kind == 'input':
            c.fill = fill(FILL_INPUT)
        elif total:
            c.fill = fill(FILL_SUB)
    for y, cl in YC.items():
        ms = YMONTHS[y]
        if annual == 'sum':
            v = f'=SUM({ms[0]}{row}:{ms[-1]}{row})'
        elif annual == 'end':
            v = f'={ms[-1]}{row}'
        elif annual == 'avg':
            v = f'=IFERROR(AVERAGE({ms[0]}{row}:{ms[-1]}{row}),0)'
        elif annual == 'min':
            v = f'=MIN({ms[0]}{row}:{ms[-1]}{row})'
        else:
            v = None
        if v is None:
            continue
        c = ws[f'{cl}{row}']
        c.value = v; c.number_format = fmt
        c.font = f(bold=True, color=WHITE if grand else KIND_COLOR.get(kind, INK))
        c.alignment = R
        c.fill = fill(FILL_BLACK) if grand else (fill(FILL_SUB) if total else fill(FILL_WHITE))
    if note:
        n = ws[f'{NOTE}{row}']
        n.value = note
        n.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)
        n.alignment = L
        n.number_format = TEXT
    return row

print('style and grid helpers ready')

# ===========================================================================
# ASSUMPTIONS
# ===========================================================================
AS = wb.create_sheet('Assumptions')
AS.sheet_view.showGridLines = False
for cl, w in (('A', 3), ('B', 54), ('C', 15), ('D', 14), ('E', 14),
              ('F', 14), ('G', 14), ('H', 14), ('I', 2), ('J', 66)):
    AS.column_dimensions[cl].width = w
AS['B1'] = 'Tarnoc B.V.  Assumptions'
AS['B1'].font = f(bold=True, color=WHITE); AS['B1'].fill = fill(FILL_BLACK)
for c in range(3, 11):
    AS.cell(1, c).fill = fill(FILL_BLACK)
AS['B2'] = ('Every driver lives here. Pale yellow cells are inputs, everything else is '
            'calculated. The two switches on rows 5 and 6 drive the whole model.')
AS['B2'].font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)

A, AY = {}, {}
_ar = 4

def a_bar(text):
    global _ar
    AS.cell(_ar, 2, text).font = f(bold=True, color=WHITE)
    for c in range(2, 11):
        AS.cell(_ar, c).fill = fill(FILL_BLACK)
    _ar += 1

def a_head(cols):
    global _ar
    for cc in range(2, 11):
        AS.cell(_ar, cc).fill = fill(FILL_SUBSEC)
    for col, txt in cols:
        c = AS[f'{col}{_ar}']
        c.value = txt; c.font = f(bold=True); c.alignment = C
    _ar += 1

def a_switch(key, lbl, value, note=None):
    """A standalone switch: one input cell, no Base/Aggressive split."""
    global _ar
    AS.cell(_ar, 2, lbl).font = f(bold=True); AS.cell(_ar, 2).alignment = L
    AS.cell(_ar, 3, 'switch').font = f(color=GREY, size=9, name=NOTE_FONT)
    c = AS[f'E{_ar}']
    c.value = value; c.number_format = NUM; c.font = f(bold=True)
    c.fill = fill(FILL_INPUT); c.alignment = R
    if note:
        n = AS[f'J{_ar}']; n.value = note; n.number_format = TEXT
        n.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT); n.alignment = L
    A[key] = _ar; _ar += 1
    return _ar - 1

# the two switches come first so their addresses are fixed and quotable
a_bar('SCENARIO SWITCHES')
CASE_ROW = a_switch('case', 'Case   1 = Base (EUR3m raise),  2 = Aggressive (EUR10m raise)', 1,
                    'the master switch. Every Live column on this tab reads it')
TIER_ROW = a_switch('tier_basis',
                    'BOM tier basis   1 = this year only,  2 = this year plus next', 1,
                    'basis 2 commits next year volume to the supplier, so basis 1 ships as the default')
CASE = f'Assumptions!$E${CASE_ROW}'
TIER = f'Assumptions!$E${TIER_ROW}'

def a_single(key, lbl, unit, base, aggr, fmt=NUM, note=None):
    global _ar
    AS.cell(_ar, 2, lbl).font = f(); AS.cell(_ar, 2).alignment = L
    AS.cell(_ar, 3, unit).font = f(color=GREY, size=9, name=NOTE_FONT)
    for col, val in (('D', base), ('E', aggr)):
        c = AS[f'{col}{_ar}']
        c.value = val; c.number_format = fmt
        c.font = f(); c.fill = fill(FILL_INPUT); c.alignment = R
    lv = AS[f'F{_ar}']
    lv.value = f'=IF({CASE}=2,E{_ar},D{_ar})'
    lv.number_format = fmt; lv.font = f(bold=True); lv.alignment = R
    lv.fill = fill(FILL_SUB)
    if note:
        n = AS[f'J{_ar}']; n.value = note; n.number_format = TEXT
        n.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT); n.alignment = L
    A[key] = _ar; _ar += 1
    return _ar - 1

def a_calc(key, lbl, unit, formula, fmt=NUM, note=None):
    global _ar
    AS.cell(_ar, 2, lbl).font = f(); AS.cell(_ar, 2).alignment = L
    AS.cell(_ar, 3, unit).font = f(color=GREY, size=9, name=NOTE_FONT)
    c = AS[f'F{_ar}']
    c.value = formula; c.number_format = fmt; c.font = f(bold=True); c.alignment = R
    c.fill = fill(FILL_SUB)
    if note:
        n = AS[f'J{_ar}']; n.value = note; n.number_format = TEXT
        n.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT); n.alignment = L
    A[key] = _ar; _ar += 1
    return _ar - 1

def a_yeartable(key, lbl, unit, base, aggr, fmt=NUM, note=None):
    """Years across D..H with a Base row, an Aggressive row and a live row."""
    global _ar
    for cc in range(2, 11):
        AS.cell(_ar, cc).fill = fill(FILL_SUBSEC)
    AS.cell(_ar, 2, lbl).font = f(bold=True); AS.cell(_ar, 2).alignment = L
    AS.cell(_ar, 3, unit).font = f(color=GREY, size=9, name=NOTE_FONT)
    for k, y in enumerate(YEARS):
        c = AS.cell(_ar, 4 + k, y)
        c.number_format = YEAR_FMT; c.font = f(bold=True); c.alignment = R
    yr = _ar
    if note:
        n = AS[f'J{_ar}']; n.value = note; n.number_format = TEXT
        n.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT); n.alignment = L
    _ar += 1
    for tag, vals in (('Base', base), ('Aggressive', aggr)):
        AS.cell(_ar, 2, '    ' + tag).font = f()
        for k, v in enumerate(vals):
            c = AS.cell(_ar, 4 + k, v)
            c.number_format = fmt; c.font = f(); c.fill = fill(FILL_INPUT); c.alignment = R
        _ar += 1
    AS.cell(_ar, 2, '    Live (per case)').font = f(bold=True)
    for k in range(len(YEARS)):
        col = gl(4 + k)
        c = AS[f'{col}{_ar}']
        c.value = f'=IF({CASE}=2,{col}{_ar-1},{col}{_ar-2})'
        c.number_format = fmt; c.font = f(bold=True); c.alignment = R
        c.fill = fill(FILL_SUB)
    AY[key] = (yr, _ar); _ar += 1
    return AY[key]

def LV(k):
    return f'Assumptions!$F${A[k]}'

def SW(k):
    return f'Assumptions!$E${A[k]}'

def YL(k, cl):
    yr, lv = AY[k]
    return f'HLOOKUP(YEAR({cl}$3),Assumptions!$D${yr}:$H${lv},{lv - yr + 1},FALSE)'

# ---- general ---------------------------------------------------------------
a_bar('GENERAL')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('open_cash', 'Opening cash at Jan-2026', 'EUR', 853120, 853120, EUR,
         'the actual balance carried into January 2026')
a_single('sal_infl', 'Annual salary increase', '%', 0.05, 0.05, PCT)
a_single('tax', 'Corporate income tax rate', '%', 0.258, 0.258, PCT,
         'losses are carried forward until profits absorb them')
a_single('loan_rate', 'Interest on the working capital loan', '%', 0.05, 0.05, PCT)

a_single('sell_from', 'First month we can sell', 'date',
         dt.datetime(2027, 1, 1), dt.datetime(2027, 1, 1), DATE_FMT,
         'nil units before this. Product, certification and supply have to be ready first')
a_single('hire_from', 'Hiring starts from', 'date',
         dt.datetime(2026, 11, 1), dt.datetime(2026, 11, 1), DATE_FMT,
         'the month after the raise lands. Nothing changes in the committed 2026 plan before it')
a_single('freeze_to', 'Committed 2026 plan holds until', 'date',
         dt.datetime(2026, 10, 1), dt.datetime(2026, 10, 1), DATE_FMT,
         'operating costs to this month are taken from the existing plan, not from the drivers')
a_single('grant_a', 'Subsidy received', 'EUR', 106782, 106782, EUR)
a_single('grant_d', 'Subsidy received in', 'date',
         dt.datetime(2026, 8, 1), dt.datetime(2026, 8, 1), DATE_FMT)

a_bar('WORKING CAPITAL')
a_single('dso', 'Days sales outstanding', 'days', 20, 20)
a_single('dpo', 'Days payable outstanding', 'days', 45, 45)
a_single('dio', 'Days inventory outstanding', 'days', 0, 30,
         note='the aggressive case carries stock, the base case does not')

a_bar('PRICING  (ex VAT)')
a_single('p_ttk', 'Turbineketel list price', 'EUR/unit', 8525.62, 8525.62, EUR)
a_single('p_odu', 'Outdoor unit, Combi+ only', 'EUR/unit', 5309.92, 5309.92, EUR)
a_calc('p_combi', 'Combi+ price (ketel plus outdoor unit)', 'EUR/unit',
       f'={LV("p_ttk")}+{LV("p_odu")}', EUR)
a_single('i_ttk', 'Installation, TTK', 'EUR/unit', 2800, 2800, EUR,
         'pass-through: the customer is charged what the installer is paid')
a_single('i_combi', 'Installation, Combi+', 'EUR/unit', 3800, 3800, EUR,
         'pass-through, so it adds revenue and an identical cost')
a_single('mix_ttk', 'TTK share of units', '%', 0.20, 0.20, PCT)
a_calc('mix_combi', 'Combi+ share of units', '%', f'=1-{LV("mix_ttk")}', PCT)
print(f'assumptions: switches through pricing, rows 4..{_ar-1}')

def a_row3(key, lbl, v1, v2, v3, fmts=(PCT, EUR, EUR), indent=1):
    """A three-column input row, for tables whose columns are not Base/Aggressive."""
    global _ar
    AS.cell(_ar, 2, ('    ' * indent) + lbl).font = f(); AS.cell(_ar, 2).alignment = L
    for col, val, fm in zip(('D', 'E', 'F'), (v1, v2, v3), fmts):
        c = AS[f'{col}{_ar}']
        c.value = val; c.number_format = fm
        c.font = f(); c.fill = fill(FILL_INPUT); c.alignment = R
    A[key] = _ar; _ar += 1
    return _ar - 1

# ---- upsell basket ---------------------------------------------------------
a_bar('UPSELL BASKET  (attached to every unit sold)')
a_head([('D', 'Attach %'), ('E', 'Price'), ('F', 'Cost')])
u1 = a_row3('u_var50', 'Boiler variant 50', 0.49, 165.29, 80)
u2 = a_row3('u_var100', 'Boiler variant 100', 0.19, 330.58, 150)
u3 = a_row3('u_sound', 'Sound reduction', 0.79, 454.55, 190)
a_calc('ups_rev', 'Upsell revenue per unit sold', 'EUR/unit',
       f'=SUMPRODUCT($D${u1}:$D${u3},$E${u1}:$E${u3})', EUR)
a_calc('ups_cost', 'Upsell cost per unit sold', 'EUR/unit',
       f'=SUMPRODUCT($D${u1}:$D${u3},$F${u1}:$F${u3})', EUR)

# ---- service contracts ----------------------------------------------------
a_bar('SERVICE CONTRACTS  (annual, charged on the installed base)')
a_head([('D', 'Attach %'), ('E', 'Price/yr'), ('F', 'Cost/yr')])
s1 = a_row3('s_ttk_b', 'TTK, basic', 0.34, 60, 30)
s2 = a_row3('s_ttk_p', 'TTK, premium', 0.54, 90, 40)
s3 = a_row3('s_cmb_b', 'Combi+, basic', 0.34, 150, 75)
s4 = a_row3('s_cmb_p', 'Combi+, premium', 0.54, 200, 100)
a_calc('svc_rev', 'Service revenue per installed unit per year', 'EUR/unit',
       f'={LV("mix_ttk")}*SUMPRODUCT($D${s1}:$D${s2},$E${s1}:$E${s2})'
       f'+{LV("mix_combi")}*SUMPRODUCT($D${s3}:$D${s4},$E${s3}:$E${s4})', EUR)
a_calc('svc_cost', 'Service cost per installed unit per year', 'EUR/unit',
       f'={LV("mix_ttk")}*SUMPRODUCT($D${s1}:$D${s2},$F${s1}:$F${s2})'
       f'+{LV("mix_combi")}*SUMPRODUCT($D${s3}:$D${s4},$F${s3}:$F${s4})', EUR,
       'this is the field labour, so service engineers are not charged again in Personnel')
a_calc('svc_attach', 'Share of the installed base on a contract', '%',
       f'=$D${s1}+$D${s2}', PCT)

# ---- demand funnel --------------------------------------------------------
a_bar('DEMAND  (marketing spend runs the funnel)')
a_yeartable('mkt', 'Marketing spend', 'EUR/month',
            [0, 28000, 70000, 120000, 170000],
            [0, 90000, 190000, 265000, 300000], EUR,
            'fills the demand the installer partners do not bring in themselves; spend more, generate more orders')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('cpl', 'Cost per lead', 'EUR', 120, 120, EUR,
         'EUR120 a lead at a 20% close rate is EUR600 of marketing per customer')
a_single('l2q', 'Lead to qualified', '%', 0.50, 0.50, PCT)
a_single('q2w', 'Qualified to won', '%', 0.40, 0.40, PCT,
         '50% then 40% is a 20% lead to sale rate, in line with high-intent HVAC leads')

# ---- selling motion -------------------------------------------------------
a_bar('SELLING CAPACITY  (own reps early, installer channel at scale)')
a_yeartable('direct', 'Share of units sold direct', '%',
            [1.00, 0.80, 0.50, 0.35, 0.30],
            [1.00, 0.80, 0.50, 0.35, 0.30], PCT,
            'direct to consumer first, installers brought in from 2027, half the volume by 2028 and installer-led from 2029')
a_yeartable('rep_add', 'Reps hired per month', 'FTE/month',
            [0.0, 0.25, 0.25, 0.25, 0.25],
            [0.0, 1.00, 0.75, 0.25, 0.25], NUM2,
            'sales hiring starts in the first month we can sell, so 2026 is nil')
a_yeartable('ptr_add', 'Installer partners signed per month', 'partners/month',
            [0.0, 0.5, 1.5, 3.0, 3.8],
            [0.0, 2.0, 4.0, 6.0, 7.0], NUM1,
            'same gate: no partner intros before the first month we can sell; the network takes time to build')
a_yeartable('ptr_orders', 'Orders an installer partner brings in per month', 'units/month',
            [0, 1, 2, 3, 4],
            [0, 1, 2, 3, 4], NUM,
            'customers the installer finds on their own jobs, rising as the brand becomes known; these add to demand')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('rep_start', 'Reps in post at Jan-2026', 'FTE', 1, 1, NUM1)
a_single('quota', 'Quota per rep', 'units/month', 20, 20,
         note='a rep sells full quota from the month they are hired, no ramp and no attainment haircut')
a_single('ptr_start', 'Installer partners at Jan-2026', 'partners', 1, 1, NUM1)
a_single('per_ptr', 'Units per partner per month', 'units/month', 8, 8,
         note='a partner sells full volume from the month they are signed')
a_single('ptr_comm', 'Installer partner commission, share of boiler price', '%', 0.10, 0.10, PCT,
         note='paid on every unit sold through the channel, on the boiler price only, not on installation or upsell')
a_single('ptr_per_pm', 'Partners per partner manager', 'partners', 18, 18)
print(f'assumptions: upsell through selling, rows 4..{_ar-1}')

# ---- build capacity and capex ---------------------------------------------
a_bar('BUILD CAPACITY AND CAPEX  (assembly partner first, in-house takes over)')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('partner_cap', 'Assembly partner capacity', 'units/month', 650, 1000,
         note='what the partner has contracted to build for us')
a_single('line1', 'In-house line 1 producing from', 'date',
         dt.datetime(2035, 1, 1), dt.datetime(2028, 7, 1), DATE_FMT,
         'set well past the horizon to mean never, so the capex lead time cannot pull spend in')
a_single('line2', 'In-house line 2 producing from', 'date',
         dt.datetime(2035, 1, 1), dt.datetime(2029, 7, 1), DATE_FMT)
a_single('line_cap', 'Capacity per in-house line', 'units/month', 1000, 1000)
a_single('line_capex', 'Capex per in-house line', 'EUR', 2500000, 2500000, EUR)
a_single('tool_capex', 'Tooling and automation, one-off with line 1', 'EUR',
         1000000, 1000000, EUR)
a_single('lead_m', 'Months from paying for a line to it producing', 'months', 12, 12,
         note='this lag is why the raise has to land before the volume does')
a_single('ops_per_line', 'Production operators per live line', 'FTE', 35, 35,
         note='assembly, balancing, leak test, run-in and electrical test')
a_single('line_run', 'Facility and maintenance per live line', 'EUR/month',
         90000, 90000, EUR, 'the building and the machines, not the people')
a_single('dep_life', 'Depreciation life, straight line', 'years', 8, 8)

# ---- bill of materials ----------------------------------------------------
a_bar('BILL OF MATERIALS  (unit cost falls as volume crosses each tier)')
a_head([('D', 'From units/yr'), ('E', 'TTK cost'), ('F', 'Outdoor unit')])
t1 = a_row3('bom_t1', 'Tier 1', 0, 9984, 1402.60, (NUM, EUR, EUR))
t2 = a_row3('bom_t2', 'Tier 2', 5000, 7069, 1262.34, (NUM, EUR, EUR))
t3 = a_row3('bom_t3', 'Tier 3', 10000, 4998, 1136.10, (NUM, EUR, EUR))
BOM_T1, BOM_T3 = t1, t3
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('ship_combi', 'Inbound shipping, Combi+ outdoor unit', 'EUR/unit', 100, 100, EUR)

# ---- organisation ---------------------------------------------------------
a_bar('ORGANISATION  (headcount follows whatever creates the work)')
a_yeartable('rnd_add', 'R&D engineers hired in the year', 'FTE',
            [0, 2, 2, 3, 4], [0, 3, 11, 12, 12], NUM,
            'a novel turbine machine plus the Twincycle needs engineers, not a fixed team')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('rnd_start', 'R&D engineers carried into 2027', 'FTE', 10, 10,
         note='the roster at the end of the committed 2026 plan')
a_single('mkt_base', 'Marketing team floor', 'FTE', 2, 2)
a_single('u_per_sc', 'Units per supply chain and logistics FTE', 'units/yr', 900, 900,
         note='buying, planning, inbound quality and warehousing')
a_single('ib_per_sup', 'Installed units per support agent', 'units', 2500, 2500,
         note='a heating appliance peaks in winter and this is a first generation product')
a_single('ib_per_esc', 'Installed units per technical escalation FTE', 'units', 15000, 15000)
a_single('ptr_per_tr', 'New partners per year per installer trainer', 'partners', 40, 40)
a_single('u_per_desk', 'Units per order desk FTE', 'units/yr', 3000, 3000)
a_single('mkt_per_fte', 'Marketing spend per marketer', 'EUR/yr', 3000000, 3000000, EUR)
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('visits_tech', 'Boilers one field engineer can look after', 'boilers', 750, 750,
         note='sizes the field service team: boilers on a service contract divided by this number = engineers to hire. Each boiler gets one visit a year, an engineer does three to four a day')

a_bar('LOADED COST PER PERSON  (employer cost, including taxes)')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('c_rep', 'Sales rep', 'EUR/month', 7500, 7500, EUR)
a_single('c_pm', 'Partner manager', 'EUR/month', 8000, 8000, EUR)
a_single('c_comm', 'Trainer, order desk, marketing', 'EUR/month', 7000, 7000, EUR)
a_single('c_ops', 'Supply chain and production', 'EUR/month', 5000, 5000, EUR)
a_single('c_sup', 'Support and escalation', 'EUR/month', 4800, 4800, EUR)
a_single('c_ga', 'Leadership and back office', 'EUR/month', 7000, 7000, EUR)
a_single('c_rnd', 'R&D engineer', 'EUR/month', 5700, 5700, EUR,
         note='the blended cost of the seven engineers already in post')
a_single('c_tech', 'Field service engineer', 'EUR/month', 6500, 6500, EUR)

a_bar('OVERHEADS AND OTHER OPERATING COSTS')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('fac_fte', 'Offices and facilities per person', 'EUR/month', 700, 700, EUR)
a_single('it_fte', 'IT and software per person', 'EUR/month', 250, 250, EUR)
a_single('trav_fte', 'Travel per person', 'EUR/month', 300, 300, EUR)
a_single('recruit', 'Recruitment per net new hire', 'EUR', 8000, 8000, EUR)
a_single('enable', 'Installer training and demo unit per new partner', 'EUR',
         3500, 3500, EUR)
a_single('g_fin', 'Finance and legal', 'EUR/month', 7000, 7000, EUR)
a_single('g_other', 'Other general', 'EUR/month', 5000, 5000, EUR)
a_single('r_dev', 'Ongoing development', 'EUR/month', 8000, 8000, EUR)
a_single('r_third', 'Third party product development', 'EUR/month', 10000, 10000, EUR)
a_single('cost_infl', 'Annual increase on the costs above', '%', 0.10, 0.10, PCT)

a_bar('FUNDING')
a_head([('D', 'Base'), ('E', 'Aggressive'), ('F', 'Live')])
a_single('eq1_d', 'First round, money in', 'date',
         dt.datetime(2026, 5, 1), dt.datetime(2026, 5, 1), DATE_FMT)
a_single('eq1_a', 'First round, amount', 'EUR', 300000, 300000, EUR)
a_single('eq2_d', 'Second round, money in', 'date',
         dt.datetime(2026, 10, 1), dt.datetime(2026, 10, 1), DATE_FMT)
a_single('eq2_a', 'Second round, amount', 'EUR', 3000000, 10000000, EUR,
         'this is the raise the model is built to justify')
a_single('loan_a', 'Convertible loan drawn', 'EUR', 300000, 300000, EUR)
a_single('loan_d', 'Loan drawn on', 'date',
         dt.datetime(2026, 7, 1), dt.datetime(2026, 7, 1), DATE_FMT)
LAST_A = _ar - 1
print(f'assumptions complete: rows 4..{LAST_A}')

# ===========================================================================
# REVENUE FORECAST  (the unit build lives here, then prices turn it into money)
# ===========================================================================
RF = sheet('Revenue Forecast', label_w=44, freeze='E4')
title(RF, 'Tarnoc B.V.  Revenue Forecast',
      'Units sold is an output: marketing spend generates orders, and selling and build '
      'capacity cap them.')
datebar(RF)
LM = MC[-1]

bar(RF, 5, 'DEMAND')
line(RF, 6, 'Marketing spend', 'EUR/mo', lambda cl, i: '=' + YL('mkt', cl),
     EUR, 'link', note='from the Assumptions year table, per the case switch')
line(RF, 7, 'Installed base at start of month', 'units',
     lambda cl, i: '=0' if i == 0 else f'={MC[i-1]}41', annual='end')
line(RF, 8, 'Cost per lead', 'EUR',
     lambda cl, i: f'={LV("cpl")}', EUR, 'link', annual='avg',
     note='held flat: marketing per unit only falls as installers bring in more of the orders')
line(RF, 9, 'Qualified leads', 'leads',
     lambda cl, i: f'=IFERROR({cl}6/{cl}8,0)*{LV("l2q")}',
     note='marketing spend divided by cost per lead, then the lead to qualified rate')
line(RF, 10, 'Orders the funnel generates', 'units',
     lambda cl, i: f'={cl}9*{LV("q2w")}')
line(RF, 11, 'Orders installer partners bring in', 'units',
     lambda cl, i: f'={cl}20*' + YL('ptr_orders', cl),
     note='partners on the books times the orders each finds on their own jobs')
line(RF, 12, 'Demand', 'units', lambda cl, i: f'={cl}10+{cl}11', total=True)

bar(RF, 14, 'SELLING CAPACITY')
line(RF, 15, 'Share sold direct', '%', lambda cl, i: '=' + YL('direct', cl),
     PCT, 'link', annual='avg',
     note='the rest is sold by trained installer partners')
line(RF, 16, 'Reps hired', 'FTE',
     lambda cl, i: f'=IF({cl}$3<{LV("sell_from")},0,' + YL('rep_add', cl) + ')',
     NUM2, 'link', note='no sales reps before the first month we can sell')
line(RF, 17, 'Reps in post', 'FTE',
     lambda cl, i: (f'={LV("rep_start")}+{cl}16' if i == 0 else f'={MC[i-1]}17+{cl}16'),
     NUM1, annual='end')
line(RF, 18, 'Capacity from our own reps', 'units/mo',
     lambda cl, i: f'={cl}17*{LV("quota")}',
     note='every rep in post carries full quota from the month they are hired')
line(RF, 19, 'Installer partners signed', 'partners',
     lambda cl, i: f'=IF({cl}$3<{LV("sell_from")},0,' + YL('ptr_add', cl) + ')',
     NUM1, 'link', note='no partner intros before the first month we can sell')
line(RF, 20, 'Installer partners on the books', 'partners',
     lambda cl, i: (f'={LV("ptr_start")}+{cl}19' if i == 0 else f'={MC[i-1]}20+{cl}19'),
     NUM1, annual='end')
line(RF, 21, 'Capacity from the installer channel', 'units/mo',
     lambda cl, i: f'={cl}20*{LV("per_ptr")}',
     note='every partner on the books sells full volume from the month they are signed')
line(RF, 22, 'Selling capacity', 'units/mo',
     lambda cl, i: (f'=MIN(IFERROR({cl}18/{cl}15,1000000),'
                    f'IFERROR({cl}21/(1-{cl}15),1000000))'), total=True,
     note='our reps must cover the direct share and the partners the rest, so the mix caps it')
line(RF, 23, 'Partner managers needed', 'FTE',
     lambda cl, i: f'=ROUND({cl}20/{LV("ptr_per_pm")},0)', annual='end')
line(RF, 24, 'Selling capacity used', '%',
     lambda cl, i: f'=IFERROR({cl}34/{cl}22,0)', PCT, 'ratio', annual='avg')

bar(RF, 27, 'BUILD CAPACITY')
line(RF, 28, 'Assembly partner', 'units/mo', lambda cl, i: f'={LV("partner_cap")}',
     kind='link', note='the partner carries the ramp while the in-house lines are built')
line(RF, 29, 'In-house lines producing', 'lines',
     lambda cl, i: f'=IF({cl}$3>={LV("line1")},1,0)+IF({cl}$3>={LV("line2")},1,0)',
     annual='end')
line(RF, 30, 'In-house lines', 'units/mo', lambda cl, i: f'={cl}29*{LV("line_cap")}')
line(RF, 31, 'Build capacity', 'units/mo', lambda cl, i: f'={cl}28+{cl}30', total=True)

bar(RF, 33, 'UNITS SOLD')
line(RF, 34, 'Units sold', 'units',
     lambda cl, i: (f'=IF({cl}$3<{LV("sell_from")},0,'
                    f'ROUND(MIN({cl}12,{cl}22,{cl}31),0))'), grand=True,
     note='nil until the first month we can sell, then the smallest of demand, selling and build capacity')
line(RF, 35, 'Build capacity used', '%',
     lambda cl, i: f'=IFERROR({cl}34/{cl}31,0)', PCT, 'ratio', annual='avg')

bar(RF, 38, 'UNIT SPLIT AND INSTALLED BASE')
line(RF, 39, 'TTK units', 'units', lambda cl, i: f'={cl}34*{LV("mix_ttk")}')
line(RF, 40, 'Combi+ units', 'units', lambda cl, i: f'={cl}34*{LV("mix_combi")}')
line(RF, 41, 'Installed base at end of month', 'units',
     lambda cl, i: f'={cl}7+{cl}34', total=True, annual='end')

bar(RF, 43, 'REVENUE')
line(RF, 44, 'Turbineketel', 'EUR', lambda cl, i: f'={cl}39*{LV("p_ttk")}', EUR)
line(RF, 45, 'Combi+', 'EUR', lambda cl, i: f'={cl}40*{LV("p_combi")}', EUR)
line(RF, 46, 'Upsell basket', 'EUR', lambda cl, i: f'={cl}34*{LV("ups_rev")}', EUR)
line(RF, 47, 'Installation (pass-through)', 'EUR',
     lambda cl, i: f'={cl}39*{LV("i_ttk")}+{cl}40*{LV("i_combi")}', EUR,
     note='revenue equals cost, so it adds no margin')
line(RF, 48, 'Service contracts', 'EUR',
     lambda cl, i: f'={cl}7*{LV("svc_rev")}/12', EUR,
     note='charged on the installed base at the start of the month')
line(RF, 49, 'Subsidies and grants', 'EUR',
     lambda cl, i: f'=IF({cl}$3={LV("grant_d")},{LV("grant_a")},0)', EUR,
     note='the 2026 subsidy already in the committed plan')
line(RF, 50, 'Total revenue', 'EUR', lambda cl, i: f'=SUM({cl}44:{cl}49)', EUR, grand=True)
line(RF, 51, 'Revenue per unit', 'EUR/unit',
     lambda cl, i: f'=IFERROR(({cl}50-{cl}49)/{cl}34,0)', EUR, 'ratio', annual='avg')
print('revenue forecast written')

def pv(i):
    """Previous month column, or None in the first month."""
    return MC[i - 1] if i > 0 else None

def infl(cl, key):
    """A monthly cost inflated from the 2026 base."""
    return f'{LV(key)}*(1+{LV("cost_infl")})^(YEAR({cl}$3)-2026)'

# ===========================================================================
# COGS
# ===========================================================================
CG = sheet('COGS', label_w=44, freeze='E4')
title(CG, 'Tarnoc B.V.  Cost of Goods Sold',
      'Unit cost falls as annual volume crosses each tier. The tier basis switch is on '
      'Assumptions row 6.')
datebar(CG)
RFY = "'Revenue Forecast'"
YRNG = f"{RFY}!$BN$3:$BR$34"          # annual columns, year header down to units sold

bar(CG, 5, 'VOLUME TIER')
line(CG, 6, 'Units sold this calendar year', 'units',
     lambda cl, i: f'=IFERROR(HLOOKUP(YEAR({cl}$3),{YRNG},32,FALSE),0)', annual='end')
line(CG, 7, 'Units sold next calendar year', 'units',
     lambda cl, i: f'=IFERROR(HLOOKUP(YEAR({cl}$3)+1,{YRNG},32,FALSE),{cl}6)', annual='end',
     note='only used when the tier basis switch is set to 2; beyond the horizon, next year is taken as at least this year')
line(CG, 8, 'Tier key', 'units',
     lambda cl, i: f'=IF({TIER}=2,{cl}6+{cl}7,{cl}6)', total=True, annual='end')
line(CG, 9, 'Turbineketel cost per unit', 'EUR/unit',
     lambda cl, i: f'=VLOOKUP({cl}8,Assumptions!$D${BOM_T1}:$E${BOM_T3},2,TRUE)',
     EUR, annual='avg',
     note='EUR9,984 below 5,000 units, EUR7,069 from 5,000, EUR4,998 from 10,000')
line(CG, 10, 'Outdoor unit cost per unit', 'EUR/unit',
     lambda cl, i: f'=VLOOKUP({cl}8,Assumptions!$D${BOM_T1}:$F${BOM_T3},3,TRUE)',
     EUR, annual='avg')

bar(CG, 12, 'COST OF GOODS SOLD')
line(CG, 13, 'Turbineketel', 'EUR', lambda cl, i: f'={RFY}!{cl}39*{cl}9', EUR)
line(CG, 14, 'Combi+', 'EUR',
     lambda cl, i: f'={RFY}!{cl}40*({cl}9+{cl}10+{LV("ship_combi")})', EUR,
     note='ketel plus outdoor unit plus inbound shipping')
line(CG, 15, 'Upsell basket', 'EUR', lambda cl, i: f'={RFY}!{cl}34*{LV("ups_cost")}', EUR)
line(CG, 16, 'Installation (pass-through)', 'EUR', lambda cl, i: f'={RFY}!{cl}47', EUR,
     note='identical to the installation revenue, so it nets to nil')
line(CG, 17, 'Service delivery', 'EUR',
     lambda cl, i: f'={RFY}!{cl}7*{LV("svc_cost")}/12', EUR,
     note='parts, consumables and travel per contract, from the service table; the engineers are on Personnel')
line(CG, 18, 'Installer partner commission', 'EUR',
     lambda cl, i: f'=({RFY}!{cl}44+{RFY}!{cl}45)*(1-{RFY}!{cl}15)*{LV("ptr_comm")}', EUR,
     note='boiler revenue on the channel share of units, times the commission rate')
line(CG, 19, 'Total cost of goods sold', 'EUR',
     lambda cl, i: f'=SUM({cl}13:{cl}18)', EUR, grand=True)
line(CG, 21, 'Gross profit per unit', 'EUR/unit',
     lambda cl, i: f'=IFERROR(({RFY}!{cl}50-{cl}19)/{RFY}!{cl}34,0)', EUR, 'ratio',
     annual='avg')

# ===========================================================================
# PERSONNEL
# ===========================================================================
PE = sheet('Personnel', label_w=44, freeze='E4')
title(PE, 'Tarnoc B.V.  Personnel',
      'Headcount by team. Every number is driven by whatever creates the work: units, '
      'installed base, partners signed or marketing spend. The back office is typed in directly.')
datebar(PE)
yr_rnd, lv_rnd = AY['rnd_add']

bar(PE, 5, 'HEADCOUNT')
line(PE, 6, 'Sales reps', 'FTE', lambda cl, i: f'={RFY}!{cl}17', NUM1, 'link', annual='end')
line(PE, 7, 'Partner managers', 'FTE', lambda cl, i: f'={RFY}!{cl}23', 'link', annual='end')
line(PE, 8, 'Installer trainers', 'FTE',
     lambda cl, i: f'=ROUND({RFY}!{cl}19*12/{LV("ptr_per_tr")},0)', annual='end',
     note='every partner has to be trained and certified before selling anything')
line(PE, 9, 'Order desk', 'FTE',
     lambda cl, i: f'=ROUND({RFY}!{cl}34*12/{LV("u_per_desk")},0)', annual='end')
line(PE, 10, 'Marketing', 'FTE',
     lambda cl, i: f'={LV("mkt_base")}+ROUND({RFY}!{cl}6*12/{LV("mkt_per_fte")},0)',
     annual='end')
line(PE, 11, 'Sales and marketing', 'FTE',
     lambda cl, i: f'=SUM({cl}6:{cl}10)', NUM1, total=True, annual='end')
line(PE, 12, 'Supply chain and logistics', 'FTE',
     lambda cl, i: f'=ROUND({RFY}!{cl}34*12/{LV("u_per_sc")},0)', annual='end',
     note='buying, planning, inbound quality and warehousing')
line(PE, 13, 'Production operators', 'FTE',
     lambda cl, i: f'={RFY}!{cl}29*{LV("ops_per_line")}', annual='end')
line(PE, 14, 'Customer support', 'FTE',
     lambda cl, i: f'=ROUND({RFY}!{cl}41/{LV("ib_per_sup")},0)', annual='end',
     note='inbound calls follow the installed base, not sales')
line(PE, 15, 'Technical escalation', 'FTE',
     lambda cl, i: f'=ROUND({RFY}!{cl}41/{LV("ib_per_esc")},0)', annual='end')
line(PE, 16, 'R&D engineers', 'FTE',
     lambda cl, i: (f'={LV("rnd_start")}+SUMIF(Assumptions!$D${yr_rnd}:$H${yr_rnd},'
                    f'"<="&YEAR({cl}$3),Assumptions!$D${lv_rnd}:$H${lv_rnd})'),
     annual='end', note='the founding engineers plus everyone hired up to and including this year')
BACK_OFFICE = {2026: 3, 2027: 4, 2028: 5, 2029: 6, 2030: 8}
line(PE, 17, 'Leadership, finance, HR, IT and legal', 'FTE',
     lambda cl, i: BACK_OFFICE[YEARS[i // 12]], kind='input', annual='end',
     note='typed here, not driven: the leadership team, then finance, HR, IT and legal as the company grows. Same in both cases')
line(PE, 18, 'Field service engineers', 'FTE',
     lambda cl, i: f'=ROUND({RFY}!{cl}41*{LV("svc_attach")}/{LV("visits_tech")},0)',
     annual='end',
     note='salaried staff: boilers on a service contract divided by the boilers one engineer can look after')
line(PE, 19, 'Operations, support and administration', 'FTE',
     lambda cl, i: f'=SUM({cl}12:{cl}18)', NUM1, total=True, annual='end')
line(PE, 20, 'Total headcount', 'FTE', lambda cl, i: f'={cl}11+{cl}19', NUM1,
     grand=True, annual='end')

bar(PE, 25, 'COST BY DEPARTMENT')
SI = lambda cl: f'(1+{LV("sal_infl")})^(YEAR({cl}$3)-2026)'
line(PE, 26, 'Research and development', 'EUR/mo',
     lambda cl, i: f'={cl}16*{LV("c_rnd")}*{SI(cl)}', EUR)
line(PE, 27, 'Sales and marketing', 'EUR/mo',
     lambda cl, i: (f'=({cl}6*{LV("c_rep")}+{cl}7*{LV("c_pm")}'
                    f'+({cl}8+{cl}9+{cl}10)*{LV("c_comm")})*{SI(cl)}'), EUR)
line(PE, 28, 'General and administrative', 'EUR/mo',
     lambda cl, i: (f'=(({cl}12+{cl}13)*{LV("c_ops")}+({cl}14+{cl}15)*{LV("c_sup")}'
                    f'+{cl}17*{LV("c_ga")}+{cl}18*{LV("c_tech")})*{SI(cl)}'), EUR)
line(PE, 29, 'Total people cost', 'EUR/mo', lambda cl, i: f'=SUM({cl}26:{cl}28)',
     EUR, total=True)
line(PE, 31, 'Average cost per person', 'EUR/mo',
     lambda cl, i: f'=IFERROR({cl}29/{cl}20,0)', EUR, 'ratio', annual='avg')
print('cogs and personnel written')

# ===========================================================================
# OPEX
# ===========================================================================
OP = sheet('OPEX', label_w=44, freeze='E4')
title(OP, 'Tarnoc B.V.  Operating Expenses',
      'People come from the Personnel tab. Everything else scales with headcount, '
      'partners signed or the production lines that are live.')
datebar(OP)

# ---- the committed 2026 plan, held fixed line by line --------------------
FROZEN = {
    38: ('People, research and development',
         [50700, 50700, 50700, 50700, 50700, 50700, 50700, 50700, 54700, 61620]),
    39: ('People, sales and marketing',
         [0, 10800, 10800, 10800, 10800, 10800, 10800, 10800, 10800, 19800]),
    40: ('People, general and administrative',
         [11820, 11820, 11820, 11820, 11820, 11820, 11820, 11820, 25740, 21420]),
    41: ('Everything else, research and development',
         [25000, 25000, 25000, 25000, 25000, 25000, 25000, 25000, 25000, 125000]),
    42: ('Everything else, sales and marketing',
         [21000, 21000, 21000, 21000, 31000, 31000, 21000, 21000, 21000, 24000]),
    43: ('Everything else, general and administrative',
         [20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000]),
}
bar(OP, 37, 'COMMITTED 2026 PLAN, HELD FIXED  (January to October, from the existing model)')
for r, (lbl, vals) in FROZEN.items():
    line(OP, r, lbl, 'EUR', lambda cl, i, v=vals: (v[i] if i < 10 else 0), EUR, 'input',
         note=('taken from the existing model and not recalculated' if r == 38 else None))
OP[f'{NOTE}44'] = ('these ten months are committed. The drivers take over from November, '
                   'the month after the raise lands.')
OP[f'{NOTE}44'].font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)

bar(OP, 5, 'PEOPLE')
line(OP, 6, 'Research and development', 'EUR',
     lambda cl, i: f'=IF({cl}$3<={LV("freeze_to")},{cl}38,Personnel!{cl}26)', EUR, 'link')
line(OP, 7, 'Sales and marketing', 'EUR',
     lambda cl, i: f'=IF({cl}$3<={LV("freeze_to")},{cl}39,Personnel!{cl}27)', EUR, 'link')
line(OP, 8, 'General and administrative', 'EUR',
     lambda cl, i: f'=IF({cl}$3<={LV("freeze_to")},{cl}40,Personnel!{cl}28)', EUR, 'link')
line(OP, 9, 'Total people cost', 'EUR', lambda cl, i: f'=SUM({cl}6:{cl}8)', EUR, total=True)

bar(OP, 11, 'SALES AND MARKETING, EVERYTHING ELSE')
line(OP, 12, 'Performance marketing', 'EUR', lambda cl, i: f'={RFY}!{cl}6', EUR, 'link',
     note='the same spend that drives the funnel, so it can never be double counted')
line(OP, 13, 'Installer training and demo units', 'EUR',
     lambda cl, i: f'={RFY}!{cl}19*{LV("enable")}', EUR)
line(OP, 14, 'Total', 'EUR',
     lambda cl, i: f'=IF({cl}$3<={LV("freeze_to")},{cl}42,SUM({cl}12:{cl}13))', EUR, total=True)

bar(OP, 16, 'RESEARCH AND DEVELOPMENT, EVERYTHING ELSE')
line(OP, 17, 'Ongoing development', 'EUR', lambda cl, i: '=' + infl(cl, 'r_dev'), EUR)
line(OP, 18, 'Third party product development', 'EUR',
     lambda cl, i: '=' + infl(cl, 'r_third'), EUR)
line(OP, 19, 'Total', 'EUR',
     lambda cl, i: f'=IF({cl}$3<={LV("freeze_to")},{cl}41,SUM({cl}17:{cl}18))', EUR, total=True)

bar(OP, 21, 'GENERAL AND ADMINISTRATIVE, EVERYTHING ELSE')
line(OP, 22, 'Offices, IT and travel', 'EUR',
     lambda cl, i: (f'=Personnel!{cl}20*({LV("fac_fte")}+{LV("it_fte")}'
                    f'+{LV("trav_fte")})'), EUR,
     note='scales with the number of people on the payroll')
line(OP, 23, 'Recruitment', 'EUR',
     lambda cl, i: ('=0' if i == 0 else
                    f'=MAX(0,Personnel!{cl}20-Personnel!{pv(i)}20)*{LV("recruit")}'), EUR)
line(OP, 24, 'Production line facility and maintenance', 'EUR',
     lambda cl, i: f'={RFY}!{cl}29*{LV("line_run")}', EUR,
     note='only once a line is actually producing')
line(OP, 25, 'Finance and legal', 'EUR', lambda cl, i: '=' + infl(cl, 'g_fin'), EUR)
line(OP, 26, 'Other general', 'EUR', lambda cl, i: '=' + infl(cl, 'g_other'), EUR)
line(OP, 27, 'Total', 'EUR',
     lambda cl, i: f'=IF({cl}$3<={LV("freeze_to")},{cl}43,SUM({cl}22:{cl}26))', EUR, total=True)

bar(OP, 29, 'TOTAL OPERATING EXPENSES')
line(OP, 30, 'Research and development', 'EUR', lambda cl, i: f'={cl}6+{cl}19', EUR)
line(OP, 31, 'Sales and marketing', 'EUR', lambda cl, i: f'={cl}7+{cl}14', EUR)
line(OP, 32, 'General and administrative', 'EUR', lambda cl, i: f'={cl}8+{cl}27', EUR)
line(OP, 33, 'Total operating expenses', 'EUR', lambda cl, i: f'=SUM({cl}30:{cl}32)',
     EUR, grand=True)
line(OP, 35, 'Operating expenses as a share of revenue', '%',
     lambda cl, i: f'=IFERROR({cl}33/{RFY}!{cl}50,0)', PCT1, 'ratio', annual='avg')

# ===========================================================================
# FINANCIAL STATEMENTS
# ===========================================================================
FS = sheet('Financial Statements', label_w=46, freeze='E4')
title(FS, 'Tarnoc B.V.  Financial Statements',
      'Three statements, monthly. The check on row 58 must read zero in every column.')
datebar(FS)

bar(FS, 5, 'PROFIT AND LOSS')
line(FS, 6, 'Revenue', 'EUR', lambda cl, i: f'={RFY}!{cl}50', EUR, 'link')
line(FS, 7, 'Cost of goods sold', 'EUR', lambda cl, i: f'=-COGS!{cl}19', EUR, 'link')
line(FS, 8, 'Gross profit', 'EUR', lambda cl, i: f'={cl}6+{cl}7', EUR, grand=True)
line(FS, 9, 'Gross margin', '%', lambda cl, i: f'=IFERROR({cl}8/{cl}6,0)', PCT1,
     'ratio', annual='avg')
line(FS, 11, 'Research and development', 'EUR', lambda cl, i: f'=-OPEX!{cl}30', EUR, 'link')
line(FS, 12, 'Sales and marketing', 'EUR', lambda cl, i: f'=-OPEX!{cl}31', EUR, 'link')
line(FS, 13, 'General and administrative', 'EUR', lambda cl, i: f'=-OPEX!{cl}32', EUR, 'link')
line(FS, 14, 'Total operating expenses', 'EUR', lambda cl, i: f'=SUM({cl}11:{cl}13)',
     EUR, total=True)
line(FS, 16, 'EBITDA', 'EUR', lambda cl, i: f'={cl}8+{cl}14', EUR, grand=True)
line(FS, 17, 'EBITDA margin', '%', lambda cl, i: f'=IFERROR({cl}16/{cl}6,0)', PCT1,
     'ratio', annual='avg')
line(FS, 19, 'Depreciation', 'EUR',
     lambda cl, i: ('=0' if i == 0 else
                    f'=-{pv(i)}63/({LV("dep_life")}*12)'), EUR,
     note='straight line on the gross book value carried into the month')
line(FS, 20, 'Interest on the loan', 'EUR',
     lambda cl, i: ('=0' if i == 0 else f'=-{pv(i)}51*{LV("loan_rate")}/12'), EUR)
line(FS, 21, 'Profit before tax', 'EUR', lambda cl, i: f'={cl}16+{cl}19+{cl}20', EUR)
line(FS, 22, 'Tax', 'EUR',
     lambda cl, i: (f'=-IF({cl}21>0,MAX(0,{cl}21-'
                    + ('0' if i == 0 else f'{pv(i)}64') + f')*{LV("tax")},0)'), EUR,
     note='losses are carried forward and absorb profit before any tax is paid')
line(FS, 23, 'Net income', 'EUR', lambda cl, i: f'={cl}21+{cl}22', EUR, grand=True)
line(FS, 24, 'Net margin', '%', lambda cl, i: f'=IFERROR({cl}23/{cl}6,0)', PCT1,
     'ratio', annual='avg')

bar(FS, 26, 'CASH FLOW')
line(FS, 27, 'Net income', 'EUR', lambda cl, i: f'={cl}23', EUR)
line(FS, 28, 'Depreciation added back', 'EUR', lambda cl, i: f'=-{cl}19', EUR)
line(FS, 29, 'Movement in receivables', 'EUR',
     lambda cl, i: f'=-({cl}44-' + ('0' if i == 0 else f'{pv(i)}44') + ')', EUR)
line(FS, 30, 'Movement in inventory', 'EUR',
     lambda cl, i: f'=-({cl}45-' + ('0' if i == 0 else f'{pv(i)}45') + ')', EUR)
line(FS, 31, 'Movement in payables', 'EUR',
     lambda cl, i: f'={cl}50-' + ('0' if i == 0 else f'{pv(i)}50'), EUR)
line(FS, 32, 'Cash from operations', 'EUR', lambda cl, i: f'=SUM({cl}27:{cl}31)',
     EUR, total=True)
line(FS, 33, 'Capital expenditure', 'EUR', lambda cl, i: f'=-{cl}61', EUR)
line(FS, 34, 'Cash from investing', 'EUR', lambda cl, i: f'={cl}33', EUR, total=True)
line(FS, 35, 'Equity raised', 'EUR',
     lambda cl, i: (f'=IF({cl}$3={LV("eq1_d")},{LV("eq1_a")},0)'
                    f'+IF({cl}$3={LV("eq2_d")},{LV("eq2_a")},0)'), EUR,
     note='dates and amounts are on the Assumptions funding block')
line(FS, 36, 'Loan drawn', 'EUR',
     lambda cl, i: f'=IF({cl}$3={LV("loan_d")},{LV("loan_a")},0)', EUR)
line(FS, 37, 'Cash from financing', 'EUR', lambda cl, i: f'={cl}35+{cl}36', EUR, total=True)
line(FS, 38, 'Movement in cash', 'EUR', lambda cl, i: f'={cl}32+{cl}34+{cl}37', EUR)
line(FS, 39, 'Cash at the start of the month', 'EUR',
     lambda cl, i: (f'={LV("open_cash")}' if i == 0 else f'={pv(i)}40'), EUR, annual='end')
line(FS, 40, 'Cash at the end of the month', 'EUR', lambda cl, i: f'={cl}39+{cl}38',
     EUR, total=True, annual='end')

bar(FS, 42, 'BALANCE SHEET')
line(FS, 43, 'Cash', 'EUR', lambda cl, i: f'={cl}40', EUR, annual='end')
line(FS, 44, 'Receivables', 'EUR', lambda cl, i: f'={LV("dso")}/30*{cl}6', EUR, annual='end')
line(FS, 45, 'Inventory', 'EUR', lambda cl, i: f'={LV("dio")}/30*-{cl}7', EUR, annual='end')
line(FS, 46, 'Total current assets', 'EUR', lambda cl, i: f'=SUM({cl}43:{cl}45)', EUR,
     total=True, annual='end')
line(FS, 47, 'Property, plant and equipment, net', 'EUR',
     lambda cl, i: f'={cl}63-{cl}62', EUR, annual='end')
line(FS, 48, 'Total assets', 'EUR', lambda cl, i: f'={cl}46+{cl}47', EUR,
     total=True, annual='end')
line(FS, 50, 'Payables', 'EUR', lambda cl, i: f'={LV("dpo")}/30*-{cl}7', EUR, annual='end')
line(FS, 51, 'Loan outstanding', 'EUR',
     lambda cl, i: (f'={cl}36' if i == 0 else f'={pv(i)}51+{cl}36'), EUR, annual='end')
line(FS, 52, 'Total liabilities', 'EUR', lambda cl, i: f'={cl}50+{cl}51', EUR,
     total=True, annual='end')
line(FS, 53, 'Share capital and premium', 'EUR',
     lambda cl, i: (f'={LV("open_cash")}+{cl}35' if i == 0 else f'={pv(i)}53+{cl}35'),
     EUR, annual='end', note='opening equity matches the opening cash brought forward')
line(FS, 54, 'Retained earnings', 'EUR',
     lambda cl, i: (f'={cl}23' if i == 0 else f'={pv(i)}54+{cl}23'), EUR, annual='end')
line(FS, 55, 'Total equity', 'EUR', lambda cl, i: f'={cl}53+{cl}54', EUR,
     total=True, annual='end')
line(FS, 56, 'Total liabilities and equity', 'EUR', lambda cl, i: f'={cl}52+{cl}55', EUR,
     total=True, annual='end')
line(FS, 58, 'Check   assets less liabilities and equity, must be nil', 'EUR',
     lambda cl, i: f'={cl}48-{cl}56', NUM2, 'check', annual='end')

bar(FS, 60, 'SUPPORTING SCHEDULES')
line(FS, 61, 'Capital expenditure in the month', 'EUR',
     lambda cl, i: (f'=IF(AND(YEAR({cl}$3)=YEAR(EDATE({LV("line1")},-{LV("lead_m")})),'
                    f'MONTH({cl}$3)=MONTH(EDATE({LV("line1")},-{LV("lead_m")}))),'
                    f'{LV("line_capex")}+{LV("tool_capex")},0)'
                    f'+IF(AND(YEAR({cl}$3)=YEAR(EDATE({LV("line2")},-{LV("lead_m")})),'
                    f'MONTH({cl}$3)=MONTH(EDATE({LV("line2")},-{LV("lead_m")}))),'
                    f'{LV("line_capex")},0)'), EUR,
     note='a line is paid for the lead time before it can build anything')
line(FS, 62, 'Accumulated depreciation', 'EUR',
     lambda cl, i: (f'=-{cl}19' if i == 0 else f'={pv(i)}62-{cl}19'), EUR, annual='end')
line(FS, 63, 'Gross book value', 'EUR',
     lambda cl, i: (f'={cl}61' if i == 0 else f'={pv(i)}63+{cl}61'), EUR, annual='end')
line(FS, 64, 'Tax losses carried forward', 'EUR',
     lambda cl, i: (f'=MAX(0,-{cl}21)' if i == 0 else
                    f'=MAX(0,{pv(i)}64-MAX(0,{cl}21))+MAX(0,-{cl}21)'), EUR, annual='end')
print('opex and financial statements written')

# ===========================================================================
# DASHBOARD
# ===========================================================================
DB = wb.create_sheet('Dashboard')
DB.sheet_view.showGridLines = False
for cl, w in (('A', 3), ('B', 50), ('C', 13), ('D', 15), ('E', 15), ('F', 15),
              ('G', 15), ('H', 15), ('I', 2), ('J', 64)):
    DB.column_dimensions[cl].width = w
DB['B1'] = 'Tarnoc B.V.  Dashboard'
DB['B1'].font = f(bold=True, color=WHITE); DB['B1'].fill = fill(FILL_BLACK)
for c in range(3, 11):
    DB.cell(1, c).fill = fill(FILL_BLACK)
DB['B2'] = 'Every figure is a formula off the other tabs. Change a driver on Assumptions and this moves.'
DB['B2'].font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)
DCOL = {y: gl(4 + k) for k, y in enumerate(YEARS)}
for y, cl in DCOL.items():
    c = DB[f'{cl}3']
    c.value = y; c.number_format = YEAR_FMT
    c.font = f(bold=True, color=WHITE); c.fill = fill(FILL_BLACK); c.alignment = R
for cc in (2, 3, 9, 10):
    DB.cell(3, cc).fill = fill(FILL_BLACK)

def d_bar(row, text):
    DB.cell(row, 2, text).font = f(bold=True, color=WHITE)
    for c in range(2, 11):
        DB.cell(row, c).fill = fill(FILL_BLACK)

def d_line(row, lbl, fn, fmt=NUM, kind='formula', total=False, note=None):
    cell = DB.cell(row, 2, lbl); cell.font = f(bold=total); cell.alignment = L
    for y, cl in DCOL.items():
        c = DB[f'{cl}{row}']
        c.value = fn(y); c.number_format = fmt
        c.font = f(bold=total, color=KIND_COLOR.get(kind, INK)); c.alignment = R
        if total:
            c.fill = fill(FILL_SUB)
    if note:
        n = DB[f'J{row}']; n.value = note; n.number_format = TEXT
        n.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT); n.alignment = L

def yc(sh, row, y):
    return f"='{sh}'!{YC[y]}{row}" if ' ' in sh else f'={sh}!{YC[y]}{row}'

d_bar(5, 'VOLUME AND CUSTOMERS')
d_line(6, 'Units sold', lambda y: yc('Revenue Forecast', 34, y), total=True)
d_line(7, 'Installed base at year end', lambda y: yc('Revenue Forecast', 41, y))
d_line(8, 'Build capacity used', lambda y: yc('Revenue Forecast', 35, y), PCT1, 'ratio')
d_line(9, 'Installer partners at year end', lambda y: yc('Revenue Forecast', 20, y), NUM1)
d_line(10, 'Marketing cost per unit sold',
       lambda y: f"=IFERROR('Revenue Forecast'!{YC[y]}6/'Revenue Forecast'!{YC[y]}34,0)", EUR)

d_bar(13, 'PROFIT AND LOSS')
d_line(14, 'Revenue', lambda y: yc('Financial Statements', 6, y), EUR, total=True)
d_line(15, 'Gross profit', lambda y: yc('Financial Statements', 8, y), EUR)
d_line(16, 'Gross margin', lambda y: yc('Financial Statements', 9, y), PCT1, 'ratio')
d_line(17, 'Operating expenses', lambda y: yc('Financial Statements', 14, y), EUR)
d_line(18, 'EBITDA', lambda y: yc('Financial Statements', 16, y), EUR, total=True)
d_line(19, 'EBITDA margin', lambda y: yc('Financial Statements', 17, y), PCT1, 'ratio')
d_line(20, 'Net income', lambda y: yc('Financial Statements', 23, y), EUR)

d_bar(22, 'PEOPLE')
d_line(23, 'Field service engineers', lambda y: yc('Personnel', 18, y))
d_line(24, 'Total headcount', lambda y: yc('Personnel', 20, y), NUM1, total=True)
d_line(26, 'Revenue per person',
       lambda y: f"=IFERROR({DCOL[y]}14/{DCOL[y]}25,0)", EUR,
       note='Viessmann runs at about EUR276k and Vaillant about EUR200k')

d_bar(28, 'CASH AND FUNDING')
d_line(29, 'Equity raised in the year', lambda y: yc('Financial Statements', 35, y), EUR)
d_line(30, 'Capital expenditure', lambda y: yc('Financial Statements', 61, y), EUR)
d_line(31, 'Cash at year end', lambda y: yc('Financial Statements', 40, y), EUR, total=True)
d_line(32, 'Lowest cash during the year',
       lambda y: (f"=MIN('Financial Statements'!{YMONTHS[y][0]}40:"
                  f"'Financial Statements'!{YMONTHS[y][-1]}40)"), EUR,
       note='the number that decides whether the raise is big enough')
# ---- where the raise actually goes ---------------------------------------
FSQ = "'Financial Statements'"
DATES = f'{FSQ}!$E$3:${MC[-1]}$3'
CASHR = f'{FSQ}!$E$40:${MC[-1]}$40'
POSTR = f'{FSQ}!$N$40:${MC[-1]}$40'          # from October 2026 onwards
TROUGH = 'D38'

def su_line(row, lbl, formula, fmt=EUR, tot=False, note=None, indent=0):
    c2 = DB.cell(row, 2, ('    ' * indent) + lbl)
    c2.font = f(bold=tot)
    c = DB['D' + str(row)]
    c.value = formula; c.number_format = fmt
    c.font = f(bold=tot); c.alignment = R
    if tot:
        c.fill = fill(FILL_SUB); c2.fill = fill(FILL_SUB)
    if note:
        nn = DB[f'J{row}']; nn.value = note; nn.number_format = TEXT
        nn.font = f(italic=True, color=GREY, size=9, name=NOTE_FONT); nn.alignment = L

d_bar(36, 'WHERE THE MONEY GOES')
su_line(37, 'Lowest the cash balance ever gets, after the raise',
        f'=MIN({POSTR})', note='the moment the plan is closest to running out')
su_line(38, 'The month it happens',
        f'=INDEX({DATES},MATCH(D37,{POSTR},0)+9)', DATE_FMT)

TO_TROUGH = lambda row: (f'SUMIF({DATES},"<="&{TROUGH},'
                         f'{FSQ}!$E${row}:${MC[-1]}${row})')
su_line(40, 'Cash at the start of January 2026', f'={LV("open_cash")}')
su_line(41, 'Equity raised up to that month', f'={TO_TROUGH(35)}')
su_line(42, 'Convertible loan drawn', f'={TO_TROUGH(36)}')
su_line(43, 'Total money available', '=SUM(D40:D42)', tot=True)
su_line(44, 'Production lines and tooling', f'={TO_TROUGH(61)}',
        note='two assembly lines plus tooling, each paid twelve months before it produces')
su_line(45, 'Absorbed by operations and working capital', f'=-{TO_TROUGH(32)}',
        note='trading losses, plus the receivables and stock the ramp ties up')
su_line(46, 'Cash still in the bank at the low point', f'=D37')
su_line(47, 'Total', '=SUM(D44:D46)', tot=True,
        note='ties to total money available')

d_bar(49, 'IS THE RAISE THE RIGHT SIZE')
su_line(50, 'Cash the moment the raise lands', f'={FSQ}!{MC[9]}40')
su_line(51, 'Most of the raise ever drawn down', '=D50-D37')
su_line(52, 'Share of the raise the plan actually uses',
        f'=IFERROR(D51/SUM({FSQ}!{MC[9]}35:{MC[-1]}35),0)', PCT1, tot=True,
        note='well under 100% means the raise is bigger than this plan needs')
su_line(53, 'Months of operating cost left at the low point',
        f'=IFERROR(D37/(-{FSQ}!BP14/12),0)', NUM1,
        note='a plan of this size wants three months or more here')

DB['B34'] = 'Balance sheet check, worst month across the whole model'
DB['B34'].font = f(bold=True)
DB['D34'] = "=MAX(ABS(MIN('Financial Statements'!$E$58:$BL$58)),ABS(MAX('Financial Statements'!$E$58:$BL$58)))"
DB['D34'].number_format = NUM2; DB['D34'].font = f(bold=True); DB['D34'].alignment = R
DB['J34'] = 'must read nil'
DB['J34'].font = f(italic=True, color=GREY, size=9, name=NOTE_FONT)

# ===========================================================================
# HOW TO READ ME
# ===========================================================================
HR = wb.create_sheet('How to read me')
HR.sheet_view.showGridLines = False
for cl, w in (('A', 3), ('B', 16), ('C', 118)):
    HR.column_dimensions[cl].width = w
HR['B1'] = 'Tarnoc B.V.  How to read this model'
HR['B1'].font = f(bold=True, color=WHITE); HR['B1'].fill = fill(FILL_BLACK)
HR['C1'].fill = fill(FILL_BLACK)

def h_bar(row, text):
    HR.cell(row, 2, text).font = f(bold=True, color=WHITE)
    for c in (2, 3):
        HR.cell(row, c).fill = fill(FILL_BLACK)
    HR.cell(row, 2).value = text

def h_line(row, left, right, lfont=None, lfill=None):
    c = HR.cell(row, 2, left)
    c.font = lfont or f(bold=True)
    if lfill:
        c.fill = fill(lfill)
    r = HR.cell(row, 3, right); r.font = f(); r.alignment = L

h_bar(3, 'COLOUR CODING')
h_line(4, 'Sample', 'Hardcoded input. These are the only cells to change. They live on Assumptions, plus the back-office headcount row on Personnel.',
       f(bold=True), FILL_INPUT)
h_line(5, 'Sample', 'Calculated cell. Do not type over these.', f(bold=True))
h_line(6, 'Sample', 'Subtotal or total row.', f(bold=True), FILL_SUB)
h_line(7, 'Sample', 'Section header.', f(bold=True, color=WHITE), FILL_BLACK)
h_line(8, 'Sample', 'Margin or ratio row.', f(bold=True, color=GREY))
h_line(9, 'Sample', 'Check row. Must read nil.', f(bold=True, color=CHECK_GREY))

h_bar(12, 'HOW THE MODEL FITS TOGETHER')
for r, txt in enumerate([
    'Assumptions   ->   every other tab. The only other typed numbers are the back-office headcount on Personnel and the committed 2026 plan on OPEX.',
    'Revenue Forecast   ->   COGS, OPEX, Personnel, Financial Statements.',
    'Personnel   ->   OPEX   ->   Financial Statements.',
    'Everything   ->   Dashboard.',
], start=13):
    HR.cell(r, 3, txt).font = f()

h_bar(18, 'THE TWO SWITCHES')
HR.cell(19, 3, f'Assumptions row {CASE_ROW}, cell E{CASE_ROW}. Case: 1 is the base plan on a EUR3m raise, '
               '2 is the aggressive plan on EUR10m. Every Live column on Assumptions follows it.').font = f()
HR.cell(20, 3, f'Assumptions row {TIER_ROW}, cell E{TIER_ROW}. BOM tier basis: 1 prices the bill of materials '
               'off this year volume alone, 2 off this year plus next. Basis 2 means committing next year '
               'volume to the supplier, so basis 1 is the default.').font = f()

h_bar(22, 'HOW UNITS SOLD IS DECIDED')
for r, txt in enumerate([
    'Units sold is not typed in. Three numbers are worked out for every month and the smallest one wins:',
    '    1.  Demand.  Marketing spend divided by cost per lead, through the two conversion rates, plus the orders installer partners bring in on their own jobs.',
    '    2.  What we can sell.  Our own reps times quota, plus installer partners times units each, capped by the direct and channel mix. No ramp-up: reps and partners sell at full rate from the month they join.',
    '    3.  What we can build.  The assembly partner, plus any in-house line that is producing.',
    'There is no market size or market share anywhere in the model.',
], start=23):
    HR.cell(r, 3, txt).font = f()

h_bar(30, 'THINGS WORTH KNOWING')
for r, txt in enumerate([
    'Installation is a pass-through. It is charged to the customer at exactly what the installation partner is paid, '
    'so it adds revenue and an identical cost and no margin.',
    'Field service engineers are salaried staff on the Personnel tab. The service contract cost per unit in COGS '
    'covers parts, consumables and travel, not their salaries.',
    'Unit cost falls in steps, not smoothly. The turbineketel bill of materials is EUR9,984 below 5,000 units a year, '
    'EUR7,069 from 5,000 and EUR4,998 from 10,000. Almost all of the profit in the later years comes from crossing '
    'the second step, so that assumption carries more weight than any other in the model.',
    'A production line is paid for twelve months before it can build anything. That lag is why the timing of the raise '
    'matters as much as the size.',
    'Every month is a real column. There are no annual-only columns, so nothing can be hardcoded in a year that the '
    'monthly build does not see.',
], start=31):
    c = HR.cell(r, 3, txt); c.font = f(); c.alignment = Alignment(horizontal='left', wrap_text=False)

# ===========================================================================
# COVER
# ===========================================================================
CV = wb.create_sheet('Cover')
CV.sheet_view.showGridLines = False
for cl, w in (('A', 3), ('B', 70), ('C', 30)):
    CV.column_dimensions[cl].width = w
for r in range(1, 30):
    for c in range(1, 8):
        CV.cell(r, c).fill = fill(FILL_BLACK)
def cv(row, text, size=10, bold=False, color=WHITE):
    c = CV.cell(row, 2, text)
    c.font = Font(name=FONT, size=size, bold=bold, color=color)
    c.alignment = L
cv(6, 'TARNOC B.V.', 20, True)
cv(8, 'Financial Model', 12)
cv(9, 'Built from scratch, September 2026', 9, color=GREY)
cv(12, 'Currency', 9, color=GREY); CV.cell(12, 3, 'EUR, ex VAT unless stated').font = Font(name=FONT, size=9, color=WHITE)
cv(13, 'Period', 9, color=GREY); CV.cell(13, 3, 'January 2026 to December 2030, monthly').font = Font(name=FONT, size=9, color=WHITE)
cv(14, 'Fiscal year', 9, color=GREY); CV.cell(14, 3, 'Calendar').font = Font(name=FONT, size=9, color=WHITE)
cv(15, 'Units', 9, color=GREY); CV.cell(15, 3, 'Whole euros').font = Font(name=FONT, size=9, color=WHITE)
cv(18, 'Read the How to read me tab first.', 9, color=GREY)
cv(21, 'STRICTLY CONFIDENTIAL', 10, True, RED)

# ---- tab order, matching the house layout ---------------------------------
order = ['Cover', 'How to read me', 'Assumptions', 'Financial Statements',
         'Revenue Forecast', 'COGS', 'OPEX', 'Personnel', 'Dashboard']
wb._sheets = [wb[t] for t in order]

os.makedirs(os.path.dirname(OUT) or '.', exist_ok=True)
wb.save(OUT)
print('saved', OUT)
