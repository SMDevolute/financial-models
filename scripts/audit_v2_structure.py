"""audit2.py -- structural checks: hardcodes, cross-sheet targets, live column."""
import openpyxl, re, sys
from openpyxl.utils import get_column_letter as gl
P=sys.argv[1]; F=openpyxl.load_workbook(P)
M0,NM=5,60; AC0=M0+NM+1; YEARS=[2026,2027,2028,2029,2030]
CALC=['Revenue Forecast','COGS','Personnel','OPEX','Financial Statements']
fails=[]

print('=== 1. hardcoded numbers in the monthly grid (only the frozen 2026 block is allowed) ===')
ALLOWED={('OPEX',r) for r in (38,39,40,41,42,43)}
hard=[]
for name in CALC:
    ws=F[name]
    for r in range(4,70):
        if not ws.cell(r,2).value: continue
        for i in range(NM):
            v=ws.cell(r,M0+i).value
            if isinstance(v,(int,float)) and (name,r) not in ALLOWED:
                hard.append(f'{name}!{gl(M0+i)}{r} = {v}  ({ws.cell(r,2).value})')
if hard:
    fails.append('hardcodes')
    for h in hard[:15]: print('  FAIL ',h)
    print(f'  ... {len(hard)} total')
else:
    print('  pass  every monthly cell outside the frozen block is a formula')

print('\n=== 2. cross-sheet references point at a labelled row ===')
pat=re.compile(r"(?:'([^']+)'|\b(Revenue Forecast|COGS|Personnel|OPEX|Financial Statements|Assumptions|Dashboard))!\$?([A-Z]{1,2})\$?(\d+)")
bad=[]; mapping={}
for name in CALC+['Dashboard']:
    ws=F[name]
    for r in range(1,80):
        for c in list(range(2,M0+3))+[AC0]:
            v=ws.cell(r,c).value
            t=getattr(v,'text',v)
            if not isinstance(t,str) or '!' not in t: continue
            for q,plain,col,tr in pat.findall(t):
                tgt=q or plain
                if tgt not in F.sheetnames: continue
                trow=int(tr)
                tlbl=F[tgt].cell(trow,2).value
                if tgt=='Assumptions':
                    tlbl=F[tgt].cell(trow,2).value
                src=ws.cell(r,2).value
                if trow == 3:
                    continue
                if not tlbl:
                    bad.append(f'{name}!{gl(c)}{r} ({src}) -> {tgt} row {trow} has NO LABEL')
                else:
                    mapping.setdefault((name,r,str(src)),set()).add((tgt,trow,str(tlbl)))
if bad:
    fails.append('unlabelled targets')
    for b in bad[:12]: print('  FAIL ',b)
else:
    print('  pass  every cross-sheet reference lands on a labelled row')

print('\n  reference map, for review:')
for (sn,sr,sl),tgts in sorted(mapping.items(), key=lambda x:(CALC.index(x[0][0]) if x[0][0] in CALC else 9, x[0][1])):
    if sn=='Dashboard': continue
    for tgt,trow,tlbl in sorted(tgts):
        if tgt=='Assumptions': continue
        print(f'    {sn:<21} r{sr:<3} {sl[:34]:<34} -> {tgt:<21} r{trow:<3} {tlbl[:34]}')

print('\n=== 3. the live column on Assumptions ===')
AS=F['Assumptions']; live_bad=[]
for r in range(4,200):
    v=AS.cell(r,6).value
    if not isinstance(v,str) or not v.startswith('='): continue
    if str(AS.cell(r,2).value or '').strip() == 'Live (per case)':
        continue
    if v.startswith('=IF(Assumptions!$E$5=2,'):
        m=re.match(r'=IF\(Assumptions!\$E\$5=2,E(\d+),D(\d+)\)$',v)
        if not m or int(m.group(1))!=r or int(m.group(2))!=r:
            live_bad.append(f'Assumptions!F{r} = {v}  (should reference E{r} and D{r})')
if live_bad:
    fails.append('live column')
    for b in live_bad: print('  FAIL ',b)
else:
    print('  pass  every live cell reads the Base and Aggressive cell on its own row')

ytab_bad=[]
for r in range(4,200):
    lbl=AS.cell(r,2).value
    if isinstance(lbl,str) and lbl.strip()=='Live (per case)':
        for k in range(5):
            col=gl(4+k); v=AS[f'{col}{r}'].value
            want=f'=IF(Assumptions!$E$5=2,{col}{r-1},{col}{r-2})'
            if v!=want: ytab_bad.append(f'Assumptions!{col}{r} = {v!r} want {want!r}')
if ytab_bad:
    fails.append('year tables')
    for b in ytab_bad[:8]: print('  FAIL ',b)
else:
    print('  pass  every year table live row reads its own Base and Aggressive rows')

print('\n=== 4. orphans: labels with no numbers, numbers with no label ===')
orph=[]
for name in CALC:
    ws=F[name]
    for r in range(4,70):
        lbl=ws.cell(r,2).value
        hasv=any(ws.cell(r,M0+i).value is not None for i in range(NM))
        isbar=ws.cell(r,2).fill.fgColor.rgb=='FF000000' if ws.cell(r,2).fill.fill_type=='solid' else False
        if lbl and not hasv and not isbar and str(lbl).strip():
            orph.append(f'{name}!{r} "{lbl}" has a label but no values')
        if hasv and not lbl:
            orph.append(f'{name}!{r} has values but no label')
if orph:
    fails.append('orphans')
    for o in orph: print('  FAIL ',o)
else:
    print('  pass  no orphan rows')

print('\n=== 5. percentage rows are formatted as percentages ===')
fmt=[]
for name in CALC:
    ws=F[name]
    for r in range(4,70):
        lbl=str(ws.cell(r,2).value or '')
        u=str(ws.cell(r,3).value or '')
        nf=ws.cell(r,M0).number_format
        if u=='%' and '%' not in nf: fmt.append(f'{name}!{r} "{lbl}" unit is % but format is {nf!r}')
        if u.startswith('EUR') and '€' not in nf and '#,##0' not in nf:
            fmt.append(f'{name}!{r} "{lbl}" unit is {u} but format is {nf!r}')
if fmt:
    fails.append('formats')
    for x in fmt: print('  FAIL ',x)
else:
    print('  pass  units and number formats agree')

print(f'\n{"ALL STRUCTURAL CHECKS PASS" if not fails else "FAILURES: "+", ".join(fails)}')
