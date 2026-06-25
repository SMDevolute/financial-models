"""
house_style.py — the visual house style for financial models.

The point: models that look like a deal team built them, not like a default
spreadsheet dump. Gridlines off, deliberate typography, a real number-format
system, parenthesised negatives, blue inputs / black formulas, a spacer margin,
section bars, freeze panes and print setup — applied consistently via helpers
so every model comes out coherent.

Palette is themed to Evolute's brand (deep navy ink) and lives in one block
below — swap PALETTE to re-theme everything.

Usage (in a build script):

    import sys, os
    sys.path.insert(0, os.path.join(os.path.dirname(__file__)))  # if needed
    import house_style as hs

    wb = hs.workbook()
    ws = hs.sheet(wb, "Model")
    hs.title_block(ws, "Acme Inc.", "Three-statement operating model · $000 · FY")
    r = 6
    hs.section(ws, r, "Revenue build", span=6); r += 2
    hs.col_headers(ws, r, ["FY24", "FY25", "FY26", "FY27", "FY28"], start_col=3); r += 1
    hs.label(ws, r, "Units sold"); hs.row_inputs(ws, r, [1000, 1200, 1450], start_col=3, fmt="num")
    ...
    hs.set_widths(ws, {1: 2.5, 2: 30, 3: 12, 4: 12, 5: 12})
    hs.finish(ws, freeze="C7")
    wb.save("models/acme.xlsx")
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette (hex, no '#') — re-theme by editing here ----------------------
PALETTE = dict(
    ink="18222F",        # primary text — deep slate
    navy="2A3A55",       # section bars / strong headers (Evolute house blue)
    muted="6B7686",      # secondary text / notes
    input="2F6FB0",      # hardcoded inputs (blue)
    link="2F8F63",       # cross-sheet links (green)
    neg="C75B45",        # negatives where colour is wanted
    band="F7F8FA",       # subtle zebra band
    subtotal="ECEFF3",   # subtotal/!=total fill
    rule="D7DBE0",       # hairlines
    white="FFFFFF",
)
P = PALETTE
BASE_FONT = "Arial"      # banker-standard, cross-platform; avoids the Calibri "AI tell"
BASE_SIZE = 10

# ---- number formats — one system, used everywhere --------------------------
FMT = dict(
    num="#,##0;(#,##0)",
    num1="#,##0.0;(#,##0.0)",
    num2="#,##0.00;(#,##0.00)",
    usd='_-$* #,##0_-;_-$* (#,##0);_-$* "–"_-;_-@_-',          # accounting $
    usd1='_-$* #,##0.0_-;_-$* (#,##0.0);_-$* "–"_-;_-@_-',
    pct="0.0%;(0.0%)",
    pct0="0%;(0%)",
    mult='0.0"x"',
    year="0",
)

# ---- shared style atoms ----------------------------------------------------
def _font(color=None, bold=False, italic=False, size=BASE_SIZE, name=BASE_FONT):
    return Font(name=name, size=size, bold=bold, italic=italic, color=(color or P["ink"]))

_thin = lambda c=P["rule"]: Side(style="thin", color=c)
_med = lambda c=P["navy"]: Side(style="medium", color=c)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def workbook():
    """A workbook whose default (Normal) cell font is the house base font."""
    wb = Workbook()
    try:  # make even untouched cells use the house font, not Calibri 11
        wb._named_styles["Normal"].font = _font()
    except Exception:
        pass
    # drop the default empty sheet; callers add sheets via sheet()
    wb.remove(wb.active)
    return wb


def sheet(wb, title, margin_width=2.5):
    ws = wb.create_sheet(title=title[:31])
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = margin_width  # left margin / spacer
    # print setup — fit to one page wide, tidy margins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False
    for side in ("left", "right", "top", "bottom"):
        setattr(ws.page_margins, side, 0.5)
    return ws


def title_block(ws, title, subtitle=None, row=2, col=2):
    c = ws.cell(row=row, column=col, value=title)
    c.font = _font(color=P["ink"], bold=True, size=17)
    c.alignment = LEFT
    if subtitle:
        s = ws.cell(row=row + 1, column=col, value=subtitle)
        s.font = _font(color=P["muted"], size=9.5)
        s.alignment = LEFT
    # a thin rule under the title block, a couple rows down
    rule_row = row + 2
    for cc in range(col, col + 14):
        ws.cell(row=rule_row, column=cc).border = Border(bottom=_thin(P["navy"]))
    ws.row_dimensions[row].height = 24
    return rule_row


def section(ws, row, text, start_col=2, span=8):
    """A navy section bar with white label — the model's structure at a glance."""
    fill = PatternFill("solid", fgColor=P["navy"])
    for i in range(span):
        cell = ws.cell(row=row, column=start_col + i)
        cell.fill = fill
    head = ws.cell(row=row, column=start_col, value=text.upper())
    head.font = _font(color=P["white"], bold=True, size=9.5)
    head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row


def col_headers(ws, row, labels, start_col=3, label_col=2, label="", align="right"):
    """Period / column headers: bold, underlined with a hairline."""
    if label:
        lc = ws.cell(row=row, column=label_col, value=label)
        lc.font = _font(color=P["muted"], bold=True, size=9)
    al = RIGHT if align == "right" else (CENTER if align == "center" else LEFT)
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font = _font(color=P["ink"], bold=True, size=9.5)
        c.alignment = al
        c.border = Border(bottom=_thin(P["navy"]))
    return row


def label(ws, row, text, col=2, indent=0, bold=False, muted=False, italic=False):
    c = ws.cell(row=row, column=col, value=text)
    color = P["muted"] if muted else P["ink"]
    c.font = _font(color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return c


def _kind_color(kind):
    return {"input": P["input"], "link": P["link"], "formula": P["ink"], "note": P["muted"]}.get(kind, P["ink"])


def cell(ws, row, col, value, fmt="num", kind="formula", bold=False, top=False, double_top=False):
    """One numeric cell. kind drives colour: input=blue, formula=black, link=green."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=_kind_color(kind), bold=bold)
    c.alignment = RIGHT
    c.number_format = FMT.get(fmt, fmt)
    if top or double_top:
        c.border = Border(top=(_med(P["navy"]) if double_top else _thin(P["ink"])))
    return c


def row_inputs(ws, row, values, start_col=3, fmt="num", kind="input", bold=False, top=False):
    for i, v in enumerate(values):
        cell(ws, row, start_col + i, v, fmt=fmt, kind=kind, bold=bold, top=top)


def total_row(ws, row, start_col, end_col, fmt="num", kind="formula"):
    """Mark a row as a total: bold + top rule across the number columns."""
    for cc in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=cc)
        c.font = _font(color=_kind_color(kind), bold=True)
        c.border = Border(top=_thin(P["ink"]))
        if cc > start_col - 1:
            c.number_format = FMT.get(fmt, fmt)


def band_rows(ws, rows, start_col=2, end_col=9):
    """Subtle zebra banding on the given data rows."""
    fill = PatternFill("solid", fgColor=P["band"])
    for r in rows:
        for cc in range(start_col, end_col + 1):
            if ws.cell(row=r, column=cc).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=r, column=cc).fill = fill


def note(ws, row, text, col=2):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(color=P["muted"], italic=True, size=8.5)
    return c


def check(ws, row, col, formula, label_text="Check — balances"):
    """A pass/fail check cell. Formula should yield 'OK' or a gap."""
    label(ws, row, label_text, col=col, muted=True, italic=True)
    c = ws.cell(row=row, column=col + 1, value=formula)
    c.font = _font(color=P["link"], bold=True, size=9)
    c.alignment = LEFT
    return c


def set_widths(ws, widths):
    for col, w in widths.items():
        letter = col if isinstance(col, str) else get_column_letter(col)
        ws.column_dimensions[letter].width = w


def finish(ws, freeze=None):
    """Freeze panes (e.g. 'C7' keeps labels + headers visible)."""
    if freeze:
        ws.freeze_panes = freeze
    return ws


# A legend any model can drop in so the colour convention is self-documenting.
def legend(ws, row, col=2):
    label(ws, row, "Key:", col=col, muted=True, bold=True)
    items = [("Input", "input"), ("Formula", "formula"), ("Link", "link")]
    for i, (txt, kind) in enumerate(items):
        c = ws.cell(row=row, column=col + 1 + i, value="● " + txt)
        c.font = _font(color=_kind_color(kind), size=8.5)
    return row
